"""
Gera dados/fornecedores.xlsx v2 com os 9 fornecedores reais mapeados dos templates oficiais.

Substitui a lista antiga (5) por:
PEX:  Astra, TF, Emmeti, Ultrapexx, Barbi, Amanco
PPR:  TopFusion, TF, Amanco
PVC:  Tigre, Krona, Amanco

Demetiri foi removido (nao apareceu em nenhum template oficial).

Uso:
    python 01_gerar_fornecedores.py            (recusa sobrescrever arquivo existente)
    python 01_gerar_fornecedores.py --force    (regenera com backup automatico do atual)
"""
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import argparse
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DADOS_DIR = BASE_DIR / "dados"
DADOS_DIR.mkdir(exist_ok=True)

AZUL_TECNO = "1F3A5F"
CINZA_CLARO = "F2F2F2"
BRANCO = "FFFFFF"
AMARELO = "FFF2CC"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=BRANCO)
HEADER_FILL = PatternFill("solid", fgColor=AZUL_TECNO)
ALT_FILL = PatternFill("solid", fgColor=CINZA_CLARO)
PENDENTE_FILL = PatternFill("solid", fgColor=AMARELO)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 9 fornecedores reais extraidos dos templates oficiais
# Coluna "Fornecedor" deve bater EXATO com a coluna 'fornecedor' do equivalencias.xlsx
# (e com os cabecalhos dos templates oficiais). "Nome Comercial" e o display.
# Cobertura calculada do catalogo_pecas v2 (1883 pecas):
FORNECEDORES = [
    # id, nome_chave, nome_comercial, cobertura_pct, contato, email, copia, sistemas, observacao, status
    ("FOR-AMANCO",    "Amanco",     "Amanco",         60.0,  "",  "", "", "PEX/PPR/PVC/CPVC/ESGOTO", "Lider universal - 60% do catalogo. Fornecedor estrategico.", "ATIVO"),
    ("FOR-TIGRE",     "Tigre",      "Tigre",          37.0,  "",  "", "", "PVC/CPVC/ESGOTO/PVC_MARROM", "Dominante em PVC/CPVC", "ATIVO"),
    ("FOR-KRONA",     "Krona",      "Krona",          32.6,  "",  "", "", "PVC/CPVC/ESGOTO/PVC_MARROM", "Alternativa ao Tigre em PVC", "ATIVO"),
    ("FOR-TOPFUSION", "TopFusion",  "TopFusion",      28.6,  "",  "", "", "PPR/AR_COMPRIMIDO", "Dono do PPR e AR COMPRIMIDO. 305 pecas exclusivas (sem alternativa).", "ATIVO"),
    ("FOR-TF",        "TF",         "Tecno Fluidos",  27.7,  "Jamil", "", "", "PEX/PPR", "Especialista PEX/PPR.", "ATIVO"),
    ("FOR-ASTRA",     "Astra",      "Astra",          14.8,  "Otavio", "", "", "PASSANTE_LAJE/PEX", "Exclusivo em passante de laje (corta-fogo, modular).", "ATIVO"),
    ("FOR-EMMETI",    "Emmeti",     "Emmeti",         12.4,  "",  "", "", "PEX", "Fornecedor entrante (provavel preferencial Cyrela Joao Dias).", "ATIVO"),
    ("FOR-BARBI",     "Barbi",      "Barbi",          9.9,   "Joyce", "", "", "PEX", "Nicho PEX. Tratado em aba separada em algumas obras (codigo proprio).", "ATIVO"),
    ("FOR-ULTRAPEXX", "Ultrapexx",  "Ultrapexx",      9.3,   "",  "", "", "PEX", "PEX alternativo. So 2 pecas exclusivas - quase substituivel.", "ATIVO"),
    # Removido:
    # FOR-DEMETIRI - nao apareceu em nenhum template oficial (Hederson confirmar saida)
]

HEADERS = ["ID", "Fornecedor", "Nome Comercial", "Cobertura %", "Contato", "E-mail", "Cc", "Sistemas", "Observacao", "Status"]
LARGURAS = [14, 14, 18, 12, 16, 30, 22, 28, 50, 10]


def main():
    parser = argparse.ArgumentParser(description="Gera dados/fornecedores.xlsx com os 9 fornecedores mapeados.")
    parser.add_argument("--force", action="store_true",
                        help="Regenera mesmo se o arquivo ja existir (cria backup automatico antes)")
    args = parser.parse_args()

    saida = DADOS_DIR / "fornecedores.xlsx"
    if saida.exists():
        if not args.force:
            print(f"ERRO: {saida} ja existe e pode conter edicoes manuais (e-mails, contatos).")
            print("Nada foi sobrescrito.")
            print("Para regenerar do zero: python 01_gerar_fornecedores.py --force  (backup automatico)")
            sys.exit(1)
        backup = saida.with_name(f"fornecedores_backup_{datetime.now():%Y%m%d-%H%M%S}.xlsx")
        shutil.copy2(saida, backup)
        print(f"Backup criado: {backup.name}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Fornecedores"

    # cabecalho
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # dados
    for i, f in enumerate(FORNECEDORES, start=2):
        for col, valor in enumerate(f, start=1):
            c = ws.cell(row=i, column=col, value=valor)
            c.alignment = LEFT
            c.border = BORDER
            if i % 2 == 0:
                c.fill = ALT_FILL
            # destaca email vazio em amarelo (pendencia)
            if col == 6 and not valor:
                c.fill = PENDENTE_FILL

    # larguras
    for col, w in enumerate(LARGURAS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    # 2a aba: pendencias de email
    ws2 = wb.create_sheet("Pendencias")
    pendencias = [
        ("E-mails dos fornecedores (TODOS pendentes - bloqueia envio automatico)",),
        ("",),
        ("Confirmar com Hederson:",),
        ("  1. Demetiri saiu da operacao? (nao aparece em templates oficiais)",),
        ("  2. Email/contato comercial dos 9 fornecedores ativos",),
        ("  3. Politica de IPI / frete / MOQ por fornecedor",),
        ("  4. Fornecedor preferencial por sistema (ex: Emmeti para Cyrela)",),
        ("  5. Vinculo Tecno Fluidos x Astra (codigos parecidos)",),
        ("  6. Barbi: formato proprio de planilha (substituicao manual no video 03)",),
    ]
    for i, p in enumerate(pendencias, start=1):
        ws2.cell(row=i, column=1, value=p[0])
    ws2.column_dimensions["A"].width = 80

    wb.save(saida)
    print(f"OK  {saida.name}  ({len(FORNECEDORES)} fornecedores)")
    print(f"\nLembrete: 9 e-mails pendentes (todos destacados em amarelo).")


if __name__ == "__main__":
    main()
