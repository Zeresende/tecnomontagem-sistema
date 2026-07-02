"""
Gera 1 planilha .xlsx por fornecedor + 1 planilha-mestre genericamente para cotacao.

Uso:
    python 03_gerar_planilhas.py NOME_OBRA

Inputs:
    dados/catalogo_pecas.xlsx     (1883 pecas canonicas)
    dados/fornecedores.xlsx       (9 fornecedores)
    dados/equivalencias.xlsx      (4377 pares peca-fornecedor)
    obras/NOME_OBRA/levantamento.xlsx  (Hederson preencheu qtd_total)

Saida:
    obras/NOME_OBRA/saida/MESTRE-generico.xlsx
    obras/NOME_OBRA/saida/<Fornecedor>.xlsx (1 por fornecedor com cobertura > 0)
    obras/NOME_OBRA/saida/_relatorio.txt
"""
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import argparse
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DADOS_DIR = BASE_DIR / "dados"
OBRAS_DIR = BASE_DIR / "obras"

AZUL = "1F3A5F"
VERDE = "1E5631"
LARANJA = "C65911"
CINZA = "F2F2F2"
BRANCO = "FFFFFF"
AMARELO = "FFF2CC"
VERMELHO_CLARO = "FCE4D6"
VERDE_CLARO = "E2EFDA"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=BRANCO)
HEADER_FILL = PatternFill("solid", fgColor=AZUL)
SISTEMA_FILL = PatternFill("solid", fgColor=VERDE)
ALT_FILL = PatternFill("solid", fgColor=CINZA)
INPUT_FILL = PatternFill("solid", fgColor=AMARELO)
ALERTA_FILL = PatternFill("solid", fgColor=VERMELHO_CLARO)
OK_FILL = PatternFill("solid", fgColor=VERDE_CLARO)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ============================================================
# LOADERS
# ============================================================

def carregar_catalogo():
    """Retorna dict {id: peca}."""
    wb = load_workbook(DADOS_DIR / "catalogo_pecas.xlsx")
    ws = wb.active
    headers = [c.value for c in ws[1]]
    catalogo = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        peca = dict(zip(headers, row))
        catalogo[peca["id"]] = peca
    return catalogo


def carregar_fornecedores():
    """Retorna lista de dicts."""
    wb = load_workbook(DADOS_DIR / "fornecedores.xlsx")
    ws = wb["Fornecedores"]
    headers = [c.value for c in ws[1]]
    fornecedores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        f = dict(zip(headers, row))
        if f.get("Status") == "ATIVO":
            fornecedores.append(f)
    return fornecedores


def carregar_equivalencias():
    """Retorna dict {peca_id: {fornecedor_nome: codigo}}."""
    wb = load_workbook(DADOS_DIR / "equivalencias.xlsx")
    ws = wb.active
    headers = [c.value for c in ws[1]]
    eq = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        d = dict(zip(headers, row))
        eq[d["peca_id"]][d["fornecedor"]] = d["codigo"]
    return dict(eq)


def carregar_levantamento(caminho):
    """Retorna (metadados, {peca_id: qtd_total} para qtd > 0, avisos de validacao)."""
    wb = load_workbook(caminho)
    ws = wb.active
    # encontra cabecalho (linha que tem 'PECA_ID' na coluna 1)
    linha_header = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and "PECA_ID" in str(v).upper():
            linha_header = r
            break
    if linha_header is None:
        raise ValueError(f"Cabecalho 'peca_id' nao encontrado em {caminho}")

    headers = [ws.cell(linha_header, c).value for c in range(1, ws.max_column + 1)]
    col_id = headers.index("PECA_ID") + 1 if "PECA_ID" in headers else 1
    col_qtd = headers.index("QTD_TOTAL") + 1 if "QTD_TOTAL" in headers else 5

    levantamento = {}
    avisos = []
    for r in range(linha_header + 1, ws.max_row + 1):
        peca_id = ws.cell(r, col_id).value
        if peca_id is None:
            continue
        if not isinstance(peca_id, (int, float)):
            # sub-header de sistema (">>> PEX <<<") e esperado; o resto e linha suspeita
            if not str(peca_id).strip().startswith(">>>"):
                avisos.append(f"linha {r}: PECA_ID '{peca_id}' nao e numero - linha ignorada")
            continue
        qtd = ws.cell(r, col_qtd).value
        if qtd is None or qtd == 0 or (isinstance(qtd, str) and qtd.strip() == ""):
            continue
        if not isinstance(qtd, (int, float)):
            avisos.append(f"linha {r}: QTD_TOTAL '{qtd}' nao e numero - peca {int(peca_id)} ignorada (corrigir a celula)")
            continue
        if qtd < 0:
            avisos.append(f"linha {r}: QTD_TOTAL negativa ({qtd}) - peca {int(peca_id)} ignorada")
            continue
        levantamento[int(peca_id)] = float(qtd)

    # cabecalho da obra (linhas 1-4)
    metadados = {
        "obra": ws.cell(1, 2).value or "OBRA",
        "construtora": ws.cell(2, 2).value or "",
        "data": ws.cell(3, 2).value or "",
        "responsavel": ws.cell(4, 2).value or "",
    }
    return metadados, levantamento, avisos


# ============================================================
# GERADORES DE SAIDA
# ============================================================

def escrever_cabecalho_obra(ws, metadados, fornecedor=None):
    ws.cell(row=1, column=1, value="OBRA").font = Font(bold=True, size=13)
    ws.cell(row=1, column=2, value=metadados["obra"]).font = Font(size=13)
    ws.cell(row=2, column=1, value="Construtora:").font = Font(bold=True)
    ws.cell(row=2, column=2, value=metadados["construtora"])
    ws.cell(row=3, column=1, value="Data:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=metadados["data"])
    if fornecedor:
        nome_display = fornecedor.get("Nome Comercial") or fornecedor["Fornecedor"]
        ws.cell(row=4, column=1, value="Fornecedor:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=nome_display).font = Font(bold=True, size=12, color=AZUL)
        ws.cell(row=5, column=1, value="Contato:").font = Font(bold=True)
        ws.cell(row=5, column=2, value=fornecedor.get("Contato", ""))


def gerar_planilha_fornecedor(metadados, fornecedor, itens_para_cotar, caminho):
    """
    Gera planilha para um fornecedor especifico.
    itens_para_cotar = lista de dicts {peca, qtd, codigo}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotacao"

    escrever_cabecalho_obra(ws, metadados, fornecedor)

    LINHA_HEADER = 7
    headers = ["CODIGO", "DESCRICAO", "UNIDADE", "QUANTIDADE", "PRECO UNIT.", "IPI %", "SUBTOTAL"]
    larguras = [18, 55, 10, 14, 16, 10, 16]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=LINHA_HEADER, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    for col, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # agrupar por sistema
    sistema_atual = None
    linha = LINHA_HEADER + 1
    for item in sorted(itens_para_cotar, key=lambda x: (x["peca"]["sistema"], x["peca"]["descricao"])):
        peca = item["peca"]
        if peca["sistema"] != sistema_atual:
            c = ws.cell(row=linha, column=1, value=f">>> {peca['sistema']} <<<")
            c.font = Font(bold=True, color=BRANCO)
            c.fill = SISTEMA_FILL
            c.alignment = CENTER
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(headers))
            for col in range(1, len(headers) + 1):
                ws.cell(row=linha, column=col).fill = SISTEMA_FILL
                ws.cell(row=linha, column=col).border = BORDER
            sistema_atual = peca["sistema"]
            linha += 1

        ws.cell(row=linha, column=1, value=item["codigo"]).border = BORDER
        ws.cell(row=linha, column=2, value=peca["descricao"]).border = BORDER
        ws.cell(row=linha, column=3, value=peca["unidade"]).border = BORDER
        ws.cell(row=linha, column=4, value=item["qtd"]).border = BORDER
        for col_input in (5, 6):  # preco unit. e IPI: celulas vazias para o fornecedor preencher
            c_input = ws.cell(row=linha, column=col_input)
            c_input.border = BORDER
            c_input.fill = INPUT_FILL
        # subtotal fica em branco ate o preco ser informado; IPI em branco vale 0%
        formula = f'=IF(E{linha}="","",D{linha}*E{linha}*(1+IF(F{linha}="",0,F{linha})/100))'
        ws.cell(row=linha, column=7, value=formula).border = BORDER

        if linha % 2 == 0:
            for col in [1, 2, 3, 4, 7]:
                ws.cell(row=linha, column=col).fill = ALT_FILL
        linha += 1

    # TOTAL
    linha_total = linha + 1
    c = ws.cell(row=linha_total, column=6, value="TOTAL GERAL")
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="right")
    c_val = ws.cell(row=linha_total, column=7, value=f"=SUM(G{LINHA_HEADER+1}:G{linha-1})")
    c_val.font = Font(bold=True, size=12)
    c_val.fill = OK_FILL

    # MENSAGEM PARA FORNECEDOR
    msg_linha = linha_total + 3
    ws.cell(row=msg_linha, column=1, value="INSTRUCOES PARA O FORNECEDOR:").font = Font(bold=True)
    ws.cell(row=msg_linha + 1, column=1, value="1. Preencher as colunas amarelas (Preco Unit. + IPI %)")
    ws.cell(row=msg_linha + 2, column=1, value="2. Subtotal e Total Geral sao calculados automaticamente (IPI em branco = 0%)")
    ws.cell(row=msg_linha + 3, column=1, value="3. Devolver via email respondendo a mensagem original")
    ws.cell(row=msg_linha + 4, column=1, value="4. Prazo de cotacao: 48h (especificar se diferente)")

    ws.freeze_panes = f"A{LINHA_HEADER + 1}"
    wb.save(caminho)


def gerar_planilha_mestre(metadados, levantamento, catalogo, equivalencias, fornecedores, caminho):
    """Planilha mestre generica: lista todos os itens com qtd > 0 e codigo de TODOS fornecedores."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mestre"

    escrever_cabecalho_obra(ws, metadados)

    LINHA_HEADER = 6
    nomes_forn = [f["Fornecedor"] for f in fornecedores]
    headers = ["peca_id", "sistema", "descricao", "unidade", "qtd_total"] + nomes_forn + ["n_codigos"]
    larguras = [8, 14, 50, 10, 12] + [14] * len(nomes_forn) + [10]

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=LINHA_HEADER, column=col, value=h.upper())
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    for col, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    sistema_atual = None
    linha = LINHA_HEADER + 1
    for peca_id, qtd in sorted(levantamento.items(),
                                key=lambda x: (catalogo[x[0]]["sistema"], catalogo[x[0]]["descricao"])):
        peca = catalogo.get(peca_id)
        if not peca:
            continue
        if peca["sistema"] != sistema_atual:
            c = ws.cell(row=linha, column=1, value=f">>> {peca['sistema']} <<<")
            c.font = Font(bold=True, color=BRANCO)
            c.fill = SISTEMA_FILL
            c.alignment = CENTER
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(headers))
            for col in range(1, len(headers) + 1):
                ws.cell(row=linha, column=col).fill = SISTEMA_FILL
                ws.cell(row=linha, column=col).border = BORDER
            sistema_atual = peca["sistema"]
            linha += 1

        ws.cell(row=linha, column=1, value=peca["id"]).border = BORDER
        ws.cell(row=linha, column=2, value=peca["sistema"]).border = BORDER
        ws.cell(row=linha, column=3, value=peca["descricao"]).border = BORDER
        ws.cell(row=linha, column=4, value=peca["unidade"]).border = BORDER
        ws.cell(row=linha, column=5, value=qtd).border = BORDER

        n_codigos = 0
        eq_da_peca = equivalencias.get(peca_id, {})
        for i, fnome in enumerate(nomes_forn):
            cod = eq_da_peca.get(fnome, "")
            c = ws.cell(row=linha, column=6 + i, value=cod)
            c.border = BORDER
            if cod:
                n_codigos += 1
            else:
                c.fill = ALERTA_FILL
        ws.cell(row=linha, column=6 + len(nomes_forn), value=n_codigos).border = BORDER

        if linha % 2 == 0:
            for col in [1, 2, 3, 4, 5, 6 + len(nomes_forn)]:
                cell = ws.cell(row=linha, column=col)
                if cell.fill.fgColor.rgb in ("00000000", None):
                    cell.fill = ALT_FILL
        linha += 1

    ws.freeze_panes = f"F{LINHA_HEADER + 1}"
    wb.save(caminho)


# ============================================================
# ORQUESTRADOR
# ============================================================

def processar_obra(nome_obra):
    obra_dir = OBRAS_DIR / nome_obra
    levantamento_path = obra_dir / "levantamento.xlsx"
    if not levantamento_path.exists():
        print(f"ERRO: levantamento nao encontrado em {levantamento_path}")
        print(f"Rode primeiro: python 02_modelo_levantamento.py {nome_obra}")
        return

    catalogo = carregar_catalogo()
    fornecedores = carregar_fornecedores()
    equivalencias = carregar_equivalencias()
    metadados, levantamento, avisos = carregar_levantamento(levantamento_path)

    # pecas com qtd preenchida mas inexistentes no catalogo (linha inserida a mao?)
    for pid in sorted(set(levantamento) - set(catalogo)):
        avisos.append(f"peca_id {pid} (qtd {levantamento.pop(pid)}) nao existe no catalogo - ignorada")

    print(f"\n[OBRA] {metadados['obra']}")
    print(f"[CATALOGO] {len(catalogo)} pecas")
    print(f"[FORNECEDORES] {len(fornecedores)} ativos")
    print(f"[LEVANTAMENTO] {len(levantamento)} pecas com qtd > 0")
    if avisos:
        print(f"\n[ATENCAO] {len(avisos)} avisos de validacao (detalhe no _relatorio.txt):")
        for a in avisos:
            print(f"  ! {a}")

    saida_dir = obra_dir / "saida"
    saida_dir.mkdir(exist_ok=True)

    # ---- 1) Planilha mestre ----
    mestre_path = saida_dir / "MESTRE-generico.xlsx"
    gerar_planilha_mestre(metadados, levantamento, catalogo, equivalencias, fornecedores, mestre_path)
    print(f"\nOK  {mestre_path.name}")

    # ---- 2) Para cada fornecedor, pegar itens que ele cobre ----
    relatorio = []
    relatorio.append(f"RELATORIO DE COTACAO - {metadados['obra']}\n")
    relatorio.append(f"Total de pecas no levantamento: {len(levantamento)}\n\n")
    if avisos:
        relatorio.append(f"--- {len(avisos)} AVISOS DE VALIDACAO (revisar antes de enviar) ---\n")
        for a in avisos:
            relatorio.append(f"  ! {a}\n")
        relatorio.append("\n")

    itens_orfaos = []  # pecas que nenhum fornecedor cobre

    # mapa peca_id -> lista de fornecedores que cobrem
    cobertura_por_peca = {}
    for peca_id in levantamento:
        eq_da_peca = equivalencias.get(peca_id, {})
        forn_cobrem = [fnome for fnome in eq_da_peca if eq_da_peca[fnome]]
        cobertura_por_peca[peca_id] = forn_cobrem
        if not forn_cobrem:
            peca = catalogo.get(peca_id, {})
            itens_orfaos.append(f"{peca_id} - {peca.get('descricao', '?')} (sistema: {peca.get('sistema', '?')})")

    for f in fornecedores:
        fnome = f["Fornecedor"]
        itens_fornecedor = []
        for peca_id, qtd in levantamento.items():
            codigo = equivalencias.get(peca_id, {}).get(fnome)
            if codigo:
                peca = catalogo[peca_id]
                itens_fornecedor.append({
                    "peca": peca,
                    "qtd": qtd,
                    "codigo": codigo,
                })

        if not itens_fornecedor:
            relatorio.append(f"[SKIP] {fnome}: 0 itens (nao cobre nenhuma peca da obra)\n")
            continue

        caminho_forn = saida_dir / f"{fnome}.xlsx"
        gerar_planilha_fornecedor(metadados, f, itens_fornecedor, caminho_forn)
        relatorio.append(f"OK  {fnome}: {len(itens_fornecedor)} itens -> {caminho_forn.name}\n")
        print(f"OK  {caminho_forn.name} ({len(itens_fornecedor)} itens)")

    # ---- 3) Relatorio ----
    relatorio.append(f"\n--- COBERTURA ---\n")
    distribuicao = defaultdict(int)
    for peca_id, forn_lista in cobertura_por_peca.items():
        distribuicao[len(forn_lista)] += 1
    for n in sorted(distribuicao.keys()):
        relatorio.append(f"  {n} fornecedor(es) competem: {distribuicao[n]} pecas\n")

    if itens_orfaos:
        relatorio.append(f"\n--- {len(itens_orfaos)} PECAS ORFAS (sem fornecedor) ---\n")
        for o in itens_orfaos:
            relatorio.append(f"  - {o}\n")

    relatorio.append(f"\n--- POR SISTEMA ---\n")
    por_sis = defaultdict(int)
    for peca_id in levantamento:
        sis = catalogo.get(peca_id, {}).get("sistema", "?")
        por_sis[sis] += 1
    for sis, qtd in sorted(por_sis.items(), key=lambda x: -x[1]):
        relatorio.append(f"  {sis:20s}: {qtd} pecas\n")

    rel_path = saida_dir / "_relatorio.txt"
    rel_path.write_text("".join(relatorio), encoding="utf-8")
    print(f"\nOK  {rel_path.name}")
    print(f"\n>>> Saida em: {saida_dir}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("obra", help="Nome da obra (pasta em obras/)")
    args = parser.parse_args()
    processar_obra(args.obra)


if __name__ == "__main__":
    main()
