"""Inspeciona o conteudo das planilhas geradas para validacao visual."""
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from pathlib import Path

BASE = Path(__file__).parent / "obras" / "TESTE_JOAO_DIAS" / "saida"

for arquivo in ["TF.xlsx", "MESTRE-generico.xlsx"]:
    p = BASE / arquivo
    print(f"\n{'='*70}\n{arquivo}\n{'='*70}")
    wb = openpyxl.load_workbook(p, data_only=False)
    ws = wb.active
    print(f"Dimensoes: {ws.max_row} linhas x {ws.max_column} colunas")
    for r in range(1, min(ws.max_row + 1, 20)):
        vals = []
        for c in range(1, min(ws.max_column + 1, 10)):
            v = ws.cell(r, c).value
            if v is not None:
                vals.append(f"{ws.cell(r,c).coordinate}={str(v)[:35]}")
        if vals:
            print(f"  L{r:2d}: " + " | ".join(vals))
