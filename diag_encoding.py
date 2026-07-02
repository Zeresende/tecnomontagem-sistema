"""
Diagnostico: verificar se o problema de encoding esta nos BYTES do xlsx
ou apenas na exibicao do console.

openpyxl le e grava em UTF-8 interno. Se a string tem caractere bom, e ok.
Se tem '�' (REPLACEMENT CHARACTER), foi corrompido na leitura.
"""
import openpyxl
from pathlib import Path

CAT = Path(r"C:\Users\resen\Clientes\Tecnomontagem\automatização kits\sistema\dados\catalogo_pecas.xlsx")
TEMPLATES = Path(r"C:\Users\resen\Clientes\Tecnomontagem\automatização kits\Levantamento de Informações _ Hederson\16.05.2026\extraido\Projetos")

print("=" * 70)
print("CATALOGO v2 - amostra de descricoes (com codepoints)")
print("=" * 70)
wb = openpyxl.load_workbook(CAT, data_only=True)
ws = wb.active
amostras_problema = []
for r in range(2, min(20, ws.max_row + 1)):
    desc = ws.cell(r, 3).value
    if desc:
        tem_problema = "�" in desc
        # imprime so codepoints nao-ASCII
        codepoints = [(c, ord(c), hex(ord(c))) for c in desc if ord(c) > 127]
        if codepoints:
            marker = "BUG!" if tem_problema else "ok  "
            print(f"  {marker} L{r}: {desc[:60]!r}")
            for c, cp, hx in codepoints[:5]:
                nome = ""
                try:
                    import unicodedata
                    nome = unicodedata.name(c, "")
                except:
                    pass
                print(f"        char={c!r} U+{cp:04X} {nome}")
            if tem_problema:
                amostras_problema.append((r, desc))

print(f"\nDescricoes com \\ufffd: {len(amostras_problema)}")
if amostras_problema:
    print("Exemplo:", amostras_problema[0])

print("\n" + "=" * 70)
print("TEMPLATE ORIGINAL - mesma celula no xlsx do Hederson")
print("=" * 70)
wb_orig = openpyxl.load_workbook(TEMPLATES / "OBRA-QUANTITATIVO PEX-RXX.xlsx", data_only=True)
ws_orig = wb_orig["RAMAL"]
# linha 4 do RAMAL = primeira peca = TUBO PEX 16
for r in range(4, 15):
    desc = ws_orig.cell(r, 6).value  # col descricao
    if desc:
        tem_problema = "�" in desc
        codepoints = [(c, ord(c)) for c in desc if ord(c) > 127]
        marker = "BUG!" if tem_problema else "ok  "
        print(f"  {marker} L{r}: {desc[:60]!r}")
        if codepoints:
            for c, cp in codepoints[:3]:
                print(f"        char={c!r} U+{cp:04X}")

print("\n" + "=" * 70)
print("CONCLUSAO")
print("=" * 70)
if amostras_problema:
    print("PROBLEMA REAL nos arquivos. Tem REPLACEMENT CHARACTER (U+FFFD).")
    print("Esses caracteres ja foram perdidos - precisam ser reconstruidos.")
else:
    print("Conteudo dos xlsx esta OK. Problema e so de exibicao no console")
    print("(stdout Windows usa cp1252 que nao suporta caracteres do UTF-8).")
