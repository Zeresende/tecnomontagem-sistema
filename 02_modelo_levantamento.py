"""
Gera modelo de levantamento por OBRA, filtrado por sistemas usados.

Uso:
    python 02_modelo_levantamento.py NOME_OBRA --sistemas PEX,PPR
    python 02_modelo_levantamento.py NOME_OBRA --sistemas PVC_MARROM
    python 02_modelo_levantamento.py NOME_OBRA  (todos os sistemas)
    python 02_modelo_levantamento.py NOME_OBRA --force  (sobrescreve com backup automatico)

Saida:
    obras/NOME_OBRA/levantamento.xlsx
"""
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import argparse
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DADOS_DIR = BASE_DIR / "dados"
OBRAS_DIR = BASE_DIR / "obras"

AZUL = "1F3A5F"
VERDE_ESCURO = "1E5631"
CINZA = "F2F2F2"
BRANCO = "FFFFFF"
AMARELO = "FFF2CC"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=BRANCO)
HEADER_FILL = PatternFill("solid", fgColor=AZUL)
SISTEMA_FILL = PatternFill("solid", fgColor=VERDE_ESCURO)
ALT_FILL = PatternFill("solid", fgColor=CINZA)
INPUT_FILL = PatternFill("solid", fgColor=AMARELO)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def carregar_catalogo():
    """Le catalogo_pecas.xlsx v2. Retorna lista de dicts."""
    wb = load_workbook(DADOS_DIR / "catalogo_pecas.xlsx")
    ws = wb.active
    headers = [c.value for c in ws[1]]
    pecas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        peca = dict(zip(headers, row))
        pecas.append(peca)
    return pecas


def gerar_modelo(nome_obra, sistemas_filtro=None, force=False):
    obra_dir = OBRAS_DIR / nome_obra
    caminho = obra_dir / "levantamento.xlsx"
    if caminho.exists():
        if not force:
            print(f"ERRO: {caminho} ja existe.")
            print("Pode haver quantidades preenchidas pelo Hederson - nada foi sobrescrito.")
            print(f"Para regenerar: python 02_modelo_levantamento.py {nome_obra} --force  (backup automatico)")
            return False
        backup = caminho.with_name(f"levantamento_backup_{datetime.now():%Y%m%d-%H%M%S}.xlsx")
        shutil.copy2(caminho, backup)
        print(f"Backup criado: {backup.name}")

    pecas = carregar_catalogo()
    if sistemas_filtro:
        pecas = [p for p in pecas if p["sistema"] in sistemas_filtro]

    obra_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Levantamento"

    # CABECALHO DA OBRA
    ws["A1"] = "OBRA"
    ws["A1"].font = Font(bold=True, size=14)
    ws["B1"] = nome_obra
    ws["B1"].font = Font(size=14)

    ws["A2"] = "Construtora:"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = ""

    ws["A3"] = "Data:"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = ""

    ws["A4"] = "Responsavel:"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = "Hederson"

    ws["A5"] = "Grupo de finais (Marcelo):"
    ws["A5"].font = Font(bold=True)
    ws["B5"] = ""
    ws["B5"].fill = INPUT_FILL
    ws["C5"] = ("declarado na abertura da obra, ex.: FINAL 1/2 x FINAL 3/4/5/6 "
                "— nao ha auto-deteccao (decisao 2026-08-28, ver PADRAO_SPHE_ARQUIVOS.yaml secao 10)")
    ws["C5"].font = Font(italic=True, size=9, color="808080")

    # CABECALHO DA TABELA
    headers = ["peca_id", "sistema", "descricao", "unidade", "qtd_total", "obs"]
    larguras = [10, 14, 55, 10, 14, 30]
    LINHA_HEADER = 7
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=LINHA_HEADER, column=col, value=h.upper())
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    for col, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # DADOS - agrupar por sistema com sub-header
    sistema_atual = None
    linha = LINHA_HEADER + 1
    for peca in sorted(pecas, key=lambda p: (p["sistema"], p["descricao"])):
        if peca["sistema"] != sistema_atual:
            # sub-header de sistema
            c = ws.cell(row=linha, column=1, value=f">>> {peca['sistema']} <<<")
            c.font = Font(bold=True, color=BRANCO)
            c.fill = SISTEMA_FILL
            c.alignment = CENTER
            # merge ate ultima coluna
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(headers))
            for col in range(1, len(headers) + 1):
                ws.cell(row=linha, column=col).fill = SISTEMA_FILL
                ws.cell(row=linha, column=col).border = BORDER
            sistema_atual = peca["sistema"]
            linha += 1

        ws.cell(row=linha, column=1, value=peca["id"]).border = BORDER
        ws.cell(row=linha, column=2, value=peca["sistema"]).border = BORDER
        ws.cell(row=linha, column=3, value=peca["descricao"]).border = BORDER
        ws.cell(row=linha, column=4, value=peca["unidade"]).border = BORDER
        c_qtd = ws.cell(row=linha, column=5, value=0)
        c_qtd.border = BORDER
        c_qtd.fill = INPUT_FILL  # destaca celula de input
        ws.cell(row=linha, column=6, value="").border = BORDER

        if linha % 2 == 0:
            for col in [1, 2, 3, 4, 6]:
                ws.cell(row=linha, column=col).fill = ALT_FILL
        linha += 1

    ws.freeze_panes = f"A{LINHA_HEADER + 1}"

    wb.save(caminho)
    print(f"OK  {caminho}")
    print(f"    {len(pecas)} pecas listadas (filtro: {sistemas_filtro or 'todos sistemas'})")
    print(f"    Hederson preenche coluna QTD_TOTAL (em amarelo)")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("obra", help="Nome da obra")
    parser.add_argument("--sistemas", help="Lista separada por virgula (ex: PEX,PPR). Padrao: todos")
    parser.add_argument("--force", action="store_true",
                        help="Sobrescreve levantamento existente (cria backup automatico antes)")
    args = parser.parse_args()

    sistemas = None
    if args.sistemas:
        sistemas = [s.strip().upper() for s in args.sistemas.split(",")]

    if gerar_modelo(args.obra, sistemas, force=args.force) is False:
        sys.exit(1)


if __name__ == "__main__":
    main()
