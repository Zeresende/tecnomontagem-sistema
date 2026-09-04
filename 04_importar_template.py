"""
Importa o quantitativo dos templates oficiais (OBRA-QUANTITATIVO *.xlsx,
inclusive a aba KITS das Cyrelas) direto para o levantamento da obra.

Elimina a re-digitacao: le as abas de quantitativo do arquivo que o Hederson
ja preenche, casa cada linha com o catalogo e gera o levantamento.xlsx
pronto para o 03_gerar_planilhas.py. Abas ocultas tambem sao varridas
(ha obras com dados preenchidos em aba oculta).

Regras de match, em ordem:
    1. codigo de fornecedor -> equivalencias (deterministico)
    2. descricao exata normalizada (mesmo sistema)
    3. fuzzy: >= 0.85 aceita com registro; abaixo vai para revisao manual

Uso:
    python 04_importar_template.py NOME_OBRA "caminho\\QUANTITATIVO.xlsx"
    python 04_importar_template.py NOME_OBRA template.xlsx --construtora Cyrela
    python 04_importar_template.py NOME_OBRA template.xlsx --force

Saida em obras/NOME_OBRA/:
    levantamento.xlsx           formato v2 com QTD_TOTAL preenchida
    importacao_relatorio.txt    como cada item foi casado
    importacao_revisao.xlsx     pendencias com top-3 candidatos do catalogo
"""
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import argparse
import importlib
import re
from collections import defaultdict
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

BASE_DIR = Path(__file__).parent
sys.path.insert(0, str(BASE_DIR))
m02 = importlib.import_module("02_modelo_levantamento")
m03 = importlib.import_module("03_gerar_planilhas")

FUZZY_ACEITA = 0.85

# Guarda do fuzzy (04/09/2026): SequenceMatcher pontua alto entre descricoes que so
# diferem na bitola/rosca — caso real: 'DISTRIBUIDOR C/ REG. ABERTO 1" PEX 25/25/25MM'
# (Living, 164 un) casou a 0,85+ com o id 850 (1.1/4" 20/20/20MM), peca errada aceita
# em silencio. Fuzzy so vale se os numeros (16/20/25, 1/2", 3/4", 1.1/4"...) forem os
# mesmos nos dois lados; senao a linha vai pra revisao manual com os candidatos.
RX_NUMEROS = re.compile(r"[0-9]+(?:[.,][0-9]+)?(?:/[0-9]+)?")


def tokens_numericos(s):
    return set(RX_NUMEROS.findall(norm_txt(s)))

# nome da aba -> sistema do catalogo (abas fora do mapa: match sem filtro de sistema)
ABA_SISTEMA = {
    "KITS": "PEX", "RAMAL": "PEX", "PEX": "PEX", "PEX-AMANCO": "PEX",
    "KIT-PEX": "PEX", "BARBI": "PEX",
    "PASSANTE DE LAJE": "PASSANTE_LAJE",
    "PPR": "PPR",
    "AR COMPRIMIDO": "AR_COMPRIMIDO",
    "ESGOTO": "ESGOTO",
    "PVC ESGOTO": "PVC_ESGOTO",
    "PVC MARROM": "PVC_MARROM",
    "CPVC": "CPVC",
}

# nome no cabecalho do template -> chave canonica de fornecedor (equivalencias.xlsx)
ALIAS_FORNECEDOR = {
    "ASTRA": "Astra", "TF": "TF", "TECNOFLUIDOS": "TF", "TECNO FLUIDOS": "TF",
    "EMMETI": "Emmeti", "ULTRAPEXX": "Ultrapexx", "BARBI": "Barbi",
    "TOPFUSION": "TopFusion", "AMANCO": "Amanco", "TIGRE": "Tigre", "KRONA": "Krona",
}


def norm_txt(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().upper()


def norm_codigo(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip().upper()
    return "" if s in ("", "-") else s


def mapear_aba(ws):
    """Se a aba tem estrutura de quantitativo (linha 1 com Descricao + Qtde),
    retorna o mapa de colunas; senao retorna None."""
    headers = [norm_txt(ws.cell(1, c).value) for c in range(1, min(ws.max_column, 30) + 1)]
    col_desc = col_unid = col_qtd = None
    col_forn = {}
    for i, h in enumerate(headers, start=1):
        if not h:
            continue
        if col_desc is None and ("DESCRI" in h or "DESRI" in h):
            col_desc = i
        elif col_qtd is None and "QTDE" in h:
            col_qtd = i
        elif col_unid is None and "UNID" in h:
            col_unid = i
        elif ("COD" in h or "CÓD" in h) and col_desc is None:  # colunas de codigo vem antes da descricao
            nome = re.sub(r"C[OÓ]DIGOS?", "", h).strip()
            forn = ALIAS_FORNECEDOR.get(nome)
            if forn is None and nome == "":
                # coluna "Codigo" sem nome de fornecedor: inferir pelo nome da aba
                for alias, canon in ALIAS_FORNECEDOR.items():
                    if alias in norm_txt(ws.title):
                        forn = canon
                        break
            if forn:
                col_forn[i] = forn
    if col_desc is None or col_qtd is None:
        return None
    return {"col_desc": col_desc, "col_unid": col_unid, "col_qtd": col_qtd,
            "col_forn": col_forn, "sistema": ABA_SISTEMA.get(norm_txt(ws.title))}


def construir_indices(catalogo, equivalencias):
    cod_idx = defaultdict(set)    # (fornecedor, codigo) -> {peca_id}
    for pid, por_forn in equivalencias.items():
        for forn, cod in por_forn.items():
            c = norm_codigo(cod)
            if c:
                cod_idx[(forn, c)].add(pid)
    desc_idx = defaultdict(list)  # (sistema, descricao normalizada) -> [peca_id]
    for pid, p in catalogo.items():
        desc_idx[(p["sistema"], norm_txt(p["descricao"]))].append(pid)
    return cod_idx, desc_idx


def casar_linha(codigos, desc, sistema, cod_idx, desc_idx, catalogo):
    """Retorna (peca_id ou None, metodo, top3_candidatos)."""
    # 1) por codigo de fornecedor (voto entre os codigos presentes na linha)
    votos = defaultdict(int)
    for forn, cod in codigos:
        ids = cod_idx.get((forn, cod), set())
        if len(ids) == 1:
            votos[next(iter(ids))] += 1
    if votos:
        melhor = max(votos.values())
        ids_melhor = [pid for pid, v in votos.items() if v == melhor]
        if len(ids_melhor) == 1:
            return ids_melhor[0], "codigo", []
    # 2) descricao exata normalizada dentro do sistema
    dn = norm_txt(desc)
    if sistema:
        ids = desc_idx.get((sistema, dn), [])
        if len(ids) == 1:
            return ids[0], "descricao", []
    # 3) fuzzy
    scores = []
    for pid, p in catalogo.items():
        if sistema is not None and p["sistema"] != sistema:
            continue
        scores.append((SequenceMatcher(None, dn, norm_txt(p["descricao"])).ratio(), pid))
    scores.sort(reverse=True)
    top3 = [(pid, catalogo[pid]["descricao"], round(s, 3)) for s, pid in scores[:3]]
    if scores and scores[0][0] >= FUZZY_ACEITA:
        melhor_pid = scores[0][1]
        if tokens_numericos(dn) == tokens_numericos(catalogo[melhor_pid]["descricao"]):
            return melhor_pid, "fuzzy", top3
        return None, "pendente_numeros", top3
    return None, "pendente", top3


def importar_template(caminho_template, catalogo, equivalencias):
    wb = load_workbook(caminho_template, data_only=True)
    cod_idx, desc_idx = construir_indices(catalogo, equivalencias)
    itens = {}                   # peca_id -> qtd acumulada
    origem = defaultdict(list)   # peca_id -> ["KITS L4 (codigo)", ...]
    pendentes = []
    avisos = []
    stats = []
    for ws in wb.worksheets:
        mapa = mapear_aba(ws)
        if mapa is None:
            continue
        n = {"aba": ws.title, "estado": ws.sheet_state, "linhas": 0,
             "codigo": 0, "descricao": 0, "fuzzy": 0, "pendente": 0}
        if mapa["sistema"] is None:
            avisos.append(f"aba '{ws.title}' nao mapeada para um sistema - match sem filtro (conferir)")
        for r in range(2, ws.max_row + 1):
            desc = ws.cell(r, mapa["col_desc"]).value
            if desc is None or norm_txt(desc) == "":
                continue
            qtd = ws.cell(r, mapa["col_qtd"]).value
            if qtd is None or qtd == 0:
                continue
            if not isinstance(qtd, (int, float)):
                avisos.append(f"{ws.title} L{r}: qtde '{qtd}' nao e numero - linha ignorada (corrigir no template)")
                continue
            if qtd < 0:
                avisos.append(f"{ws.title} L{r}: qtde negativa ({qtd}) - linha ignorada")
                continue
            n["linhas"] += 1
            codigos = []
            for col, forn in mapa["col_forn"].items():
                c = norm_codigo(ws.cell(r, col).value)
                if c:
                    codigos.append((forn, c))
            pid, metodo, top3 = casar_linha(codigos, desc, mapa["sistema"], cod_idx, desc_idx, catalogo)
            if pid is None:
                n["pendente"] += 1
                if metodo == "pendente_numeros":
                    avisos.append(f"{ws.title} L{r}: '{str(desc).strip()[:50]}' parecia com o id "
                                  f"{top3[0][0]} ({top3[0][2]:.2f}) mas a bitola/rosca difere - foi pra revisao")
                pendentes.append({"aba": ws.title, "linha": r, "descricao": str(desc).strip(),
                                  "unid": ws.cell(r, mapa["col_unid"]).value if mapa["col_unid"] else "",
                                  "qtd": float(qtd), "top3": top3})
                continue
            n[metodo] += 1
            if pid in itens:
                avisos.append(f"peca {pid} repetida ({origem[pid][0]} + {ws.title} L{r}) - qtds somadas")
            itens[pid] = itens.get(pid, 0.0) + float(qtd)
            origem[pid].append(f"{ws.title} L{r} ({metodo})")
        stats.append(n)
    return itens, origem, pendentes, avisos, stats


def preencher_levantamento(caminho, itens, construtora):
    wb = load_workbook(caminho)
    ws = wb.active
    ws.cell(row=2, column=2, value=construtora)
    ws.cell(row=3, column=2, value=date.today().isoformat())
    restantes = dict(itens)
    for r in range(7, ws.max_row + 1):
        pid = ws.cell(r, 1).value
        if isinstance(pid, (int, float)) and int(pid) in restantes:
            ws.cell(row=r, column=5, value=restantes.pop(int(pid)))
    wb.save(caminho)
    return restantes  # se nao vazio, erro interno (peca casada sem linha no modelo)


def gerar_revisao(pendentes, caminho):
    wb = Workbook()
    ws = wb.active
    ws.title = "Revisao"
    ws.append(["Aba", "Linha", "Descricao no template", "Unid", "Qtd",
               "Cand.1 id", "Cand.1 descricao", "Score 1",
               "Cand.2 id", "Cand.2 descricao", "Score 2",
               "Cand.3 id", "Cand.3 descricao", "Score 3", "PECA_ID escolhido"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3A5F")
    for p in pendentes:
        linha = [p["aba"], p["linha"], p["descricao"], p["unid"], p["qtd"]]
        for pid, desc, score in p["top3"]:
            linha += [pid, desc, score]
        linha += [""] * (14 - len(linha)) + [""]
        ws.append(linha)
    for col, w in zip(("A", "B", "C", "D", "E"), (16, 7, 55, 7, 10)):
        ws.column_dimensions[col].width = w
    for col in ("G", "J", "M"):
        ws.column_dimensions[col].width = 45
    ws.freeze_panes = "A2"
    wb.save(caminho)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("obra", help="Nome da obra (cria pasta em obras/)")
    parser.add_argument("template", help="Caminho do OBRA-QUANTITATIVO *.xlsx preenchido")
    parser.add_argument("--construtora", default="")
    parser.add_argument("--force", action="store_true",
                        help="Sobrescreve levantamento existente (backup automatico)")
    args = parser.parse_args()

    caminho_template = Path(args.template)
    if not caminho_template.exists():
        print(f"ERRO: template nao encontrado: {caminho_template}")
        sys.exit(1)

    catalogo = m03.carregar_catalogo()
    equivalencias = m03.carregar_equivalencias()
    print(f"[CATALOGO] {len(catalogo)} pecas")
    print(f"[TEMPLATE] {caminho_template.name}")

    itens, origem, pendentes, avisos, stats = importar_template(caminho_template, catalogo, equivalencias)
    if not itens and not pendentes:
        print("ERRO: nenhuma linha com quantidade > 0 nas abas de quantitativo do template.")
        sys.exit(1)

    sistemas = sorted({catalogo[pid]["sistema"] for pid in itens})
    if m02.gerar_modelo(args.obra, sistemas or None, force=args.force) is False:
        sys.exit(1)

    obra_dir = BASE_DIR / "obras" / args.obra
    restantes = preencher_levantamento(obra_dir / "levantamento.xlsx", itens, args.construtora)

    rel = [f"RELATORIO DE IMPORTACAO - {args.obra}\n",
           f"Template: {caminho_template}\n", f"Data: {date.today().isoformat()}\n\n",
           "--- ABAS PROCESSADAS ---\n"]
    for n in stats:
        rel.append(f"  {n['aba']:20s} ({n['estado']}): {n['linhas']} itens com qtd>0 | "
                   f"codigo: {n['codigo']} | descricao: {n['descricao']} | "
                   f"fuzzy: {n['fuzzy']} | pendentes: {n['pendente']}\n")
    rel.append(f"\nTotal importado: {len(itens)} pecas (sistemas: {', '.join(sistemas)})\n")
    fuzzy_usados = {pid: o for pid, o in origem.items() if any("fuzzy" in x for x in o)}
    if fuzzy_usados:
        rel.append(f"\n--- {len(fuzzy_usados)} MATCHES POR FUZZY (conferir antes de cotar) ---\n")
        for pid, o in fuzzy_usados.items():
            rel.append(f"  peca {pid} [{itens[pid]:g}] = {'; '.join(o)}\n")
    if avisos:
        rel.append(f"\n--- {len(avisos)} AVISOS ---\n")
        rel.extend(f"  ! {a}\n" for a in avisos)
    if pendentes:
        rel.append(f"\n--- {len(pendentes)} PENDENTES DE REVISAO ---\n")
        rel.append("  Abrir importacao_revisao.xlsx, escolher o PECA_ID correto e digitar a qtd no levantamento.xlsx\n")
    if restantes:
        rel.append(f"\nERRO INTERNO: {len(restantes)} pecas casadas sem linha no modelo: {sorted(restantes)[:10]}\n")
    (obra_dir / "importacao_relatorio.txt").write_text("".join(rel), encoding="utf-8")

    if pendentes:
        gerar_revisao(pendentes, obra_dir / "importacao_revisao.xlsx")

    print(f"\nOK  {len(itens)} pecas importadas -> levantamento.xlsx")
    if fuzzy_usados:
        print(f"[CONFERIR] {len(fuzzy_usados)} matches por fuzzy (lista no importacao_relatorio.txt)")
    if pendentes:
        print(f"[REVISAR] {len(pendentes)} itens sem match -> importacao_revisao.xlsx")
    if avisos:
        print(f"[ATENCAO] {len(avisos)} avisos (detalhe no importacao_relatorio.txt)")
    print(f"\nProximo passo: python 03_gerar_planilhas.py {args.obra}")


if __name__ == "__main__":
    main()
