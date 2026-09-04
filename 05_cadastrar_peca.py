"""
Cadastra pecas NOVAS em dados/catalogo_pecas.xlsx + dados/equivalencias.xlsx.

Motivo de existir: peca que aparece no quantitativo de uma obra mas nao esta no
catalogo (ex.: as 2 travessas do Living em 28/08) era cadastrada a mao nos 2 xlsx,
sem trilha. Aqui o id sai sequencial (max+1), a equivalencia vai junto e os dois
arquivos ganham backup automatico antes de gravar.

Uso:
    python 05_cadastrar_peca.py --sistema PEX --unidade UN \
        --descricao "DISTRIBUIDOR C/ REG. ABERTO 1\"  PEX 25/25/25MM" \
        --codigo Astra=DL/003R1 --fonte "Marcelo (Tecnomontagem) via Karina, 2026-09-04"

    --codigo pode repetir (um por fornecedor). Fornecedores aceitos: os da coluna
    'fornecedor' do equivalencias.xlsx (Astra, TF, Emmeti, Ultrapexx, Barbi,
    TopFusion, Amanco, Tigre, Krona).
    --dry-run mostra o que faria sem gravar.

Recusa: descricao ja existente no mesmo sistema (normalizada) ou codigo ja usado
pelo mesmo fornecedor - nos dois casos aponta o id existente.
"""
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import argparse
import re
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent
DADOS_DIR = BASE_DIR / "dados"
CATALOGO = DADOS_DIR / "catalogo_pecas.xlsx"
EQUIVALENCIAS = DADOS_DIR / "equivalencias.xlsx"

# nome canonico do fornecedor (equivalencias.xlsx) -> coluna no catalogo
COLUNA_FORNECEDOR = {
    "Astra": "cod_astra", "TF": "cod_tf", "Emmeti": "cod_emmeti",
    "Ultrapexx": "cod_ultrapexx", "Barbi": "cod_barbi", "TopFusion": "cod_topfusion",
    "Amanco": "cod_amanco", "Tigre": "cod_tigre", "Krona": "cod_krona",
}


def norm_txt(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().upper()


def parse_codigo(texto):
    if "=" not in texto:
        raise argparse.ArgumentTypeError(f"--codigo espera Fornecedor=CODIGO, veio '{texto}'")
    forn, cod = texto.split("=", 1)
    forn = forn.strip()
    canon = next((k for k in COLUNA_FORNECEDOR if k.lower() == forn.lower()), None)
    if canon is None:
        raise argparse.ArgumentTypeError(
            f"fornecedor '{forn}' desconhecido; use um de {', '.join(COLUNA_FORNECEDOR)}")
    return canon, cod.strip()


def backup(caminho):
    destino = caminho.with_name(f"{caminho.stem}_backup_{datetime.now():%Y%m%d-%H%M%S}{caminho.suffix}")
    shutil.copy2(caminho, destino)
    return destino


def copiar_estilo(ws, linha_modelo, linha_nova, n_cols):
    for c in range(1, n_cols + 1):
        origem = ws.cell(linha_modelo, c)
        destino = ws.cell(linha_nova, c)
        if origem.has_style:
            destino._style = copy(origem._style)


def validar(ws_cat, ws_eq, headers_cat, sistema, descricao, codigos):
    """Devolve lista de conflitos (vazia = pode cadastrar)."""
    conflitos = []
    i_sis = headers_cat.index("sistema")
    i_desc = headers_cat.index("descricao")
    alvo = norm_txt(descricao)
    for row in ws_cat.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        if row[i_sis] == sistema and norm_txt(row[i_desc]) == alvo:
            conflitos.append(f"descricao ja existe no sistema {sistema}: id {row[0]}")
    usados = {}
    for row in ws_eq.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        usados[(row[4], norm_txt(row[5]))] = row[0]
    for forn, cod in codigos:
        pid = usados.get((forn, norm_txt(cod)))
        if pid is not None:
            conflitos.append(f"codigo {cod} ja e usado por {forn} na peca id {pid}")
    return conflitos


def cadastrar(sistema, descricao, unidade, codigos, fonte, dry_run=False):
    wb_cat = load_workbook(CATALOGO)
    ws_cat = wb_cat.active
    wb_eq = load_workbook(EQUIVALENCIAS)
    ws_eq = wb_eq.active
    headers_cat = [c.value for c in ws_cat[1]]

    conflitos = validar(ws_cat, ws_eq, headers_cat, sistema, descricao, codigos)
    if conflitos:
        print("RECUSADO:")
        for c in conflitos:
            print(f"  - {c}")
        return None

    ids = [row[0] for row in ws_cat.iter_rows(min_row=2, values_only=True)
           if isinstance(row[0], (int, float))]
    novo_id = int(max(ids)) + 1
    ultima = max(r for r in range(2, ws_cat.max_row + 1) if ws_cat.cell(r, 1).value is not None)

    registro = {h: None for h in headers_cat}
    registro.update({"id": novo_id, "sistema": sistema, "descricao": descricao,
                     "unidade": unidade, "n_fornecedores": len(codigos), "fontes": fonte})
    for forn, cod in codigos:
        registro[COLUNA_FORNECEDOR[forn]] = cod

    print(f"id {novo_id} | {sistema} | {descricao} | {unidade} | "
          f"{', '.join(f'{f}={c}' for f, c in codigos) or 'sem codigo'}")
    if dry_run:
        print("  (dry-run: nada gravado)")
        return novo_id

    b1, b2 = backup(CATALOGO), backup(EQUIVALENCIAS)
    linha = ultima + 1
    copiar_estilo(ws_cat, ultima, linha, len(headers_cat))
    for col, h in enumerate(headers_cat, start=1):
        ws_cat.cell(linha, col, registro[h])

    ultima_eq = max(r for r in range(2, ws_eq.max_row + 1) if ws_eq.cell(r, 1).value is not None)
    for forn, cod in codigos:
        ultima_eq += 1
        copiar_estilo(ws_eq, ultima_eq - 1, ultima_eq, 6)
        for col, v in enumerate((novo_id, sistema, descricao, unidade, forn, cod), start=1):
            ws_eq.cell(ultima_eq, col, v)

    wb_cat.save(CATALOGO)
    wb_eq.save(EQUIVALENCIAS)
    print(f"  gravado (backups: {b1.name}, {b2.name})")
    return novo_id


def main():
    parser = argparse.ArgumentParser(description="Cadastra uma peca nova no catalogo + equivalencias")
    parser.add_argument("--sistema", required=True, help="ex.: PEX, PPR, PVC_MARROM")
    parser.add_argument("--descricao", required=True, help="exatamente como aparece no quantitativo")
    parser.add_argument("--unidade", default="UN")
    parser.add_argument("--codigo", action="append", default=[], type=parse_codigo,
                        metavar="Fornecedor=CODIGO", help="pode repetir")
    parser.add_argument("--fonte", required=True, help="quem informou e quando")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    novo = cadastrar(args.sistema.upper(), args.descricao.strip(), args.unidade.upper(),
                     args.codigo, args.fonte, dry_run=args.dry_run)
    sys.exit(0 if novo else 1)


if __name__ == "__main__":
    main()
