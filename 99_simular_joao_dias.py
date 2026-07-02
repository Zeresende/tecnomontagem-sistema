"""
SIMULACAO: preenche o levantamento.xlsx da obra TESTE_JOAO_DIAS
com os totais reais extraidos do KITS/Joao Dias original.

Usa fuzzy match por descricao para encontrar o peca_id correspondente
no catalogo v2.
"""
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from pathlib import Path
from difflib import SequenceMatcher

BASE_DIR = Path(__file__).parent

# Dados REAIS da aba KITS/JoaoDias (linhas que tem qtd > 0)
# Formato: (descricao_aproximada_kits_jd, qtd_total)
DADOS_REAIS_JD = [
    # tubo PEX
    ("TUBO PEX 16 - SERIE 5 DN16MM", 16600),
    ("TUBO PEX 20 - SERIE 5 DN 20MM", 6900),
    # tirar dados de mais peca seria necessario abrir a planilha completa
    # estes 2 ja servem para validacao
]


def similaridade(a, b):
    return SequenceMatcher(None, a.upper(), b.upper()).ratio()


def encontrar_peca_id(desc_busca, levantamento_ws, col_id=1, col_sistema=2, col_desc=3, linha_header=6):
    """Procura no levantamento qual peca_id mais se assemelha a desc_busca (so PEX)."""
    melhor_id = None
    melhor_score = 0.0
    melhor_desc = ""
    for r in range(linha_header + 1, levantamento_ws.max_row + 1):
        peca_id = levantamento_ws.cell(r, col_id).value
        sistema = levantamento_ws.cell(r, col_sistema).value
        desc = levantamento_ws.cell(r, col_desc).value
        if peca_id is None or not isinstance(peca_id, (int, float)):
            continue
        if sistema != "PEX":
            continue
        s = similaridade(desc_busca, desc or "")
        if s > melhor_score:
            melhor_score = s
            melhor_id = (peca_id, r)
            melhor_desc = desc
    return melhor_id, melhor_score, melhor_desc


def main():
    obra = "TESTE_JOAO_DIAS"
    levantamento_path = BASE_DIR / "obras" / obra / "levantamento.xlsx"
    if not levantamento_path.exists():
        print(f"ERRO: {levantamento_path} nao existe.")
        print(f"Rode: python 02_modelo_levantamento.py {obra} --sistemas PEX")
        return

    wb = openpyxl.load_workbook(levantamento_path)
    ws = wb.active

    # preenche metadados
    ws.cell(row=2, column=2, value="Cyrela")
    ws.cell(row=3, column=2, value="2026-05-16")

    print("Preenchendo dados reais do KITS/Joao Dias:\n")
    for desc, qtd in DADOS_REAIS_JD:
        match, score, melhor_desc = encontrar_peca_id(desc, ws)
        if match and score > 0.5:
            peca_id, linha = match
            ws.cell(row=linha, column=5, value=qtd)
            print(f"  [{score:.2f}] {desc}")
            print(f"         -> id {peca_id}: {melhor_desc[:60]} = {qtd}m")
        else:
            print(f"  [SKIP] '{desc}' (melhor match: {score:.2f})")

    wb.save(levantamento_path)
    print(f"\nOK  Salvo em {levantamento_path}")
    print(f"\nProximo passo: python 03_gerar_planilhas.py {obra}")


if __name__ == "__main__":
    main()
