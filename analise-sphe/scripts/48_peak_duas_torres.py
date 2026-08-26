# -*- coding: utf-8 -*-
"""PEAK — MEDICAO REFEITA SOMANDO AS DUAS TORRES (item 13.1, 11/08/2026).

O que muda em relacao a 10/08. Naquele dia achamos que o pavimento do Peak estava
desenhado duas vezes (as notas vinham em pares deslocados 50 em Y, e 94,5% dos trechos
da metade de cima tinham gemeo identico 50 abaixo) e concluimos que medir o modelspace
inteiro contava o mesmo tubo duas vezes. **O Hederson corrigiu no item 13.1: "sao duas
torres do mesmo empreendimento".** Ambas contam, nao ha divisao por 2 a fazer.

E ha uma segunda coisa errada naquela medicao, achada agora: **cortar o desenho pela
metade do Y nao separa as duas torres.** A prancha tem seis aglomerados de tubo, nao
dois — as duas plantas de torre, um par de plantas de FURO (uma por torre, e a planilha
tem aba `PASSANTE DE LAJE` para elas), um bloco estrutural a direita e um terceiro
pavimento quase sem tubo. Somar tudo mistura ramal com furo.

Por isso este script recorta por REGIAO (x, y) e diz o que cada aglomerado e, antes de
medir. Depois compara cada torre com a SUA aba de ramal, em metro por apartamento, e
quebra por sufixo de camada — que foi o que fez a Living fechar em 03/07 (EXO-TET = o
ramal aereo; o total de todas as camadas da 2,45x a receita, na Living tambem).

Uso: python 48_peak_duas_torres.py [obra]
"""
import sys, os, re, glob, math, logging
from collections import defaultdict, Counter
from pathlib import Path
import ezdxf
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
logging.getLogger("ezdxf").setLevel(logging.ERROR)

AQUI = Path(__file__).resolve().parent
sys.path.insert(0, str(AQUI))
m22 = __import__("22_ramal_por_dn")
m25 = __import__("25_avaliar_dn")
p17 = __import__("17_parser_contagens")
BASE = m25.BASE

TUBO_LAYER = re.compile(r"H(AF|AQ)-TUB", re.I)
EXCL_LAYER = re.compile(r"LEG|REF|TAB|DET|PREVIS", re.I)
DESLOC = 50.0          # deslocamento entre as duas torres, medido em 10/08
TOL_GEMEO = 0.02
CEL = (50.0, 25.0)     # celula do mapa de regioes
FINAIS = 10            # finais por pavimento tipo, pela matriz PREDIO do Peak

# o que cada aglomerado e, lido pelos textos de dentro (probe de 11/08)
RX_FURO = re.compile(r"FURO|PASSANTE|VIGA|PILAR|ALV-TETO", re.I)
RX_TIPOLOGIA = re.compile(r"^TIPO [A-Z]\d", re.I)


def andar(entidades, prof=0):
    for e in entidades:
        if e.dxftype() == "INSERT" and prof < 6:
            try:
                yield from andar(e.virtual_entities(), prof + 1)
            except Exception:
                pass
            continue
        yield e


def coletar(doc):
    """(a, b, comprimento, sistema, sufixo) de cada segmento de tubo."""
    saida = []
    textos = []
    for e in andar(doc.modelspace()):
        t = e.dxftype()
        if t in ("TEXT", "MTEXT"):
            s = (e.plain_text() if t == "MTEXT" else str(e.dxf.text)).strip().replace("\n", " ")
            if s and len(s) < 70:
                try:
                    textos.append((s, tuple(e.dxf.insert)[:2]))
                except Exception:
                    pass
            continue
        ly = str(getattr(e.dxf, "layer", ""))
        mm = TUBO_LAYER.search(ly)
        if not mm or EXCL_LAYER.search(ly):
            continue
        suf = re.sub(r"[_\-]+", "-", ly[mm.end():]).strip("-").upper() or "(base)"
        pts = m22.pontos_de(e)
        for i in range(len(pts) - 1):
            a, b = pts[i], pts[i + 1]
            c = math.dist(a, b)
            if c > 0:
                saida.append((a, b, c, mm.group(1).upper(), suf))
    return saida, textos


def regioes(dados, textos):
    """Aglomera o tubo em celulas e classifica cada aglomerado pelos textos de dentro."""
    cel = defaultdict(list)
    for i, (a, b, c, sis, suf) in enumerate(dados):
        x = int(((a[0] + b[0]) / 2) // CEL[0] * CEL[0])
        y = int(((a[1] + b[1]) / 2) // CEL[1] * CEL[1])
        cel[(x, y)].append(i)
    saida = []
    for (x, y), idx in cel.items():
        m = sum(dados[i][2] for i in idx)
        if m < 1.0:
            continue
        dentro = [s for s, p in textos
                  if x <= p[0] <= x + CEL[0] and y <= p[1] <= y + CEL[1]]
        n_furo = sum(1 for s in dentro if RX_FURO.search(s))
        n_tipo = sum(1 for s in dentro if RX_TIPOLOGIA.match(s))
        # planta de torre = tem tipologia de apartamento; planta de furo = so furo
        if n_tipo >= 3:
            classe = "PLANTA DE TORRE"
        elif n_furo and n_furo / max(len(dentro), 1) > 0.05:
            classe = "PLANTA DE FURO/ESTRUTURA"
        else:
            classe = "(indefinido)"
        saida.append({"x": x, "y": y, "m": m, "idx": idx, "classe": classe,
                      "n_tipo": n_tipo, "n_furo": n_furo, "n_txt": len(dentro)})
    saida.sort(key=lambda r: -r["m"])
    return saida


def planilha_por_torre(obra):
    x = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
         if "PRE-" not in a.upper() and "LEVANTAMENTO" not in a.upper()
         and not os.path.basename(a).startswith("~$")][0]
    wb = openpyxl.load_workbook(x, data_only=True)
    rec = {}
    for ws in wb.worksheets:
        if not ws.title.upper().startswith("RAMAL"):
            continue
        cr = p17.linha_contagem(ws)
        tot = defaultdict(float)
        for r in range(cr + 1, ws.max_row + 1):
            e, u, g = ws.cell(r, 5).value, ws.cell(r, 6).value, ws.cell(r, 7).value
            if not (isinstance(e, str) and e.upper().startswith("TUBO")):
                continue
            if str(u).strip().upper() != "RL" or not isinstance(g, (int, float)) or not g:
                continue
            mm = m25.RX_TUBO.search(e.upper())
            if not mm:
                continue
            tams = [int(t) for t in m25.TAM_ROLO.findall(e.upper())
                    if int(t) in (50, 100, 200)]
            # no Peak o sistema vem na COR do PERT: vermelho = quente, azul = fria
            cor = ("AQ" if "VERMELH" in e.upper()
                   else "AF" if "AZUL" in e.upper() else "?")
            tot[(cor, int(mm.group(2)))] += g * (tams[-1] if tams else 100) / 1.07
        rec[ws.title] = dict(tot)
    apt = {}
    for ws in wb.worksheets:
        if "PR" not in ws.title.upper() or "HID" not in ws.title.upper():
            continue
        n, pavs = 0, set()
        for r in range(1, ws.max_row + 1):
            pav = ws.cell(r, 4).value
            marcou = False
            for c in range(5, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and re.fullmatch(r"TP\d+", v.strip(), re.I):
                    n += 1
                    marcou = True
            if marcou and isinstance(pav, (int, float)):
                pavs.add(int(pav))
        apt["A" if ws.title.strip().upper().endswith("A") else "B"] = (n, len(pavs))
    wb.close()
    return rec, apt


def main():
    obra = sys.argv[1] if len(sys.argv) > 1 else "20251533"
    d = AQUI / "dxf" / obra
    caminho = [p for p in sorted(d.glob("*.dxf"))
               if re.search(r"PVTIPO|TIPO|TIP\b|-TIP-", p.name, re.I)
               and not re.search(r"DET", p.name, re.I)][0]
    print(f"== {caminho.name}\n", flush=True)

    doc = ezdxf.readfile(str(caminho))
    dados, textos = coletar(doc)
    print(f"-- {len(dados)} segmentos · {sum(s[2] for s in dados):.1f} m de tubo "
          f"no modelspace inteiro\n")

    regs = regioes(dados, textos)
    print("-- OS AGLOMERADOS DA PRANCHA (por que somar tudo estava errado)")
    print(f"   {'x':>6} {'y':>6} {'metro':>9}  {'tipologias':>10} {'furos':>6}  classe")
    for r in regs:
        print(f"   {r['x']:>6} {r['y']:>6} {r['m']:>9.1f}  {r['n_tipo']:>10} "
              f"{r['n_furo']:>6}  {r['classe']}")

    # a tipologia (TIPO D1, D3, B1...) aparece em TRES copias na prancha, mas a
    # terceira tem 6,8 m de tubo — e planta de referencia, nao pavimento medivel.
    # Sem este corte o denominador vira 30 apartamentos e o resultado muda 50%.
    todas = [r for r in regs if r["classe"] == "PLANTA DE TORRE"]
    maior = max((r["m"] for r in todas), default=0.0)
    torres = [r for r in todas if r["m"] >= 0.10 * maior]
    descartadas = [r for r in todas if r not in torres]
    torres.sort(key=lambda r: r["y"])
    for r in descartadas:
        print(f"\n   planta com tipologia em y={r['y']} descartada: so {r['m']:.1f} m de "
              f"tubo contra {maior:.1f} m da maior — referencia, nao pavimento medivel")
    fora = [r for r in regs if r["classe"] != "PLANTA DE TORRE"]
    print(f"\n   entram na conta: {len(torres)} plantas de torre · "
          f"{sum(r['m'] for r in torres):.1f} m")
    print(f"   ficam de fora  : {len(fora)} aglomerados · {sum(r['m'] for r in fora):.1f} m"
          f"  (furo/estrutura — a planilha tem aba PASSANTE DE LAJE para isso)")

    if len(torres) != 2:
        print("\n   ATENCAO: nao foram encontradas exatamente 2 plantas de torre.")

    # teste do gemeo, agora entre as duas plantas de torre e nao entre metades do Y
    if len(torres) == 2:
        base, alto = torres[0], torres[1]
        dy = alto["y"] - base["y"]
        cj = set((round(dados[i][0][0] / TOL_GEMEO), round(dados[i][0][1] / TOL_GEMEO),
                  round(dados[i][1][0] / TOL_GEMEO), round(dados[i][1][1] / TOL_GEMEO))
                 for i in base["idx"])
        casa = 0
        for i in alto["idx"]:
            a, b = dados[i][0], dados[i][1]
            ch = (round(a[0] / TOL_GEMEO), round((a[1] - dy) / TOL_GEMEO),
                  round(b[0] / TOL_GEMEO), round((b[1] - dy) / TOL_GEMEO))
            if ch in cj:
                casa += 1
        print(f"\n-- SEMELHANCA ENTRE AS DUAS TORRES (deslocamento {dy:.0f} m em Y)")
        print(f"   {casa} de {len(alto['idx'])} segmentos da torre de cima tem gemeo "
              f"exato na de baixo ({100*casa/max(len(alto['idx']),1):.1f}%)")
        print(f"   metragem: baixo {base['m']:.1f} m · cima {alto['m']:.1f} m "
              f"(diferenca {100*abs(base['m']-alto['m'])/max(base['m'],alto['m']):.1f}%)")
        print("   torres parecidas e nao iguais — exatamente o que o item 13.1 descreve")

    rec, apt = planilha_por_torre(obra)
    print("\n-- RECEITA REAL POR TORRE (aba RAMAL, metro liquido)")
    for aba, t in sorted(rec.items()):
        letra = "A" if aba.strip().upper().endswith("A") else "B"
        n, pav = apt[letra]
        s = sum(t.values())
        print(f"   {aba:<18} {s:9.1f} m · {n} aptos em {pav} pav · "
              f"{s/n:5.2f} m/apto  ("
              + " · ".join(f"{k[0]}{k[1]} {v:.0f}" for k, v in sorted(t.items())) + ")")
    soma_rec = sum(sum(t.values()) for t in rec.values())
    soma_apt = sum(n for n, _ in apt.values())
    print(f"   {'SOMA':<18} {soma_rec:9.1f} m · {soma_apt} aptos · "
          f"{soma_rec/soma_apt:5.2f} m/apto")

    # ---- quebra por sufixo de camada, que foi o que fez a Living fechar
    print("\n-- DESENHO POR SUFIXO DE CAMADA (as 2 torres somadas, m/apto do tipo)")
    idx_torre = [i for r in torres for i in r["idx"]]
    por_suf = defaultdict(float)
    for i in idx_torre:
        a, b, c, sis, suf = dados[i]
        por_suf[(sis, suf)] += c
    n_tipo_aptos = FINAIS * len(torres)
    # receita separada por sistema — no Peak o AQ e so DN20 e vale 1/4 do AF
    rec_sis = defaultdict(float)
    for t in rec.values():
        for (cor, dn), v in t.items():
            rec_sis[cor] += v
    alvo_sis = {k: v / soma_apt for k, v in rec_sis.items()}
    print(f"   base: {n_tipo_aptos} apartamentos ({FINAIS} finais x {len(torres)} torres)")
    print("   receita por sistema: "
          + " · ".join(f"{k} {v:.2f} m/apto" for k, v in sorted(alvo_sis.items())))
    print(f"   {'sistema':<8}{'sufixo':<16}{'metro':>9}{'m/apto':>9}"
          f"{'x receita do sistema':>22}")
    for (sis, suf), m in sorted(por_suf.items(), key=lambda kv: -kv[1]):
        mp = m / n_tipo_aptos
        raz = mp / alvo_sis.get(sis, float("nan"))
        print(f"   {sis:<8}{suf:<16}{m:>9.1f}{mp:>9.2f}{raz:>22.2f}")
    print(f"\n   {'sistema':<8}{'EXO-TET m/apto':>16}{'receita m/apto':>16}{'razao':>8}")
    for sis in ("AF", "AQ"):
        mp = por_suf.get((sis, "EXO-TET"), 0.0) / n_tipo_aptos
        a = alvo_sis.get(sis, 0.0)
        print(f"   {sis:<8}{mp:>16.2f}{a:>16.2f}{(mp/a if a else 0):>8.2f}")
    tot = sum(por_suf.values()) / n_tipo_aptos
    print(f"   {'TODAS AS CAMADAS':<8}{tot:>16.2f}{soma_rec/soma_apt:>16.2f}"
          f"{tot/(soma_rec/soma_apt):>8.2f}")

    print("\n-- CADA TORRE CONTRA A SUA ABA (m/apto)")
    print(f"   {'planta':<22}{'desenho':>9}{'torre A':>10}{'torre B':>10}")
    for r in torres:
        med = r["m"] / FINAIS
        col = []
        for letra in ("A", "B"):
            aba = [k for k in rec if k.strip().upper().endswith(letra)][0]
            col.append(med / (sum(rec[aba].values()) / apt[letra][0]))
        print(f"   y={r['y']:<20}{med:>9.2f}{col[0]:>10.2f}{col[1]:>10.2f}")


if __name__ == "__main__":
    main()
