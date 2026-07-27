# -*- coding: utf-8 -*-
"""Levantamento da estrutura dos 6 quantitativos PEX (Analise projetista)."""
import sys, glob, os
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def survey(path):
    print("=" * 90)
    print("OBRA:", os.path.basename(os.path.dirname(path)))
    print("XLSX:", os.path.basename(path))
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    for ws in wb.worksheets:
        vis = ws.sheet_state
        dim = f"{ws.max_row}x{ws.max_column}"
        flag = "" if vis == "visible" else f"  <{vis}>"
        print(f"   - {ws.title!r:32} {dim:>10}{flag}")
    wb.close()

def main():
    obras = sorted(d for d in os.listdir(BASE)
                   if os.path.isdir(os.path.join(BASE, d)) and not d.startswith("_"))
    for o in obras:
        xs = glob.glob(os.path.join(BASE, o, "*.xlsx"))
        if not xs:
            print("SEM XLSX:", o); continue
        try:
            survey(xs[0])
        except Exception as e:
            print("ERRO em", o, "->", e)
    print("=" * 90)

if __name__ == "__main__":
    main()
