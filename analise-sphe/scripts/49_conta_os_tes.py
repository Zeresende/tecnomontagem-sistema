# -*- coding: utf-8 -*-
"""OS TES ESTAO SENDO CONTADOS? (pergunta do Jose, 11/08/2026)

A pergunta tem duas leituras e as duas valem numero:

  A) como PECA — o te e uma conexao. Ele entra no levantamento pela linha de conexao
     da planilha, nao pela metragem de tubo. Aqui o script so confere se a nossa
     cadeia ja reproduz essas linhas (o script 14 fechou 77/77 em 03/07);

  B) como METRO — e a leitura que pode estar nos mordendo. O desenho CORTA o tubo em
     cada conexao: sobra um vao entre a ponta que chega e a ponta que sai. Esse vao
     nao esta em camada nenhuma, entao ninguem o mede. Se houver muitos tes, o
     somatorio dos vaos e metro real que falta no nosso numero. A "ponte de 15 a 25 cm"
     do script 30 existia para RELIGAR o percurso, mas ela nunca somou o vao a
     metragem — religava e seguia.

Este script mede B: quantos vaos existem, quanto medem, e quanto isso pesa contra a
receita. E conta os nos de grau >= 3, que sao os tes efetivamente desenhados como
bifurcacao.

Uso: python 49_conta_os_tes.py [obra ...]
"""
import sys, os, re, glob, math, logging
from collections import defaultdict, Counter
from pathlib import Path
import ezdxf
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
logging.getLogger("ezdxf").setLevel(logging.ERROR)

AQUI = Path(__file__).resolve().parent
sys.path.insert(0, str(AQUI))
m22 = __import__("22_ramal_por_dn")
m25 = __import__("25_avaliar_dn")
m29 = __import__("29_grafo_ramal")
p17 = __import__("17_parser_contagens")
BASE = m25.BASE

TUBO_LAYER = re.compile(r"H(AF|AQ)-TUB", re.I)
EXCL_LAYER = re.compile(r"LEG|REF|TAB|DET|PREVIS", re.I)
TOL_NO = 0.02
VAO_MIN = 0.03      # abaixo disso e ruido de snap, nao vao de conexao
VAO_MAX = 0.60      # acima disso nao e conexao, e trecho que falta mesmo
RX_CONEX = re.compile(r"\bTE\b|TÊ|TEE|JOELHO|COTOVELO|LUVA|CONEX|CURVA|REDU", re.I)


def andar(entidades, prof=0):
    for e in entidades:
        if e.dxftype() == "INSERT":
            nome = ""
            try:
                nome = str(e.dxf.name)
            except Exception:
                pass
            yield ("INSERT", nome, e)
            if prof < 6:
                try:
                    yield from andar(e.virtual_entities(), prof + 1)
                except Exception:
                    pass
            continue
        yield ("ENT", "", e)


def coletar(doc):
    segs, blocos = [], Counter()
    for tipo, nome, e in andar(doc.modelspace()):
        if tipo == "INSERT":
            blocos[nome] += 1
            continue
        ly = str(getattr(e.dxf, "layer", ""))
        mm = TUBO_LAYER.search(ly)
        if not mm or EXCL_LAYER.search(ly):
            continue
        pts = m22.pontos_de(e)
        for i in range(len(pts) - 1):
            a, b = pts[i], pts[i + 1]
            c = math.dist(a, b)
            if c > 0:
                segs.append({"a": a, "b": b, "m": c, "tipo": mm.group(1).upper()})
    return segs, blocos


def sistema_da_aresta(arestas, i):
    return arestas[i]["tipo"]


def vaos(adj, arestas):
    """Pares de pontas livres proximas = corte do desenho na conexao.
    Devolve [(distancia, sistema)] e o total de pontas livres."""
    livres = [no for no, viz in adj.items() if len(viz) == 1]
    cel = defaultdict(list)
    R = VAO_MAX
    pos = {}
    for no in livres:
        i, _ = adj[no][0]
        a, b = arestas[i]["pts"][0], arestas[i]["pts"][-1]
        p = a if (round(a[0] / TOL_NO), round(a[1] / TOL_NO)) == no else b
        pos[no] = p
        cel[(int(p[0] / R), int(p[1] / R))].append(no)
    usados, pares = set(), []
    for no in livres:
        if no in usados:
            continue
        p = pos[no]
        cx, cy = int(p[0] / R), int(p[1] / R)
        melhor, dist = None, 1e18
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                for outro in cel.get((cx + dx, cy + dy), ()):
                    if outro == no or outro in usados:
                        continue
                    d = math.dist(p, pos[outro])
                    if VAO_MIN <= d <= VAO_MAX and d < dist:
                        melhor, dist = outro, d
        if melhor is not None:
            usados.add(no)
            usados.add(melhor)
            pares.append((dist, sistema_da_aresta(arestas, adj[no][0][0])))
    return pares, len(livres)


# Denominadores explicitos, com fonte. Os parsers genericos erram aqui e o erro e caro:
# `planilha_ramal` soma AF e AQ como se fossem apartamentos diferentes (da 328 na Living,
# que tem 164) e `aptos_pav` conta as tres abas PREDIO do Peak (da 30, sao 10 por torre).
# Comparar receita de PREDIO com medicao de PAVIMENTO ja derrubou um resultado em 10/08.
OBRAS = {
    "20241385": {   # Living — receita do script 18, validada em 03/07
        "rec_apto": 49.75 + 35.22, "aptos_predio": 164, "aptos_tipo": 8,
        "regiao": None,
        "fonte": "planilha_ramal do script 18 (AF 49,75 + AQ 35,22 m/apto x 164 aptos)"
    },
    "20251533": {   # Peak — receita e recorte do script 48, 11/08
        "rec_apto": 10373.8 / 466, "aptos_predio": 466, "aptos_tipo": 20,
        "regiao": (100, 150, 75, 150),
        "fonte": "abas RAMAL TORRE A+B do script 48 (10.373,8 m / 466 aptos); "
                 "so as 2 plantas de torre entram, sem as plantas de furo"
    },
}


def receita(obra):
    x = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
         if "PRE-" not in a.upper() and "LEVANTAMENTO" not in a.upper()
         and not os.path.basename(a).startswith("~$")][0]
    wb = openpyxl.load_workbook(x, data_only=True)
    tubo = 0.0
    conex = Counter()
    n_aptos = 0
    for ws in wb.worksheets:
        if not ws.title.upper().startswith("RAMAL"):
            continue
        cr = p17.linha_contagem(ws)
        for c in range(8, ws.max_column + 1):
            v = ws.cell(cr, c).value
            if isinstance(v, (int, float)):
                n_aptos += v
        for r in range(cr + 1, ws.max_row + 1):
            e, u, g = ws.cell(r, 5).value, ws.cell(r, 6).value, ws.cell(r, 7).value
            if not isinstance(e, str) or not isinstance(g, (int, float)) or not g:
                continue
            un = str(u).strip().upper()
            if e.upper().startswith("TUBO") and un == "RL":
                mm = m25.RX_TUBO.search(e.upper())
                if mm:
                    tams = [int(t) for t in m25.TAM_ROLO.findall(e.upper())
                            if int(t) in (50, 100, 200)]
                    tubo += g * (tams[-1] if tams else 100) / 1.07
            elif un == "UN" and RX_CONEX.search(e.upper()):
                chave = "TÊ" if re.search(r"\bTE\b|TÊ|TEE", e.upper()) else "outra conexão"
                conex[chave] += g
    wb.close()
    return tubo, conex, n_aptos


def main():
    obras = sys.argv[1:] or ["20241385", "20251533"]
    for obra in obras:
        d = AQUI / "dxf" / obra
        cand = [p for p in sorted(d.glob("*.dxf"))
                if re.search(r"PVTIPO|TIPO|TIP\b|-TIP-", p.name, re.I)
                and not re.search(r"DET", p.name, re.I)]
        if not cand:
            continue
        print("=" * 78)
        print(f"OBRA {obra} · {cand[0].name}", flush=True)
        cfg = OBRAS.get(obra)
        doc = ezdxf.readfile(str(cand[0]))
        segs, blocos = coletar(doc)
        if cfg and cfg["regiao"]:
            x0, x1, y0, y1 = cfg["regiao"]
            antes = len(segs)
            segs = [s for s in segs
                    if x0 <= (s["a"][0] + s["b"][0]) / 2 <= x1
                    and y0 <= (s["a"][1] + s["b"][1]) / 2 <= y1]
            print(f"   recorte de regiao {cfg['regiao']}: {antes} -> {len(segs)} segmentos")
        total = sum(s["m"] for s in segs)
        adj, arestas = m29.construir(segs, TOL_NO)
        graus = Counter(len(v) for v in adj.values())
        pares, n_livres = vaos(adj, arestas)

        print(f"   tubo medido            : {total:9.1f} m em {len(segs)} segmentos")
        print(f"   nos do grafo           : {len(adj)}  "
              + " · ".join(f"grau {g}: {n}" for g, n in sorted(graus.items())[:5]))
        print(f"   nos de grau >= 3 (tês) : {sum(n for g, n in graus.items() if g >= 3)}")
        print(f"   pontas livres          : {n_livres}")
        print(f"   VAOS de conexao        : {len(pares)} pares entre "
              f"{VAO_MIN} e {VAO_MAX} m")
        if pares:
            s = sum(v for v, _ in pares)
            ord_ = sorted(v for v, _ in pares)
            print(f"      metro nao desenhado : {s:8.2f} m  "
                  f"({100*s/total:.2f}% do medido)")
            print(f"      vao medio {s/len(pares):.3f} m · mediana "
                  f"{ord_[len(ord_)//2]:.3f} m · maior {ord_[-1]:.3f} m")
            por_sis = defaultdict(float)
            for v, sis in pares:
                por_sis[sis] += v
            print("      por sistema:",
                  " · ".join(f"{k} {v:.1f} m" for k, v in sorted(por_sis.items())))
            faixa = Counter(round(v, 1) for v, _ in pares)
            print("      distribuicao:", dict(sorted(faixa.items())))

        bl = {k: v for k, v in blocos.items() if RX_CONEX.search(k)}
        print(f"   blocos de conexao no desenho: {sum(bl.values())} "
              f"{dict(Counter(bl).most_common(6)) if bl else '(nenhum com nome de conexao)'}")

        try:
            _, conex_rec, _ = receita(obra)
            print(f"   conexoes na planilha (aba RAMAL): "
                  + (" · ".join(f"{k} {v:.0f}" for k, v in conex_rec.items())
                     if conex_rec else "nenhuma linha UN de conexao"))
            if not cfg:
                continue
            # ESCALA: receita e do PREDIO, medicao e de UM pavimento tipo.
            rec_pav = cfg["rec_apto"] * cfg["aptos_tipo"]
            print(f"   receita: {cfg['rec_apto']:.2f} m/apto x {cfg['aptos_tipo']} aptos "
                  f"do tipo = {rec_pav:.1f} m")
            print(f"      fonte: {cfg['fonte']}")
            if pares:
                s = sum(v for v, _ in pares)
                print(f"   >> O VAO VALE {s:.1f} m = {100*s/rec_pav:.1f}% da receita "
                      f"do mesmo pavimento")
            if conex_rec.get("TÊ"):
                te_pav = conex_rec["TÊ"] / cfg["aptos_predio"] * cfg["aptos_tipo"]
                print(f"      tês da planilha, equivalentes a este pavimento: "
                      f"{te_pav:.0f} · vaos encontrados no desenho: {len(pares)}")
        except Exception as erro:
            print(f"   receita: falhou ({erro})")


if __name__ == "__main__":
    main()
