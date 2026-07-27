# -*- coding: utf-8 -*-
"""Reconciliacao generalizada SPHE: contagens das abas de kits x derivacao do RESUMO.
Testa 2 regras por coluna de kit com FINAIS no header:
  strict  = soma rep*aptos dos grupos cujos finais estao CONTIDOS no header
  overlap = soma rep*|finais do grupo ∩ finais do header|  (formaliza a "regra de dobra":
            terreo/20o/duplex sao absorvidos na coluna de receita mais parecida)
Para kits nomeados (chuveiro/travessa/...) deriva de totais de ambiente do RESUMO.
Tambem mede o que cada planilha tem de recuperavel (receitas preenchidas, G>0).
"""
import sys, os, glob, re
from collections import defaultdict
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OBRAS = ["20241385", "20241390", "20251670", "20251430", "20251533"]
FINAIS_RE = re.compile(r"FINAIS?\s*([\d/,\sE]+)", re.I)
DN_RE = re.compile(r"PEX\s*(\d{2})", re.I)


def finais_de(texto):
    m = FINAIS_RE.search(texto or "")
    return sorted({int(t) for t in re.findall(r"\d", m.group(1))}) if m else None


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def resumo_grupos(wb):
    ws = wb["RESUMO"]
    ambs = {}
    for c in range(5, ws.max_column + 1):
        v = ws.cell(2, c).value
        if isinstance(v, str) and v.strip():
            ambs[c] = v.strip().upper()
    grupos = []
    for r in range(3, ws.max_row + 1):
        pav = ws.cell(r, 1).value
        if not isinstance(pav, str) or not pav.strip():
            continue
        rep = num(ws.cell(r, 2).value)
        fins = finais_de(str(ws.cell(r, 3).value or "")) or []
        apt = num(ws.cell(r, 4).value)
        amb = {ambs[c]: num(ws.cell(r, c).value) for c in ambs if isinstance(ws.cell(r, c).value, (int, float))}
        grupos.append((pav.strip(), rep, fins, apt, amb))
    return grupos


def linha_contagem(ws):
    best, br = -1, 3
    for r in (2, 3, 4, 5):
        n = sum(1 for c in range(8, ws.max_column + 1) if isinstance(ws.cell(r, c).value, (int, float)))
        if n > best:
            best, br = n, r
    return br


def abas_kit(wb):
    for ws in wb.worksheets:
        t = ws.title.upper()
        if t.startswith("RAMAL") or t.startswith("KIT") or t.startswith("CHICOTE"):
            yield ws


def deriva(nome, grupos, tot_amb, aptos_total):
    """Retorna dict {regra: valor} candidatos para a contagem da coluna."""
    up = nome.upper()
    fins = finais_de(nome)
    cand = {}
    if fins:
        cand["strict"] = sum(r * a for _, r, gf, a, _ in grupos if gf and set(gf) <= set(fins))
        cand["overlap"] = sum(r * len(set(gf) & set(fins)) for _, r, gf, a, _ in grupos if gf)
    for amb, tot in tot_amb.items():
        chave = amb.replace("BNH", "BANHO")
        if any(p in up for p in (chave, chave.replace("BANHO TIPO", "CHUVEIRO"))):
            cand[f"amb:{amb}"] = tot
    if "CHUVEIRO" in up or "BANHO" in up:
        cand["banhos"] = tot_amb.get("BANHO TIPO", 0) + tot_amb.get("BANHO MASTER", 0)
    if "TRAVESSA" in up and "MANIFOLD" not in up:
        cand["aptos"] = aptos_total
    return cand


def analisa(obra):
    xs = glob.glob(os.path.join(BASE, obra, "*.xlsx"))
    if not xs:
        print(f"== {obra}: SEM XLSX")
        return
    x = xs[0]
    wb = openpyxl.load_workbook(x, data_only=True)
    print("=" * 90)
    print(f"OBRA {obra} | {os.path.basename(x)}")
    grupos = resumo_grupos(wb)
    tot_amb = defaultdict(float)
    aptos_total = 0.0
    for _, rep, fins, apt, amb in grupos:
        aptos_total += rep * apt
        for a, q in amb.items():
            tot_amb[a] += rep * apt * q
    print(f"  RESUMO: {len(grupos)} grupos | aptos={aptos_total:.0f} | " +
          " | ".join(f"{k}={v:.0f}" for k, v in tot_amb.items()))
    if not grupos:
        print("  (RESUMO vazio/quebrado - sem base p/ derivar)")

    ok = tot = 0
    soma_cont = defaultdict(float)
    print(f"  {'aba':16} {'kit':46} {'cont':>6}  match")
    for ws in abas_kit(wb):
        cr = linha_contagem(ws)
        for c in range(8, ws.max_column + 1):
            cnt = ws.cell(cr, c).value
            if not isinstance(cnt, (int, float)):
                continue
            h = " / ".join(str(ws.cell(r, c).value or "").replace("\n", " ").strip()
                           for r in range(1, cr) if ws.cell(r, c).value)
            cand = deriva(h, grupos, tot_amb, aptos_total)
            hit = [k for k, v in cand.items() if abs(v - cnt) < 0.5]
            perto = min(cand.items(), key=lambda kv: abs(kv[1] - cnt))[0] if cand else None
            stat = f"= {hit[0]}" if hit else (f"~ {perto}({cand[perto]:.0f})" if perto else "sem regra")
            tot += 1
            ok += bool(hit)
            soma_cont[ws.title] += cnt
            print(f"  {ws.title[:16]:16} {h[:46]:46} {cnt:6.0f}  {stat}")
    print(f"  DERIVADAS: {ok}/{tot} exatas | soma contagens por aba: " +
          ", ".join(f"{k}={v:.0f}" for k, v in soma_cont.items()))

    # recuperavel: receitas preenchidas e G>0 nas linhas de tubo
    rec = defaultdict(lambda: [0, 0, 0])  # dn -> [linhas, com_receita, rolos_G]
    for ws in abas_kit(wb):
        cr = linha_contagem(ws)
        for r in range(cr + 1, ws.max_row + 1):
            e = ws.cell(r, 5).value
            if not isinstance(e, str) or "TUBO PEX" not in e.upper() or "BARRA" in e.upper():
                continue
            m = DN_RE.search(e)
            if not m:
                continue
            dn = int(m.group(1))
            temrec = any(isinstance(ws.cell(r, c).value, (int, float)) and ws.cell(r, c).value
                         for c in range(8, ws.max_column + 1))
            g = num(ws.cell(r, 7).value)
            rec[dn][0] += 1
            rec[dn][1] += bool(temrec)
            rec[dn][2] += g
    print("  RECUPERAVEL (linhas de tubo): " + " | ".join(
        f"DN{dn}: {v[1]}/{v[0]} c/receita, G={v[2]:.0f} rolos" for dn, v in sorted(rec.items())))
    wb.close()


def main():
    for obra in OBRAS:
        analisa(obra)


if __name__ == "__main__":
    main()
