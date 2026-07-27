# -*- coding: utf-8 -*-
"""Le RESUMO + detecta nomenclatura de DN (PEX) de cada quantitativo."""
import sys, glob, os, re
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def dump_resumo(ws):
    print("   --- RESUMO ---")
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() for c in row if c not in (None, "")]
        if cells:
            print("   |", " | ".join(cells))

def detect_dn(wb):
    """Procura tokens tipo PEX 16, DN 20, 1/2, etc. nas abas RAMAL/KITS."""
    dns = {}
    pat = re.compile(r"\b(PEX|TUBO|DN|Ø|%%C)?\s*\b(12|13|16|19|20|22|25|28|32|40|50)\b")
    for ws in wb.worksheets:
        t = ws.title.upper()
        if not ("RAMAL" in t or "KIT" in t or "CHICOTE" in t):
            continue
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 330), values_only=True):
            for c in row:
                if isinstance(c, str) and ("PEX" in c.upper() or "TUBO" in c.upper()):
                    for m in re.finditer(r"\b(12|13|16|19|20|22|25|28|32|40|50)\b", c):
                        dns.setdefault(m.group(1), c.strip()[:50])
    return dns

def main():
    obras = sorted(d for d in os.listdir(BASE)
                   if os.path.isdir(os.path.join(BASE, d)) and not d.startswith("_"))
    for o in obras:
        xs = glob.glob(os.path.join(BASE, o, "*.xlsx"))
        if not xs: continue
        print("=" * 90)
        print("OBRA:", o, "|", os.path.basename(xs[0]))
        wb = openpyxl.load_workbook(xs[0], data_only=True)
        for ws in wb.worksheets:
            if ws.title.upper().startswith("RESUMO"):
                dump_resumo(ws)
        dns = detect_dn(wb)
        print("   --- DNs de PEX/TUBO detectados ---")
        for k in sorted(dns, key=int):
            print(f"   DN {k:>3}  ex: {dns[k]}")
        wb.close()
    print("=" * 90)

if __name__ == "__main__":
    main()
