# -*- coding: utf-8 -*-
"""Receita completa (tubo + conexao) dos kits CHUVEIRO / BANHO / LAVABO / COZINHA
nas obras SPHE, para entregar a Karina (pedido de 12/08/2026).

A `biblioteca_receitas.txt` do script 08 so guarda TUBO por DN. A receita de CONEXAO
existe nas mesmas colunas da planilha, mas nunca foi exportada — o script 14 so
conferiu que ela fecha (77/77), sem despejar os valores. Este script despeja.

Estrutura da planilha, ja decodificada em 19/06:
  col 5 = descricao da peca · col 6 = unidade (RL/BR/UN) · col 7 = G (compra)
  col 8+ = uma coluna por KIT; linha de contagem = quantas vezes o kit ocorre;
  celula [linha da peca, coluna do kit] = RECEITA (quanto da peca em 1 kit):
      conexao em UN/kit; TUBO EM METROS/KIT (nao rolos). Prova: a coluna G e
      ROUNDUP(soma(celula x contagem) / tamanho_rolo x 1,07) nas 5 obras (04/09/2026).

Camada de CORRECOES AUDITADAS (04/09/2026): a planilha-fonte pode ter erro que o
Marcelo/Hederson corrigem depois (ex.: Pamaris BANHO 2 sem o te 20-16-16). A fonte
nao e nossa pra editar, entao a correcao vive em `saida/correcoes_receita_sphe.csv`
e e aplicada por cima da extracao na hora de gravar o CSV. Assim, re-rodar este
script NAO desfaz a correcao. Linhas com status diferente de `aplicada` (ex.:
`pendente_qtd`) so sao avisadas, nunca entram no CSV.

Uso: python 51_receita_kits_4.py [--csv]
"""
import sys, os, re, glob
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
AQUI = Path(__file__).resolve().parent
sys.path.insert(0, str(AQUI))
p17 = __import__("17_parser_contagens")
BASE = AQUI.parent

OBRAS = {"20241385": "Living", "20241390": "Edition", "20251430": "Brooklyn",
         "20251533": "Peak", "20251670": "Pamaris"}

# Os 4 kits que a Karina pediu. Casamento por palavra-chave no cabecalho da coluna,
# porque cada obra nomeia diferente (CHICOTE BANHO 1, KIT CHUVEIRO H, COZ./A.S...).
ALVOS = {
    "CHUVEIRO": re.compile(r"CHUVEIRO", re.I),
    "BANHO":    re.compile(r"BANHO", re.I),
    "LAVABO":   re.compile(r"LAVABO", re.I),
    "COZINHA":  re.compile(r"COZINHA|COZ\.", re.I),
}


def planilha(obra):
    arqs = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
            if "PRE-" not in a.upper() and "LEVANTAMENTO" not in a.upper()
            and not os.path.basename(a).startswith("~$")]
    return arqs[0] if arqs else None


def cabecalho_kit(ws, col, linha_cont):
    """Nome do kit = TODAS as strings acima da linha de contagem, juntas.

    Nao basta pegar a primeira: no Peak a primeira e o codigo da folha ('F.1116') e
    o nome do kit vem na linha seguinte ('CHICOTE BANHO AQ/AF'). Ler so a primeira
    fazia os 4 kits do Peak sumirem — e viraria um 'essa obra nao tem' falso."""
    partes = []
    for r in range(1, linha_cont + 1):
        v = ws.cell(r, col).value
        if isinstance(v, str) and len(v.strip()) >= 3:
            partes.append(" ".join(v.split()))
    return " · ".join(partes)


def coletar(obra):
    caminho = planilha(obra)
    if not caminho:
        return {}
    wb = openpyxl.load_workbook(caminho, data_only=True)
    achados = {}
    for ws in wb.worksheets:
        t = ws.title.upper()
        if not (t.startswith("RAMAL") or t.startswith("KIT")
                or t.startswith("CHICOTE")):
            continue
        cr = p17.linha_contagem(ws)
        for col in range(8, ws.max_column + 1):
            cont = ws.cell(cr, col).value
            if not isinstance(cont, (int, float)) or not cont:
                continue
            nome = cabecalho_kit(ws, col, cr)
            alvo = next((k for k, rx in ALVOS.items() if rx.search(nome)), None)
            if not alvo:
                continue
            itens = []
            for r in range(cr + 1, ws.max_row + 1):
                desc = ws.cell(r, 5).value
                un = ws.cell(r, 6).value
                v = ws.cell(r, col).value
                if not isinstance(desc, str) or not isinstance(v, (int, float)) or not v:
                    continue
                itens.append((desc.strip(), str(un or "").strip().upper(), float(v)))
            if itens:
                achados.setdefault(alvo, []).append(
                    {"aba": ws.title, "coluna": nome, "contagem": cont, "itens": itens})
    wb.close()
    return achados


CORRECOES = AQUI / "saida" / "correcoes_receita_sphe.csv"


def aplicar_correcoes(linhas):
    """Aplica `correcoes_receita_sphe.csv` sobre as linhas extraidas (cabecalho incluso).

    acao=renomear: troca `peca` por `peca_nova` na linha (obra, coluna_planilha, peca).
    acao=adicionar: insere (obra, coluna_planilha, peca, unidade, receita), copiando
    kit_alvo/aba/contagem de uma linha ja existente da mesma coluna.
    So status=aplicada entra; o resto e listado como pendencia."""
    import csv as _csv
    if not CORRECOES.exists():
        return linhas, []
    with open(CORRECOES, encoding="utf-8-sig", newline="") as f:
        regras = list(_csv.DictReader(f, delimiter=";"))
    cab, corpo = linhas[0], [list(l) for l in linhas[1:]]
    avisos = []
    for rg in regras:
        chave = (rg["obra"], rg["coluna_planilha"])
        rotulo = f"{rg['obra']} · {rg['coluna_planilha']} · {rg['acao']} {rg['peca']}"
        if rg["status"] != "aplicada":
            avisos.append(f"PENDENTE [{rg['status']}] {rotulo} — {rg['motivo']}")
            continue
        mesma_coluna = [l for l in corpo if (l[0], l[2]) == chave]
        if not mesma_coluna:
            avisos.append(f"IGNORADA (coluna nao existe na extracao) {rotulo}")
            continue
        if rg["acao"] == "renomear":
            alvo = [l for l in mesma_coluna if l[5] == rg["peca"]]
            if len(alvo) != 1:
                avisos.append(f"IGNORADA ({len(alvo)} linhas casaram) {rotulo}")
                continue
            alvo[0][5] = rg["peca_nova"]
            avisos.append(f"aplicada: {rotulo} -> {rg['peca_nova']}")
        elif rg["acao"] == "adicionar":
            if any(l[5] == rg["peca"] for l in mesma_coluna):
                avisos.append(f"IGNORADA (peca ja existe na coluna) {rotulo}")
                continue
            modelo = mesma_coluna[0]
            nova = [modelo[0], modelo[1], modelo[2], modelo[3], modelo[4],
                    rg["peca"], rg["unidade"], float(rg["receita"])]
            pos = corpo.index(mesma_coluna[-1]) + 1
            corpo.insert(pos, nova)
            avisos.append(f"aplicada: {rotulo} = {rg['receita']} {rg['unidade']}")
        else:
            avisos.append(f"IGNORADA (acao desconhecida) {rotulo}")
    return [cab] + [tuple(l) for l in corpo], avisos


def main():
    csv = "--csv" in sys.argv
    linhas_csv = [("obra", "kit_alvo", "coluna_planilha", "aba", "contagem",
                   "peca", "unidade", "receita")]
    for obra, apelido in OBRAS.items():
        achados = coletar(obra)
        if not achados:
            continue
        print("=" * 78)
        print(f"OBRA {obra} · {apelido}")
        for alvo in ("CHUVEIRO", "BANHO", "LAVABO", "COZINHA"):
            for bloco in achados.get(alvo, []):
                print(f"\n  [{alvo}] coluna \"{bloco['coluna']}\" "
                      f"(aba {bloco['aba']}, ocorre {bloco['contagem']:.0f}x)")
                tubo = [i for i in bloco["itens"] if i[1] in ("RL", "M", "BR")]
                conex = [i for i in bloco["itens"] if i[1] not in ("RL", "M", "BR")]
                for rotulo, grupo in (("tubo", tubo), ("conexao", conex)):
                    if not grupo:
                        continue
                    print(f"    {rotulo}:")
                    for desc, un, v in grupo:
                        print(f"      {v:>8.3f} {un:<3} {desc[:58]}")
                for desc, un, v in bloco["itens"]:
                    linhas_csv.append((obra, alvo, bloco["coluna"], bloco["aba"],
                                       bloco["contagem"], desc, un, v))
        print()

    linhas_csv, avisos = aplicar_correcoes(linhas_csv)
    if avisos:
        print("CORRECOES AUDITADAS (saida/correcoes_receita_sphe.csv):")
        for a in avisos:
            print(f"  {a}")

    if csv:
        import csv as _csv
        saida = AQUI / "saida" / "receita_kits_4_sphe.csv"
        with open(saida, "w", newline="", encoding="utf-8-sig") as f:
            _csv.writer(f, delimiter=";").writerows(linhas_csv)
        print(f"CSV: {saida}  ({len(linhas_csv)-1} linhas)")


if __name__ == "__main__":
    main()
