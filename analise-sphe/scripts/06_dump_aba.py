# -*- coding: utf-8 -*-
"""Dump cru de uma aba: linhas 1-4 + todas as linhas de TUBO PEX com TODAS as colunas nao vazias."""
import sys, glob, os
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
obra, aba = sys.argv[1], sys.argv[2]
x = glob.glob(os.path.join(BASE, obra, "*.xlsx"))[0]
wb = openpyxl.load_workbook(x, data_only=True)
ws = next(w for w in wb.worksheets if w.title.upper().startswith(aba.upper()))
print("XLSX", os.path.basename(x), "| ABA", ws.title, f"{ws.max_row}x{ws.max_column}")
L = openpyxl.utils.get_column_letter
for r in range(1, 5):
    nz = [f"{L(c)}{r}={ws.cell(r,c).value}" for c in range(1, ws.max_column+1) if ws.cell(r,c).value not in (None,"")]
    print(f"L{r}:", " | ".join(nz[:18]))
print("--- TUBO PEX ---")
for r in range(1, ws.max_row+1):
    e = ws.cell(r,5).value
    if isinstance(e,str) and "TUBO PEX" in e.upper() and "ROLO" not in e.upper() and "BARRA" not in e.upper():
        nz = [f"{L(c)}={ws.cell(r,c).value}" for c in range(1, ws.max_column+1) if ws.cell(r,c).value not in (None,"")]
        print(f"L{r}:", " | ".join(nz))
wb.close()
