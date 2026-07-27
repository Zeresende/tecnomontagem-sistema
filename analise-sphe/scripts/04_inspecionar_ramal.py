# -*- coding: utf-8 -*-
"""Inspeciona cabecalho e linhas de TUBO PEX da aba RAMAL/KITS de uma obra."""
import sys, glob, os
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

obra = sys.argv[1]
x = glob.glob(os.path.join(BASE, obra, "*.xlsx"))[0]
print("XLSX:", os.path.basename(x))
wb = openpyxl.load_workbook(x, data_only=True)
for alvo in ("RAMAL", "KITS"):
    ws = next((w for w in wb.worksheets if w.title.upper().startswith(alvo)), None)
    if not ws: continue
    print("\n==== ABA", ws.title, f"({ws.max_row}x{ws.max_column}) ====")
    # primeiras 6 linhas (cabecalho)
    for r in range(1, 7):
        cells = [(c, ws.cell(r, c).value) for c in range(1, min(ws.max_column, 16)+1)]
        nonempty = [f"{openpyxl.utils.get_column_letter(c)}={v}" for c, v in cells if v not in (None, "")]
        if nonempty:
            print(f"  L{r}:", " | ".join(nonempty[:12]))
    # linhas com TUBO PEX
    print("  --- linhas TUBO PEX ---")
    for r in range(1, ws.max_row+1):
        desc = None
        for c in range(1, 6):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "TUBO PEX" in v.upper():
                desc = (c, v); break
        if desc:
            vals = [(openpyxl.utils.get_column_letter(c), ws.cell(r, c).value)
                    for c in range(1, min(ws.max_column, 12)+1) if ws.cell(r, c).value not in (None, "")]
            print(f"   L{r}:", " | ".join(f"{cl}={vv}" for cl, vv in vals))
wb.close()
