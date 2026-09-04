# -*- coding: utf-8 -*-
"""Gabarito da obra que ficou de fora (rodada holdout), a partir do quantitativo dela.

O quantitativo SPHE e a resposta do Hederson/Marcelo. Este script o transforma no formato
que o conector emite (PECA_ID x QTD_TOTAL, contrato de 12/08) para a comparacao ser
mecanica. Regras, conferidas na formula da coluna G das 5 obras (04/09/2026):
  - celula [peca, coluna de kit] = receita POR KIT: conexao em UN/kit, tubo em METROS/kit
  - total de conexao = soma(celula x contagem), exato
  - metros de tubo   = soma(celula x contagem); compra = ROUNDUP(metros x 1,07 / rolo)
  - QTD_TOTAL do tubo sai em METROS (unidade do catalogo, ids 1049/1052/1055 = "M");
    os rolos do Hederson (coluna G) vao numa coluna a parte
Correcoes auditadas (saida/correcoes_receita_sphe.csv, status=aplicada) entram na matriz
ANTES dos totais; as pendentes vao para a aba PENDENTES.

Uso: python 62_gabarito_holdout.py --obra 20251670
Saida: saida/holdout_<apelido>/gabarito_<apelido>.xlsx (+ .csv da aba POR_PECA_ID)
"""
import argparse
import csv
import glob
import importlib
import io
import math
import os
import re
import sys
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

sys.stdout.reconfigure(encoding="utf-8")
AQUI = Path(__file__).resolve().parent
BASE = AQUI.parent
SISTEMA = BASE.parent / "sistema"
sys.path.insert(0, str(AQUI))
sys.path.insert(0, str(SISTEMA))
p17 = importlib.import_module("17_parser_contagens")
m03 = importlib.import_module("03_gerar_planilhas")
m04 = importlib.import_module("04_importar_template")

OBRAS = {"20241385": "Living", "20241390": "Edition", "20251430": "Brooklyn",
         "20251533": "Peak", "20251670": "Pamaris"}
ROLO_PADRAO = {"16": 200, "20": 100, "25": 100, "32": 50}     # PADRAO_SPHE.yaml tamanho_rolo_m
RX_ROLO = re.compile(r"([0-9]+)[ ]*M(?:[^A-Z]|$)")
RX_DN = re.compile(r"(?:PERT|PEX)[ ]*(16|20|25|32)", re.I)
FOLGA = 1.07

CAMPOS_ABERTURA = {
    "20251670": [
        ("obra", "PAMARIS (20251670), Cyrela, projetista SPHE"),
        ("torres / pavimentos tipo / finais", "1 torre - 20 pavimentos - 39 finais = 780 aptos (aba PREDIO)"),
        ("linha de produto", "PEX Serie 5 (rolos: O16 = 200 m, O20 = 100 m, O25 = 100 m)"),
        ("grupo de finais - banho", "CHICOTE BANHO 1 = finais 1-18 e 25-39 (660) - CHICOTE BANHO 2 = finais 19-24 (120)"),
        ("area de servico", "TRAVESSA TANQUE (120) e CHICOTE ASV (120) nos finais 19-24 - QUADRO ASV (RG/AQ) em todos (780)"),
        ("registro do chuveiro", "BASE REG GAVETA 3/4 DN20-B (2/kit) + BASE REG PRESSAO MVS 1/2 DN15-B (2/kit)"),
        ("ramal", "100% DN25 (alerta do PADRAO_SPHE_OBRAS: qualquer extrator acerta o DN)"),
        ("fonte", "PAMARIS-QUANTITATIVO PEX-R00.xlsx - PROIBIDA durante a rodada; so entra no fim"),
    ],
}


def planilha(obra):
    arqs = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
            if "PRE-" not in a.upper() and "LEVANTAMENTO" not in a.upper()
            and not os.path.basename(a).startswith("~$")]
    return arqs[0] if arqs else None


def tamanho_rolo(desc, dn):
    m = RX_ROLO.findall(desc.upper())
    return int(m[-1]) if m else ROLO_PADRAO.get(dn or "", 100)


def ler_matriz(ws):
    """Uma aba de kit/ramal -> (contagem por coluna, cabecalho por coluna, linhas)."""
    cr = p17.linha_contagem(ws)
    mapa = m04.mapear_aba(ws) or {"col_forn": {}}
    cont, nome = {}, {}
    for c in range(8, ws.max_column + 1):
        v = ws.cell(cr, c).value
        if isinstance(v, (int, float)) and v:
            cont[c] = float(v)
            nome[c] = " - ".join(" ".join(str(ws.cell(r, c).value).split())
                                 for r in range(1, cr) if ws.cell(r, c).value)
    linhas = {}
    for r in range(cr + 1, ws.max_row + 1):
        desc = ws.cell(r, 5).value
        if not isinstance(desc, str) or not desc.strip():
            continue
        desc = desc.strip()
        celulas = {c: float(ws.cell(r, c).value) for c in cont
                   if isinstance(ws.cell(r, c).value, (int, float)) and ws.cell(r, c).value}
        codigos = [(f, m04.norm_codigo(ws.cell(r, col).value)) for col, f in mapa["col_forn"].items()
                   if m04.norm_codigo(ws.cell(r, col).value)]
        g = ws.cell(r, 7).value
        linhas[desc] = {"unid": str(ws.cell(r, 6).value or "").strip().upper(), "celulas": celulas,
                        "codigos": codigos, "G": float(g) if isinstance(g, (int, float)) else 0.0}
    return cont, nome, linhas


def aplicar_correcoes(obra, abas):
    """Aplica o overlay do 51 na matriz. Devolve (aplicadas, pendentes)."""
    arq = AQUI / "saida" / "correcoes_receita_sphe.csv"
    if not arq.exists():
        return [], []
    with io.open(arq, encoding="utf-8-sig", newline="") as fh:
        regras = [r for r in csv.DictReader(fh, delimiter=";") if r["obra"] == obra]
    aplicadas, pendentes = [], []
    for rg in regras:
        if rg["status"] != "aplicada":
            pendentes.append(rg)
            continue
        for aba, (cont, nome, linhas) in abas.items():
            cols = [c for c, n in nome.items() if n == rg["coluna_planilha"]]
            if not cols:
                continue
            c = cols[0]
            if rg["acao"] == "renomear":
                v = linhas[rg["peca"]]["celulas"].pop(c, None)
                if v is not None:
                    novo = linhas.setdefault(rg["peca_nova"], {"unid": rg["unidade"], "celulas": {},
                                                               "codigos": [], "G": 0.0})
                    novo["celulas"][c] = v
                    aplicadas.append(f"{aba} / {rg['coluna_planilha']}: {rg['peca']} -> {rg['peca_nova']} ({v:g}/kit)")
            elif rg["acao"] == "adicionar":
                novo = linhas.setdefault(rg["peca"], {"unid": rg["unidade"], "celulas": {},
                                                      "codigos": [], "G": 0.0})
                novo["celulas"][c] = float(rg["receita"])
                aplicadas.append(f"{aba} / {rg['coluna_planilha']}: +{rg['peca']} = {rg['receita']}/kit")
    return aplicadas, pendentes


def tubo_por_bitola(catalogo, dn):
    for p, x in catalogo.items():
        if x["sistema"] == "PEX" and str(x["descricao"]).upper().startswith(f"TUBO PEX {dn} - S"):
            return p
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--obra", required=True)
    a = ap.parse_args()
    apelido = OBRAS[a.obra]
    saida = AQUI / "saida" / f"holdout_{apelido.lower()}"
    saida.mkdir(parents=True, exist_ok=True)
    caminho = planilha(a.obra)
    catalogo = m03.carregar_catalogo()
    equivalencias = m03.carregar_equivalencias()
    cod_idx, desc_idx = m04.construir_indices(catalogo, equivalencias)

    wb = openpyxl.load_workbook(caminho, data_only=True)
    abas = {}
    for ws in wb.worksheets:
        t = ws.title.upper()
        if t.startswith(("KIT", "CHICOTE", "RAMAL")):
            abas[ws.title] = ler_matriz(ws)
    aplicadas, pendentes = aplicar_correcoes(a.obra, abas)

    por_kit, sem_id = [], []
    por_pid = defaultdict(lambda: {"qtd": 0.0, "rolos": 0.0, "origem": set()})
    bitolas = defaultdict(lambda: defaultdict(float))
    for aba, (cont, nome, linhas) in abas.items():
        for desc, L in linhas.items():
            if not L["celulas"]:
                continue
            tubo = desc.upper().startswith("TUBO") and L["unid"] in ("RL", "M", "BR")
            m_dn = RX_DN.search(desc.upper())
            dn = m_dn.group(1) if m_dn else None
            rolo = tamanho_rolo(desc, dn) if tubo else None
            total = sum(v * cont[c] for c, v in L["celulas"].items())
            pid, metodo, _ = m04.casar_linha(L["codigos"], desc, "PEX", cod_idx, desc_idx, catalogo)
            if pid is None and tubo and dn:
                pid = tubo_por_bitola(catalogo, dn)      # linha criada por correcao, sem codigo
            for c, v in L["celulas"].items():
                por_kit.append([aba, nome[c], int(cont[c]), desc, "M/KIT" if tubo else "UN/KIT", v,
                                pid or "", round(v * cont[c], 1)])
            if pid is None:
                sem_id.append([aba, desc, L["unid"], round(total, 1)])
                continue
            por_pid[pid]["qtd"] += total
            por_pid[pid]["origem"].add(aba)
            if tubo:
                rolos = L["G"] or math.ceil(total * FOLGA / rolo)
                por_pid[pid]["rolos"] += rolos
                por_pid[pid]["rolo"] = rolo
                bitolas[dn or "?"][aba] += total
                bitolas[dn or "?"]["rolos"] += rolos

    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "POR_PECA_ID"
    ws.append(["PECA_ID", "SISTEMA", "DESCRICAO", "UNIDADE", "QTD_TOTAL", "TIPO",
               "ROLOS_HEDERSON", "TAMANHO_ROLO_M", "ORIGEM_ABAS"])
    linhas_csv = []
    for pid in sorted(por_pid):
        p, d = catalogo[pid], por_pid[pid]
        tipo = "TUBO" if d.get("rolo") else "CONEXAO"
        row = [pid, p["sistema"], p["descricao"], p["unidade"], round(d["qtd"], 1), tipo,
               (int(d["rolos"]) if tipo == "TUBO" else ""), d.get("rolo", ""),
               " + ".join(sorted(d["origem"]))]
        ws.append(row)
        linhas_csv.append(row)
    ws2 = out.create_sheet("POR_KIT")
    ws2.append(["ABA", "COLUNA", "CONTAGEM", "PECA", "UNIDADE", "RECEITA_POR_KIT", "PECA_ID", "TOTAL_COLUNA"])
    for r in por_kit:
        ws2.append(r)
    ws3 = out.create_sheet("TUBO_POR_BITOLA")
    abas_nomes = sorted(abas)
    ws3.append(["BITOLA"] + [f"METROS_{x}" for x in abas_nomes]
               + ["METROS_TOTAL", "ROLOS_HEDERSON", "OBS"])
    for dn in sorted(bitolas):
        ms = [round(bitolas[dn].get(x, 0.0), 1) for x in abas_nomes]
        ws3.append([f"O{dn}"] + ms + [round(sum(ms), 1), int(bitolas[dn]["rolos"]),
                    "rolos = ROUNDUP(metros x 1,07 / rolo), coluna G do quantitativo"])
    ws4 = out.create_sheet("PENDENTES")
    ws4.append(["COLUNA", "ACAO", "PECA", "STATUS", "MOTIVO"])
    for rg in pendentes:
        ws4.append([rg["coluna_planilha"], rg["acao"], rg["peca"], rg["status"], rg["motivo"]])
    for s in sem_id:
        ws4.append(["(sem PECA_ID)", "cadastrar", s[1], "sem_id", f"{s[0]}: {s[3]} {s[2]}"])
    ws5 = out.create_sheet("CAMPOS_ABERTURA")
    ws5.append(["CAMPO", "VALOR DECLARADO"])
    for k, v in CAMPOS_ABERTURA.get(a.obra, []):
        ws5.append([k, v])
    ws6 = out.create_sheet("CORRECOES_APLICADAS")
    ws6.append(["CORRECAO"])
    for x in aplicadas:
        ws6.append([x])
    for w in out.worksheets:
        for c in w[1]:
            c.font = Font(bold=True)
        w.freeze_panes = "A2"
    out.save(saida / f"gabarito_{apelido.lower()}.xlsx")
    with io.open(saida / f"gabarito_{apelido.lower()}_por_peca.csv", "w",
                 encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow([c.value for c in ws[1]])
        w.writerows(linhas_csv)

    print(f"gabarito {apelido}: {len(por_pid)} PECA_ID / {len(por_kit)} celulas de kit / "
          f"{len(pendentes)} pendentes / {len(sem_id)} sem id")
    for x in aplicadas:
        print("  correcao aplicada:", x)
    for dn in sorted(bitolas):
        metros = sum(v for k, v in bitolas[dn].items() if k != "rolos")
        print(f"  O{dn}: {metros:.0f} m -> {int(bitolas[dn]['rolos'])} rolos (Hederson)")
    for s in sem_id:
        print("  SEM PECA_ID:", s)
    print(f"saida: {saida}")


if __name__ == "__main__":
    main()
