# -*- coding: utf-8 -*-
"""Gera LEVANTAMENTO COMPLETO (tubos + conexoes) para Brooklyn e Peak.
Extensao do 13_pre_quantitativo: em vez de somar so metros de tubo por DN,
empresta a receita ITEM A ITEM de cada kit (col E das obras completas) e aplica:
  conexoes: qtde = soma(receita_item_kit x contagem_kit)   [exata - validado no 14]
  tubos:    rolos = ROUNDUP(metros x 1,07 / tamanho_rolo)  [validado no 10/14]
Saida: LEVANTAMENTO_<obra>.xlsx (LISTA-MATERIAIS + RASTREIO-KITS).
"""
import sys, os, glob, re, math, unicodedata
from collections import defaultdict
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
COMPLETAS = ["20241385", "20241390", "20251670"]
ALVOS = ["20251430", "20251533"]
ROLO_RE = re.compile(r"(?:ROLO|X)\s*(\d+)\s*M\b", re.I)


def norm_kit(nome):
    s = unicodedata.normalize("NFKD", nome.upper()).encode("ascii", "ignore").decode()
    s = re.sub(r"F\.\d+", " ", s)
    s = re.sub(r"FINA(L|IS)?\s*[A-Z]?[\d/,\sE\-]*", " ", s)
    s = re.sub(r"\d+[ºO]\s*(AO\s*\d+[ºO])?\s*(PAVIMENTO)?", " ", s)
    s = s.replace("AGUA FRIA", "AF").replace("AGUA QUENTE", "AQ")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_item(desc):
    return re.sub(r"\s+", " ", desc.replace("\n", " ")).strip().upper()


def linha_contagem(ws):
    best, br = -1, 3
    for r in (2, 3, 4, 5):
        n = sum(1 for c in range(8, ws.max_column + 1) if isinstance(ws.cell(r, c).value, (int, float)))
        if n > best:
            best, br = n, r
    return br


def abas_kit(wb):
    for ws in wb.worksheets:
        t = ws.title.upper()
        if t.startswith("RAMAL") or t.startswith("KIT") or t.startswith("CHICOTE"):
            yield ws


def colunas_kit(ws):
    cr = linha_contagem(ws)
    for c in range(8, ws.max_column + 1):
        cnt = ws.cell(cr, c).value
        h = " ".join(str(ws.cell(r, c).value or "").replace("\n", " ").strip()
                     for r in range(1, cr) if ws.cell(r, c).value).strip()
        if h or isinstance(cnt, (int, float)):
            yield c, h, (float(cnt) if isinstance(cnt, (int, float)) else None), cr


def extrai_biblioteca():
    """{kit_norm: [(obra, kit_orig, {item_norm: (desc_orig, qtd_por_kit)})]}"""
    bib = defaultdict(list)
    for obra in COMPLETAS:
        x = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
             if "PRE-" not in os.path.basename(a).upper() and not os.path.basename(a).startswith("~$")][0]
        wb = openpyxl.load_workbook(x, data_only=True)
        for ws in abas_kit(wb):
            cols = list(colunas_kit(ws))
            if not cols:
                continue
            cr = cols[0][3]
            for c, h, cnt, _ in cols:
                if not h:
                    continue
                rec = {}
                for r in range(cr + 1, ws.max_row + 1):
                    e = ws.cell(r, 5).value
                    v = ws.cell(r, c).value
                    if isinstance(e, str) and isinstance(v, (int, float)) and v:
                        un = str(ws.cell(r, 6).value or "UN").strip().upper()
                        rec[norm_item(e)] = (e.strip(), un, float(v))
                nk = norm_kit(h)
                if rec and len(nk) >= 4:
                    bib[nk].append((obra, h, rec))
        wb.close()
    return bib


def melhor_match(nk, bib):
    if len(nk) < 4:
        return None, "nome vazio"
    if nk in bib:
        return nk, "exato"
    toks = set(nk.split())
    best, bs = None, 0.0
    for k in bib:
        s = len(toks & set(k.split())) / max(len(toks | set(k.split())), 1)
        if s > bs:
            best, bs = k, s
    return (best, f"parcial {bs:.0%}") if bs >= 0.5 else (None, "sem match")


def gera(obra, bib):
    x = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
         if "PRE-" not in os.path.basename(a).upper() and "LEVANTAMENTO" not in os.path.basename(a).upper()
         and not os.path.basename(a).startswith("~$")][0]
    wb = openpyxl.load_workbook(x, data_only=True)
    itens = defaultdict(float)   # item_norm -> qtd acumulada
    descr = {}                   # item_norm -> (descricao, unidade)
    fonte_item = defaultdict(set)  # item_norm -> {"propria" | "emprestada-kit" | "emprestada-ramal"}
    rastreio, cob, dsc = [], 0, 0
    for ws in abas_kit(wb):
        for c, h, cnt, cr in colunas_kit(ws):
            if not cnt or not h:
                continue
            # 1) receita PROPRIA da obra nesta coluna (validado: e o que o Hederson preencheu)
            rec_own = {}
            for r in range(cr + 1, ws.max_row + 1):
                e, v = ws.cell(r, 5).value, ws.cell(r, c).value
                if isinstance(e, str) and isinstance(v, (int, float)) and v:
                    un = str(ws.cell(r, 6).value or "UN").strip().upper()
                    rec_own[norm_item(e)] = (e.strip(), un, float(v))
            if rec_own:
                for ik, (d, un, q) in rec_own.items():
                    itens[ik] += q * cnt
                    descr.setdefault(ik, (d, un))
                    fonte_item[ik].add("propria")
                cob += cnt
                rastreio.append((ws.title, h, cnt, "PROPRIA", "receita da propria obra"))
                continue
            # 2) fallback: empresta da biblioteca
            nk = norm_kit(h)
            mk, tipo = melhor_match(nk, bib)
            if mk is None:
                dsc += cnt
                rastreio.append((ws.title, h, cnt, "-", "SEM RECEITA"))
                continue
            eh_ramal = "RAMAL" in nk or ws.title.upper().startswith("RAMAL")
            fontes = bib[mk]
            acc = defaultdict(list)
            for _, _, rec in fontes:
                for ik, (d, un, q) in rec.items():
                    acc[ik].append(q)
                    descr.setdefault(ik, (d, un))
                    fonte_item[ik].add("emprestada-ramal" if eh_ramal else "emprestada-kit")
            for ik, qs in acc.items():
                itens[ik] += (sum(qs) / len(qs)) * cnt
            cob += cnt
            rastreio.append((ws.title, h, cnt, tipo, "; ".join(f"{o}:{k[:26]}" for o, k, _ in fontes)))
    wb.close()

    def confianca(ik):
        f = fonte_item[ik]
        if "emprestada-ramal" in f:
            return "BAIXA - ramal e da geometria da obra; medir no desenho/validar"
        if f == {"propria"}:
            return "PROPRIA (receita real da obra)"
        return "ALTA - kit/chicote (validado: erro de um digito entre obras)"

    linhas = []
    for ik, q in sorted(itens.items(), key=lambda kv: descr[kv[0]][0]):
        d, un = descr[ik]
        m_rolo = ROLO_RE.findall(d)
        conf = confianca(ik)
        if un == "RL" and m_rolo:
            rolo = int(m_rolo[-1])
            compra = math.ceil(q * 1.07 / rolo)
            linhas.append((d, "rolo", round(q, 1), f"metros x1,07 / rolo {rolo}m", compra, conf))
        elif un == "BR":
            compra = math.ceil(q * 1.07 / 5.8)
            linhas.append((d, "barra", round(q, 1), "metros x1,07 / barra 5,80m", compra, conf))
        else:
            linhas.append((d, "un", round(q, 1), "soma exata", math.ceil(q - 1e-9), conf))

    out = openpyxl.Workbook()
    ws1 = out.active
    ws1.title = "LISTA-MATERIAIS"
    ws1.append(["Item (descricao da planilha SPHE)", "Unid", "Qtde calculada (metros p/ tubo)",
                "Regra", "COMPRA sugerida", "Confianca",
                "Receita propria da obra tem prioridade; emprestada so no que falta"])
    for l in linhas:
        ws1.append(list(l))
    ws2 = out.create_sheet("RASTREIO-KITS")
    ws2.append(["aba", "kit (header)", "contagem", "match", "receita emprestada de"])
    for r in rastreio:
        ws2.append(list(r))
    dest = os.path.join(BASE, obra, f"LEVANTAMENTO_{obra}.xlsx")
    out.save(dest)

    tubos = [l for l in linhas if l[1] == "rolo"]
    print("=" * 88)
    print(f"LEVANTAMENTO {obra} | base: {os.path.basename(x)}")
    print(f"  cobertura de contagens: {cob:.0f} cobertas | {dsc:.0f} sem receita ({cob/max(cob+dsc,1):.0%})")
    print(f"  itens na lista: {len(linhas)} ({len(tubos)} tubos, {len(linhas)-len(tubos)} conexoes/pecas)")
    for d, _, q, regra, cp, conf in tubos:
        print(f"    TUBO  {d[:48]:<48} {q:>8.0f} m -> {cp:>3} rolos | {conf.split(' -')[0]}")
    print(f"  salvo: {dest}")


def main():
    bib = extrai_biblioteca()
    n_itens = len({ik for fs in bib.values() for _, _, rec in fs for ik in rec})
    print(f"BIBLIOTECA COMPLETA: {len(bib)} kits | {n_itens} itens distintos (tubos + conexoes)")
    for obra in ALVOS:
        gera(obra, bib)


if __name__ == "__main__":
    main()
