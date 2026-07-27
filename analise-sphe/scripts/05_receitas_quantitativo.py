# -*- coding: utf-8 -*-
"""Extrai metros de PEX por DN (gabarito) de cada quantitativo, via receita.

Modelo (decifrado no Joao Dias, confirmado PAMARIS):
  aba RAMAL/KITS: L1 cabecalho, L2 nomes dos kits/finais (cols H+),
  L3 contagem por kit, demais linhas = pecas. Coluna E=descricao, F=unid,
  G=Qtde Total (em ROLOS p/ tubo). Cols H+ = metros da peca por kit.
  metros_por_DN = soma_c( cell(linha,c) * contagem(c) ) sobre cols de kit.
"""
import sys, glob, os, re
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

DN_RE = re.compile(r"PEX\s*(\d{2})", re.I)
ROLO_RE = re.compile(r"X\s*(\d+)\s*M\b", re.I)
FIRST_KIT_COL = 8  # coluna H

def num(v):
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        v = v.strip().replace(".", "").replace(",", ".")
        try: return float(v)
        except: return 0.0
    return 0.0

def detecta_linha_contagem(ws):
    """A linha de contagem e a que tem mais celulas numericas nas cols de kit."""
    melhor, best_r = -1, 3
    for r in (2, 3, 4, 5):
        n = sum(1 for c in range(FIRST_KIT_COL, ws.max_column+1)
                if isinstance(ws.cell(r, c).value, (int, float)))
        if n > melhor:
            melhor, best_r = n, r
    return best_r

def processar_aba(ws):
    crow = detecta_linha_contagem(ws)
    counts = {}
    for c in range(FIRST_KIT_COL, ws.max_column+1):
        counts[c] = num(ws.cell(crow, c).value)
    nomes = {c: ws.cell(crow-1, c).value for c in range(FIRST_KIT_COL, ws.max_column+1)}
    metros_dn = {}
    rolos_dn = {}
    for r in range(crow+1, ws.max_row+1):
        desc = ws.cell(r, 5).value
        unid = ws.cell(r, 6).value
        if not isinstance(desc, str) or "TUBO PEX" not in desc.upper():
            continue
        if "BARRA" in desc.upper() or "ROLO" in desc.upper():
            continue  # ignora variacoes barra/rolo alternativas (mantem SERIE 5)
        m = DN_RE.search(desc)
        if not m: continue
        dn = int(m.group(1))
        rolo = ROLO_RE.search(desc)
        tam = int(rolo.group(1)) if rolo else None
        metros = sum(num(ws.cell(r, c).value) * counts.get(c, 0)
                     for c in range(FIRST_KIT_COL, ws.max_column+1))
        rolos = num(ws.cell(r, 7).value)
        if metros > 0 or rolos > 0:
            metros_dn[dn] = metros_dn.get(dn, 0) + metros
            rolos_dn[dn] = rolos_dn.get(dn, 0) + rolos * (tam or 0)
    return metros_dn, rolos_dn, counts, nomes

def main():
    obras = sorted(d for d in os.listdir(BASE)
                   if os.path.isdir(os.path.join(BASE, d)) and not d.startswith("_"))
    for o in obras:
        xs = glob.glob(os.path.join(BASE, o, "*.xlsx"))
        if not xs: continue
        print("=" * 80)
        print("OBRA", o, "|", os.path.basename(xs[0]))
        wb = openpyxl.load_workbook(xs[0], data_only=True)
        tot_m = {}
        for ws in wb.worksheets:
            t = ws.title.upper()
            if not (t.startswith("RAMAL") or t.startswith("KIT") or t.startswith("CHICOTE")):
                continue
            md, rd, counts, nomes = processar_aba(ws)
            if not md: continue
            print(f"  [{ws.title}]  metros por DN (receita x contagem) | (rolos x tam = compra):")
            for dn in sorted(md):
                print(f"     DN {dn:>3}: {md[dn]:10.1f} m   | compra ~{rd.get(dn,0):10.0f} m")
                tot_m[dn] = tot_m.get(dn, 0) + md[dn]
            # mostra os kits ativos (contagem>0)
            ativos = [(nomes[c], int(counts[c])) for c in counts if counts[c] > 0 and nomes.get(c)]
            print("     kits ativos:", ", ".join(f"{n}={q}" for n, q in ativos[:12]))
        print("  >>> TOTAL OBRA por DN (m, receita):", {k: round(v) for k, v in sorted(tot_m.items())})
        wb.close()

if __name__ == "__main__":
    main()
