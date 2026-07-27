# -*- coding: utf-8 -*-
"""Reconcilia a obra-ancora LIVING (20241385 / 7409) em 3 pontas:
  A) PREDIO (matriz pavimentos x finais) -> aptos por final
  B) contagens das abas RAMAL/KITS -> quais derivam da matriz e quais nao (residual tacito)
  C) DXF PVTIPO -> blocos de prumada/tipologia que o desenho fornece por pavimento
Objetivo: provar que contagem = f(PREDIO, desenho) e isolar o que sobra de "cabeca do Hederson".
"""
import sys, os, glob, re
from collections import Counter, defaultdict
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OBRA = "20241385"
DXF = os.path.join(BASE, "_analise", "dxf", OBRA, "7409-HID-PE-0007-PVTIPO-R01.dxf")


def parte_a_predio(wb):
    ws = wb["PRÉDIO"]
    finais = {ws.cell(6, c).value: c for c in range(5, 13)}  # E6..L6 = 1..8
    aptos = {}
    for f, c in finais.items():
        n = 0
        for r in range(8, 37):  # D8..D36 (28..T)
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "X":
                n += 1
        aptos[int(f)] = n
    print("A) PRÉDIO - aptos por final (contagem de X):")
    print("   " + " | ".join(f"final {f}={n}" for f, n in sorted(aptos.items())))
    print(f"   total aptos = {sum(aptos.values())}")
    return aptos


def contagens_aba(ws):
    """(nome_kit, contagem) por coluna H+; nome = header L1+L2 concatenado."""
    out = []
    for c in range(8, ws.max_column + 1):
        h1 = ws.cell(1, c).value or ""
        h2 = ws.cell(2, c).value or ""
        cnt = ws.cell(3, c).value
        nome = " / ".join(x.replace("\n", " ").strip() for x in (str(h1), str(h2)) if x.strip())
        if nome or isinstance(cnt, (int, float)):
            out.append((nome, cnt if isinstance(cnt, (int, float)) else None))
    return out


FINAIS_RE = re.compile(r"FINAIS?\s+([\d/,\sE]+)", re.I)


def finais_do_nome(nome):
    m = FINAIS_RE.search(nome)
    if not m:
        return None
    return sorted({int(t) for t in re.findall(r"\d", m.group(1))})


def resumo_grupos(wb):
    """Le a aba RESUMO: [(grupo_pav, repeticoes, finais, aptos, ambientes{col:qtd})]."""
    ws = wb["RESUMO"]
    ambs = {c: (ws.cell(2, c).value or "").strip() for c in range(5, 11)}  # E..J
    grupos = []
    for r in range(3, ws.max_row + 1):
        pav = ws.cell(r, 1).value
        if not isinstance(pav, str) or not pav.strip():
            continue
        rep = ws.cell(r, 2).value or 0
        fins = finais_do_nome(str(ws.cell(r, 3).value or "")) or []
        apt = ws.cell(r, 4).value or 0
        amb = {ambs[c]: ws.cell(r, c).value for c in ambs if isinstance(ws.cell(r, c).value, (int, float))}
        grupos.append((pav.strip(), rep, fins, apt, amb))
    return grupos


def parte_b_contagens(wb, aptos):
    grupos = resumo_grupos(wb)
    print("\nB1) RESUMO - grupos pavimento x finais x ambientes:")
    tot = defaultdict(float)
    for pav, rep, fins, apt, amb in grupos:
        tot["APTOS"] += rep * apt
        for a, q in amb.items():
            tot[a] += rep * apt * q
        print(f"   {pav:20} rep={rep:2} finais={fins} aptos={apt} amb={amb}")
    print("   TOTAIS derivados do RESUMO: " + " | ".join(f"{k}={v:.0f}" for k, v in tot.items()))

    print("\nB2) Contagens das abas x derivação (RESUMO como fonte):")
    print(f"   {'aba':6} {'kit (header)':58} {'cont':>5} {'deriv':>5} {'diff':>5}")
    residuais = []
    # derivacao por coluna: RAMAL usa rep*aptos do grupo de pavimento+finais;
    # o Hederson DOBRA terreo e 20o regular na coluna do TIPO de mesma receita.
    derivadores = {
        # RAMAL: (finais do header, range do header) -> soma rep*apt dos grupos compatíveis
        "RAMAL": None,  # tratado abaixo via grupos
        # KITS: nome do ambiente no RESUMO
        "CHUVEIRO": tot.get("BANHO TIPO", 0) + tot.get("BANHO MASTER", 0),
        "CHICOTE BANHO": tot.get("BANHO TIPO", 0) + tot.get("BANHO MASTER", 0),
        "CHICOTE LAVABO": tot.get("LAVABO", 0),
        "CHICOTE COZINHA": None,  # so finais 1/2/7/8 -> via matriz
        "TRAVESSA": tot.get("APTOS", 0),
    }
    for aba in ("RAMAL", "KITS"):
        ws = wb[aba]
        for nome, cnt in contagens_aba(ws):
            if cnt is None:
                continue
            up = nome.upper()
            fins = finais_do_nome(nome)
            deriv = None
            if aba == "RAMAL" and fins:
                # soma rep*apt dos grupos do RESUMO cujos finais estao contidos no header
                deriv = sum(rep * apt for pav, rep, gf, apt, _ in grupos
                            if gf and set(gf) <= set(fins))
            elif "MANIFOLD" in up and fins:
                deriv = sum(aptos.get(f, 0) for f in fins)
            elif ("LAVABO" in up or "COZINHA" in up) and fins:
                deriv = sum(rep * apt for pav, rep, gf, apt, amb in grupos
                            if gf and set(gf) <= set(fins) and
                            ("LAVABO" not in up or amb.get("LAVABO")))
            else:
                for chave, val in derivadores.items():
                    if chave != "RAMAL" and val is not None and up.startswith(chave):
                        deriv = int(val)
                        break
            ds = f"{deriv:.0f}" if deriv is not None else "-"
            diff = f"{int(cnt) - deriv:+.0f}" if deriv is not None else "?"
            print(f"   {aba:6} {nome[:58]:58} {int(cnt):5d} {ds:>5} {diff:>5}")
            if deriv is None or int(cnt) != deriv:
                residuais.append((aba, nome, int(cnt), deriv))
    return residuais


def parte_c_dxf():
    import ezdxf
    doc = ezdxf.readfile(DXF)
    msp = doc.modelspace()
    counts = Counter()

    def walk(block_name, mult, depth):
        if depth > 6 or block_name not in doc.blocks:
            return
        for e in doc.blocks[block_name]:
            if e.dxftype() == "INSERT":
                counts[e.dxf.name] += mult
                walk(e.dxf.name, mult, depth + 1)

    for ins in msp.query("INSERT"):
        counts[ins.dxf.name] += 1
        walk(ins.dxf.name, 1, 1)

    prum = Counter()
    arq = Counter()
    for nome_full, q in counts.items():
        nome = nome_full.split("$")[-1].upper()
        if "PRUM" in nome or "SATT" in nome:
            prum[nome] += q
        elif "APTO" in nome or "-TIPO" in nome:
            arq[nome] += q
    print("\nC) DXF PVTIPO - o que o desenho fornece:")
    print("   blocos de prumada/insercao:")
    for n, q in prum.most_common():
        print(f"     {n:40} {q:4d}")
    print("   blocos de arquitetura (tipologias):")
    for n, q in arq.most_common(8):
        print(f"     {n:40} {q:4d}")
    return prum


def main():
    x = glob.glob(os.path.join(BASE, OBRA, "*.xlsx"))[0]
    wb = openpyxl.load_workbook(x, data_only=True)
    print("OBRA", OBRA, "|", os.path.basename(x))
    print("=" * 78)
    aptos = parte_a_predio(wb)
    residuais = parte_b_contagens(wb, aptos)
    parte_c_dxf()
    print("\nRESIDUAL (nao derivou direto da matriz = candidato a tacito):")
    for aba, nome, cnt, deriv in residuais:
        print(f"   [{aba}] {nome[:60]} cont={cnt} deriv={deriv}")
    if not residuais:
        print("   nenhum - contagens 100% derivadas do PRÉDIO")


if __name__ == "__main__":
    main()
