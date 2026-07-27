# -*- coding: utf-8 -*-
"""Gera PRE-QUANTITATIVO para Brooklyn (20251430) e Peak (20251533) — obras cujo
xlsx tem contagens completas mas ZERO receita/compra (levantamento nunca feito).
Empresta receitas [kit x DN] das 3 obras completas (Living/Edition/Pamaris) por
match de nome normalizado e calcula:
  compra_rolos[DN] = ROUNDUP( Σ(receita_media[kit,DN] x contagem[kit]) / rolo x 1,07 )
Saida: console + xlsx PRE-QUANTITATIVO_<obra>.xlsx com aba de rastreio (qual receita
veio de onde, o que ficou descoberto). NAO substitui levantamento: e ponto de partida
p/ validacao do Hederson.
"""
import sys, os, glob, re, math, unicodedata
from collections import defaultdict
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
COMPLETAS = ["20241385", "20241390", "20251670"]
ALVOS = ["20251430", "20251533"]
ROLO = {16: 200, 20: 100, 25: 100, 32: 50}
DN_RE = re.compile(r"PEX\s*(\d{2})", re.I)


def norm(nome):
    s = unicodedata.normalize("NFKD", nome.upper()).encode("ascii", "ignore").decode()
    s = re.sub(r"F\.\d+", " ", s)                       # prefixo de folha (Peak)
    s = re.sub(r"FINA(L|IS)?\s*[A-Z]?[\d/,\sE\-]*", " ", s)  # grupos de finais
    s = re.sub(r"\d+[ºO]\s*(AO\s*\d+[ºO])?\s*(PAVIMENTO)?", " ", s)
    s = s.replace("AGUA FRIA", "AF").replace("AGUA QUENTE", "AQ")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def linha_contagem(ws):
    best, br = -1, 3
    for r in (2, 3, 4, 5):
        n = sum(1 for c in range(8, ws.max_column + 1) if isinstance(ws.cell(r, c).value, (int, float)))
        if n > best:
            best, br = n, r
    return br


def abas_kit(wb):
    for ws in wb.worksheets:
        t = ws.title.upper()
        if t.startswith("RAMAL") or t.startswith("KIT") or t.startswith("CHICOTE"):
            yield ws


def colunas_kit(ws):
    cr = linha_contagem(ws)
    for c in range(8, ws.max_column + 1):
        cnt = ws.cell(cr, c).value
        h = " ".join(str(ws.cell(r, c).value or "").replace("\n", " ").strip()
                     for r in range(1, cr) if ws.cell(r, c).value).strip()
        if h or isinstance(cnt, (int, float)):
            yield c, h, (float(cnt) if isinstance(cnt, (int, float)) else None), cr


def extrai_receitas():
    """{nome_norm: [(obra, kit_original, {dn: metros})]}"""
    bib = defaultdict(list)
    for obra in COMPLETAS:
        x = glob.glob(os.path.join(BASE, obra, "*.xlsx"))[0]
        wb = openpyxl.load_workbook(x, data_only=True)
        for ws in abas_kit(wb):
            cols = list(colunas_kit(ws))
            if not cols:
                continue
            cr = cols[0][3]
            for c, h, cnt, _ in cols:
                if not h:
                    continue
                rec = defaultdict(float)
                for r in range(cr + 1, ws.max_row + 1):
                    e = ws.cell(r, 5).value
                    if not isinstance(e, str) or "TUBO PEX" not in e.upper() or "BARRA" in e.upper():
                        continue
                    m = DN_RE.search(e)
                    v = ws.cell(r, c).value
                    if m and isinstance(v, (int, float)) and v:
                        rec[int(m.group(1))] += float(v)
                nn = norm(h)
                if rec and len(nn) >= 4:
                    bib[nn].append((obra, h, dict(rec)))
        wb.close()
    return bib


def melhor_match(nome_norm, bib):
    if len(nome_norm) < 4:
        return None, "nome vazio pos-normalizacao"
    if nome_norm in bib:
        return nome_norm, "exato"
    toks = set(nome_norm.split())
    best, bs = None, 0.0
    for k in bib:
        kt = set(k.split())
        s = len(toks & kt) / max(len(toks | kt), 1)
        if s > bs:
            best, bs = k, s
    return (best, f"parcial {bs:.0%}") if bs >= 0.5 else (None, "sem match")


def gera(obra, bib):
    x = glob.glob(os.path.join(BASE, obra, "*.xlsx"))[0]
    wb = openpyxl.load_workbook(x, data_only=True)
    print("=" * 90)
    print(f"PRE-QUANTITATIVO {obra} | {os.path.basename(x)}")
    metros = defaultdict(float)
    rastreio = []
    cobertos = descob = 0
    for ws in abas_kit(wb):
        for c, h, cnt, cr in colunas_kit(ws):
            if cnt is None or cnt == 0 or not h:
                continue
            nn = norm(h)
            mk, tipo = melhor_match(nn, bib)
            if mk is None:
                descob += cnt
                rastreio.append((ws.title, h, cnt, "-", "-", "SEM RECEITA", {}))
                continue
            fontes = bib[mk]
            med = defaultdict(list)
            for _, _, rec in fontes:
                for dn, m in rec.items():
                    med[dn].append(m)
            rec_med = {dn: sum(v) / len(v) for dn, v in med.items()}
            for dn, m in rec_med.items():
                metros[dn] += m * cnt
            cobertos += cnt
            origem = "; ".join(f"{o}:{k[:28]}" for o, k, _ in fontes)
            rastreio.append((ws.title, h, cnt, tipo, origem, "OK", rec_med))
    wb.close()

    print(f"  contagens cobertas por receita emprestada: {cobertos:.0f} | descobertas: {descob:.0f} "
          f"({cobertos / max(cobertos + descob, 1):.0%} de cobertura)")
    print(f"  {'DN':>4} {'metros':>10} {'x1,07':>10} {'rolo':>5} {'COMPRA (rolos)':>14}")
    compra = {}
    for dn in sorted(metros):
        rolo = ROLO.get(dn, 100)
        rolos = math.ceil(metros[dn] * 1.07 / rolo)
        compra[dn] = rolos
        print(f"  {dn:>4} {metros[dn]:10.0f} {metros[dn]*1.07:10.0f} {rolo:5d} {rolos:14d}")

    out = openpyxl.Workbook()
    ws1 = out.active
    ws1.title = "PRE-QUANTITATIVO"
    ws1.append(["DN", "metros (receita x contagem)", "metros x1,07", "rolo (m)", "COMPRA (rolos)",
                "AVISO: receitas EMPRESTADAS de outras obras SPHE - validar com Hederson"])
    for dn in sorted(metros):
        ws1.append([dn, round(metros[dn], 1), round(metros[dn] * 1.07, 1), ROLO.get(dn, 100), compra[dn]])
    ws2 = out.create_sheet("RASTREIO")
    ws2.append(["aba", "kit (header)", "contagem", "match", "receita emprestada de", "status", "receita DN:m"])
    for row in rastreio:
        ws2.append(list(row[:6]) + ["; ".join(f"DN{d}={m:.1f}" for d, m in sorted(row[6].items()))])
    dest = os.path.join(BASE, obra, f"PRE-QUANTITATIVO_{obra}.xlsx")
    out.save(dest)
    print(f"  salvo: {dest}")


def main():
    bib = extrai_receitas()
    print(f"BIBLIOTECA: {len(bib)} kits com receita nas obras completas")
    for obra in ALVOS:
        gera(obra, bib)


if __name__ == "__main__":
    main()
