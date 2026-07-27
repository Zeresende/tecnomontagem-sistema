# -*- coding: utf-8 -*-
"""VALIDACAO LEAVE-ONE-OUT: Brooklyn (20251430) tem receitas REAIS de tubo nos
ramais (descoberta de 03/07 - nao estava vazio). Compara, kit a kit, a receita
emprestada da biblioteca (Living/Edition/Pamaris) contra a receita real do
Hederson no Brooklyn. Metrica: metros por DN por kit + erro %.
E o numero de confianca do levantamento automatico sem depender do Hederson."""
import sys, os, glob, re
from collections import defaultdict
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
lc = __import__("15_levantamento_completo")

DN_TUBO_RE = re.compile(r"(?:PEX|PERT)\s*(\d{2})", re.I)
ALVO = "20251430"


def receita_dn(rec):
    """Colapsa receita item-level p/ {dn: metros} usando so linhas de tubo (RL)."""
    out = defaultdict(float)
    for ik, (d, un, q) in rec.items():
        m = DN_TUBO_RE.search(d)
        if un == "RL" and m and d.upper().startswith("TUBO"):
            out[int(m.group(1))] += q
    return dict(out)


def main():
    bib = lc.extrai_biblioteca()
    x = [a for a in glob.glob(os.path.join(BASE, ALVO, "*.xlsx"))
         if "PRE-" not in a.upper() and "LEVANTAMENTO" not in a.upper()
         and not os.path.basename(a).startswith("~$")][0]
    wb = openpyxl.load_workbook(x, data_only=True)
    print(f"VALIDACAO LEAVE-ONE-OUT | alvo: {os.path.basename(x)}")
    print(f"biblioteca emprestada: Living + Edition + Pamaris ({len(bib)} kits)\n")
    tot_real, tot_prev = defaultdict(float), defaultdict(float)
    kits_ok = kits_sem = 0
    for ws in lc.abas_kit(wb):
        cols = list(lc.colunas_kit(ws))
        if not cols:
            continue
        cr = cols[0][3]
        for c, h, cnt, _ in cols:
            if not h or not cnt:
                continue
            # receita REAL do Brooklyn nessa coluna (so linhas de tubo)
            real = defaultdict(float)
            for r in range(cr + 1, ws.max_row + 1):
                e, v = ws.cell(r, 5).value, ws.cell(r, c).value
                if not (isinstance(e, str) and isinstance(v, (int, float)) and v):
                    continue
                un = str(ws.cell(r, 6).value or "").strip().upper()
                m = DN_TUBO_RE.search(e)
                if un == "RL" and m and e.upper().startswith("TUBO"):
                    real[int(m.group(1))] += float(v)
            if not real:
                continue  # kit sem receita real -> nao da p/ comparar
            nk = lc.norm_kit(h)
            mk, tipo = lc.melhor_match(nk, bib)
            if mk is None:
                kits_sem += 1
                print(f"[{ws.title}] {h[:50]:<50} cont={cnt:.0f}  REAL={fmt(real)}  PREVISTO=sem match")
                continue
            med = defaultdict(list)
            for _, _, rec in bib[mk]:
                for dn, q in receita_dn(rec).items():
                    med[dn].append(q)
            prev = {dn: sum(v) / len(v) for dn, v in med.items()}
            kits_ok += 1
            for dn, q in real.items():
                tot_real[dn] += q * cnt
            for dn, q in prev.items():
                tot_prev[dn] += q * cnt
            print(f"[{ws.title}] {h[:50]:<50} cont={cnt:.0f} ({tipo})")
            for dn in sorted(set(real) | set(prev)):
                rr, pp = real.get(dn, 0), prev.get(dn, 0)
                err = (pp - rr) / rr * 100 if rr else float("inf")
                print(f"      DN{dn}: real {rr:7.2f} m/kit | emprestada {pp:7.2f} m/kit | erro {err:+6.1f}%")
    wb.close()
    print("\n" + "=" * 70)
    print(f"TOTAIS (so kits comparaveis: {kits_ok} com match, {kits_sem} sem match)")
    for dn in sorted(set(tot_real) | set(tot_prev)):
        rr, pp = tot_real.get(dn, 0), tot_prev.get(dn, 0)
        err = (pp - rr) / rr * 100 if rr else float("inf")
        print(f"  DN{dn}: real {rr:9.0f} m | previsto {pp:9.0f} m | erro {err:+6.1f}%")


def fmt(d):
    return "; ".join(f"DN{k}={v:g}" for k, v in sorted(d.items()))


if __name__ == "__main__":
    main()
