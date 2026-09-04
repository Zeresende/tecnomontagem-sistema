# -*- coding: utf-8 -*-
"""Compara um resultado com o gabarito da rodada holdout e aplica os 4 criterios cravados.

Criterios (fixados ANTES da rodada, 04/09/2026 - nao mexer depois de ver o resultado):
  C1 contagem de kits           : exata (so avalia se vier --kits)
  C2 conexao por PECA_ID        : total da obra dentro de 5% do gabarito, para TODA peca
                                  do gabarito; peca a mais no resultado tambem reprova
  C3 tubo por bitola            : rolos dentro de +-1 do que o Hederson comprou
  C4 pecas fora do catalogo     : zero

Resultado aceito em 3 formatos:
  --resultado arquivo.xlsx   levantamento no contrato de 12/08 (PECA_ID + QTD_TOTAL)
  --resultado arquivo.csv    colunas PECA_ID;QTD_TOTAL (ou , como separador)
  --predicao                 braco A: predicao da biblioteca-mae (saida do 61), por papel
Tubo no resultado pode vir em metros (padrao, unidade do catalogo) ou em rolos (--tubo-em rolos).
  --receita receita_lida.csv  COLUNA;PECA_ID;QTD_POR_KIT (o que o conector leu de cada kit na DTIP):
                             decompoe a diferenca por celula e aplica a regra de 04/09 - diferenca
                             que cai inteira numa celula PENDENTE (tubos do Marcelo) e "explicada",
                             e o C3 passa a ser avaliado sem esses metros (C3 ajustado).

Uso:
  python 63_comparar_holdout.py --obra 20251670 --resultado kitflow_pamaris.xlsx [--kits kits.csv] [--receita receita_lida.csv]
  python 63_comparar_holdout.py --obra 20251670 --predicao
"""
import argparse
import csv
import importlib
import io
import math
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
AQUI = Path(__file__).resolve().parent
SISTEMA = AQUI.parent.parent / "sistema"
sys.path.insert(0, str(SISTEMA))
m03 = importlib.import_module("03_gerar_planilhas")
sys.path.insert(0, str(AQUI))
import holdout_receita_lida as hrl

OBRAS = {"20241385": "Living", "20241390": "Edition", "20251430": "Brooklyn",
         "20251533": "Peak", "20251670": "Pamaris"}
TOL_CONEXAO = 0.05
TOL_ROLOS = 1
FOLGA = 1.07


def sem_acento(s):
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()


def norm(s):
    return re.sub("[ ]+", " ", sem_acento(s).upper()).strip()


def ler_gabarito(pasta, apelido):
    arq = pasta / f"gabarito_{apelido.lower()}_por_peca.csv"
    with io.open(arq, encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.DictReader(fh, delimiter=";"))
    gab = {}
    for r in rows:
        gab[int(r["PECA_ID"])] = {
            "desc": r["DESCRICAO"], "qtd": float(r["QTD_TOTAL"]), "tipo": r["TIPO"],
            "rolos": int(r["ROLOS_HEDERSON"]) if r["ROLOS_HEDERSON"] else None,
            "rolo": int(r["TAMANHO_ROLO_M"]) if r["TAMANHO_ROLO_M"] else None,
        }
    return gab


def ler_resultado(caminho):
    caminho = Path(caminho)
    if caminho.suffix.lower() == ".xlsx":
        _, lev, avisos = m03.carregar_levantamento(caminho)
        return {int(k): float(v) for k, v in lev.items()}, avisos
    with io.open(caminho, encoding="utf-8-sig", newline="") as fh:
        texto = fh.read()
    sep = ";" if texto.count(";") >= texto.count(",") else ","
    res, avisos = {}, []
    for r in csv.DictReader(io.StringIO(texto), delimiter=sep):
        chaves = {k.upper(): k for k in r}
        try:
            pid = int(float(r[chaves["PECA_ID"]]))
            qtd = float(str(r[chaves["QTD_TOTAL"]]).replace(",", "."))
        except (KeyError, ValueError):
            avisos.append(f"linha ignorada: {r}")
            continue
        res[pid] = res.get(pid, 0.0) + qtd
    return res, avisos


def ler_predicao(pasta, apelido, catalogo, gab):
    """Braco A: predicao por papel (61) -> PECA_ID via catalogo. Tubo em metros."""
    arq = pasta / f"predicao_{apelido.lower()}.csv"
    with io.open(arq, encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.DictReader(fh, delimiter=";"))
    por_papel = defaultdict(float)
    for r in rows:
        por_papel[r["papel"]] += float(r["total"])
    desc_idx = {norm(p["descricao"]): pid for pid, p in catalogo.items() if p["sistema"] == "PEX"}
    tubo_gab = {}
    for pid, g in gab.items():
        m = re.search(r"(?:PERT|PEX)[ ]*(16|20|25|32)", g["desc"].upper())
        if g["tipo"] == "TUBO" and m:
            tubo_gab[m.group(1)] = pid
    res, sem_mapa = {}, []
    for papel, total in por_papel.items():
        if papel.startswith("TUBO O"):
            dn = papel[6:]
            pid = tubo_gab.get(dn)
            if pid is None:     # bitola que a obra nao usa: pega o tubo Serie 5 dessa bitola
                pid = next((p for p, x in catalogo.items() if x["sistema"] == "PEX"
                            and norm(x["descricao"]).startswith(f"TUBO PEX {dn} - S")), None)
        else:
            pid = desc_idx.get(papel)
        if pid is None:
            sem_mapa.append(f"{papel} ({total:g})")
            continue
        res[pid] = res.get(pid, 0.0) + total
    return res, sem_mapa


def ler_kits(caminho):
    with io.open(caminho, encoding="utf-8-sig", newline="") as fh:
        texto = fh.read()
    sep = ";" if texto.count(";") >= texto.count(",") else ","
    out = {}
    for r in csv.DictReader(io.StringIO(texto), delimiter=sep):
        chaves = {k.upper(): k for k in r}
        out[norm(r[chaves["COLUNA"]])] = float(r[chaves["CONTAGEM"]])
    return out


def kits_gabarito(pasta, apelido):
    wb = load_workbook(pasta / f"gabarito_{apelido.lower()}.xlsx", read_only=True)
    ws = wb["POR_KIT"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            out[norm(row[1])] = float(row[2])
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--obra", required=True)
    ap.add_argument("--resultado", help="xlsx (contrato 12/08) ou csv PECA_ID;QTD_TOTAL")
    ap.add_argument("--predicao", action="store_true", help="braco A: biblioteca-mae do 61")
    ap.add_argument("--tubo-em", choices=("metros", "rolos"), default="metros")
    ap.add_argument("--kits", help="csv COLUNA;CONTAGEM para o criterio C1")
    ap.add_argument("--receita", help="csv COLUNA;PECA_ID;QTD_POR_KIT lido da DTIP (decomposicao + C3 ajustado)")
    a = ap.parse_args()
    apelido = OBRAS[a.obra]
    pasta = AQUI / "saida" / f"holdout_{apelido.lower()}"
    catalogo = m03.carregar_catalogo()
    gab = ler_gabarito(pasta, apelido)

    if a.predicao:
        res, avisos = ler_predicao(pasta, apelido, catalogo, gab)
        rotulo = "braco A - biblioteca-mae (mediana das outras obras x contagem real)"
        avisos = [f"papel sem PECA_ID no catalogo: {x}" for x in avisos]
    elif a.resultado:
        res, avisos = ler_resultado(a.resultado)
        rotulo = a.resultado
    else:
        ap.error("informe --resultado ou --predicao")

    print("=" * 96)
    print(f"HOLDOUT {apelido} ({a.obra}) - resultado: {rotulo}")
    print("=" * 96)
    for x in avisos:
        print("  aviso:", x)
    veredito = {}
    kr = None

    # C1 contagens
    if a.kits:
        kg, kr = kits_gabarito(pasta, apelido), ler_kits(a.kits)
        falhas = [(c, kg[c], kr.get(c)) for c in kg if kr.get(c) != kg[c]]
        veredito["C1 contagem de kits (exata)"] = not falhas
        print("\nC1 contagem de kits:")
        for c, g, r in falhas:
            print(f"  FALHA {c}: gabarito {g:g}, resultado {r}")
        if not falhas:
            print(f"  {len(kg)} colunas exatas")
    else:
        print("\nC1 contagem de kits: nao avaliado (sem --kits)")

    # C2 conexoes
    print(f"\nC2 conexao por PECA_ID (tolerancia {TOL_CONEXAO:.0%}):")
    print(f"  {'PECA_ID':>7} {'descricao':52} {'gabarito':>9} {'resultado':>9} {'desvio':>8}  ok")
    ok_c2 = True
    for pid, g in sorted(gab.items()):
        if g["tipo"] != "CONEXAO":
            continue
        r = res.get(pid, 0.0)
        desvio = (r - g["qtd"]) / g["qtd"] if g["qtd"] else 0.0
        passa = abs(desvio) <= TOL_CONEXAO
        ok_c2 &= passa
        print(f"  {pid:>7} {g['desc'][:52]:52} {g['qtd']:>9.0f} {r:>9.0f} {desvio:>+7.1%}  {'ok' if passa else 'FALHA'}")
    extras = [(pid, q) for pid, q in res.items() if pid not in gab and pid in catalogo
              and catalogo[pid]["sistema"] == "PEX" and "TUBO" not in str(catalogo[pid]["descricao"]).upper()]
    for pid, q in sorted(extras):
        ok_c2 = False
        print(f"  {pid:>7} {str(catalogo[pid]['descricao'])[:52]:52} {0:>9} {q:>9.0f} {'a mais':>8}  FALHA")
    veredito["C2 conexao por PECA_ID (5%)"] = ok_c2

    # C3 tubo
    print(f"\nC3 tubo por bitola (rolos, tolerancia +-{TOL_ROLOS}):")
    ok_c3 = True
    for pid, g in sorted(gab.items()):
        if g["tipo"] != "TUBO":
            continue
        r = res.get(pid, 0.0)
        rolos_r = int(round(r)) if a.tubo_em == "rolos" else math.ceil(r * FOLGA / g["rolo"])
        passa = abs(rolos_r - g["rolos"]) <= TOL_ROLOS
        ok_c3 &= passa
        print(f"  {pid:>7} {g['desc'][:44]:44} gabarito {g['qtd']:>7.0f} m = {g['rolos']:>3} rolos | "
              f"resultado {r:>7.0f} {a.tubo_em[:1]} = {rolos_r:>3} rolos  {'ok' if passa else 'FALHA'}")
    tubos_extras = [pid for pid in res if pid not in gab and pid in catalogo
                    and "TUBO" in str(catalogo[pid]["descricao"]).upper()]
    for pid in tubos_extras:
        ok_c3 = False
        print(f"  {pid:>7} {str(catalogo[pid]['descricao'])[:44]:44} a mais: {res[pid]:.0f}  FALHA")
    veredito[f"C3 tubo por bitola (+-{TOL_ROLOS} rolo)"] = ok_c3

    # decomposicao por coluna de kit + C3 ajustado (regra de 04/09: celulas pendentes do Marcelo)
    if a.receita:
        lida, avisos_r = hrl.ler_receita_lida(a.receita, norm)
        for x in avisos_r:
            print("  aviso:", x)
        celulas, contagem, pend = hrl.ler_gabarito_por_kit(pasta, apelido, norm)
        pend = hrl.resolver_pendentes_por_catalogo(pend, catalogo, norm)
        linhas, explicado = hrl.decompor(lida, celulas, contagem, pend, kr, catalogo, gab)
        hrl.conferir_totais(lida, contagem, kr, res, gab, a.tubo_em, FOLGA)
        ok_c3_aj = hrl.imprimir(linhas, explicado, gab, res, a.tubo_em, FOLGA, TOL_ROLOS)
        veredito.pop(f"C3 tubo por bitola (+-{TOL_ROLOS} rolo)")
        veredito[f"C3 tubo por bitola (+-{TOL_ROLOS} rolo) - bruto: {'passa' if ok_c3 else 'falha'}; "
                 f"AJUSTADO sem celulas pendentes"] = ok_c3_aj
        veredito["DECOMPOSICAO por celula (sem DIFERENCA / A MAIS / FALTA)"] = not any(
            st in ("DIFERENCA", "A MAIS", "FALTA") for *_, st in linhas)

    # C4 fora do catalogo
    fora = sorted(pid for pid in res if pid not in catalogo)
    veredito["C4 pecas fora do catalogo (zero)"] = not fora
    print(f"\nC4 pecas fora do catalogo: {len(fora)}" + (f" -> {fora}" if fora else ""))

    print("\n" + "-" * 96)
    for k, v in veredito.items():
        print(f"  {'PASSA' if v else 'FALHA':6} {k}")
    final = all(veredito.values())
    print(f"\nVEREDITO: {'PASSA' if final else 'FALHA'} "
          f"({sum(veredito.values())}/{len(veredito)} criterios avaliados)")
    return 0 if final else 1


if __name__ == "__main__":
    raise SystemExit(main())
