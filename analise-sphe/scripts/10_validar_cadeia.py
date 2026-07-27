# -*- coding: utf-8 -*-
"""Valida a cadeia deterministica da Tecnomontagem ponta a ponta nas obras completas.
Para cada linha de tubo PEX:
  meters = Σ_c receita[c] * contagem[c]
  divisor e margem extraidos da formula da coluna G (ROUNDUP(.../div)*marg)
  rolos_previsto = ceil(meters*marg/div)   vs   rolos_real = G(data_only)
Agrega por DN e compara previsto x real (deve dar erro ~0)."""
import sys, glob, os, re, math
from collections import defaultdict
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DN_RE = re.compile(r"PEX\s*(\d{2})", re.I)
FORM_RE = re.compile(r"/\s*(\d+)\s*\)\s*\)?\s*\*\s*([\d.]+)")  # .../DIV)*MARG
COMPLETAS = ["20241385", "20241390", "20251670"]

def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0

def linha_contagem(ws):
    best, br = -1, 3
    for r in (2, 3, 4, 5):
        n = sum(1 for c in range(8, ws.max_column+1) if isinstance(ws.cell(r, c).value, (int, float)))
        if n > best: best, br = n, r
    return br

def main():
    for obra in COMPLETAS:
        x = glob.glob(os.path.join(BASE, obra, "*.xlsx"))[0]
        wbv = openpyxl.load_workbook(x, data_only=True)
        wbf = openpyxl.load_workbook(x, data_only=False)
        print("=" * 78)
        print("OBRA", obra, "|", os.path.basename(x))
        lev = defaultdict(float); prev = defaultdict(int); real = defaultdict(int)
        margens = set(); divisores = defaultdict(set)
        for ws in wbv.worksheets:
            t = ws.title.upper()
            if not (t.startswith("RAMAL") or t.startswith("KIT") or t.startswith("CHICOTE")):
                continue
            wsf = wbf[ws.title]
            cr = linha_contagem(ws)
            counts = {c: num(ws.cell(cr, c).value) for c in range(8, ws.max_column+1)}
            for r in range(cr+1, ws.max_row+1):
                e = ws.cell(r, 5).value
                if not isinstance(e, str) or "TUBO PEX" not in e.upper(): continue
                if "ROLO" in e.upper() or "BARRA" in e.upper(): continue
                m = DN_RE.search(e)
                if not m: continue
                dn = int(m.group(1))
                meters = sum(num(ws.cell(r, c).value) * counts.get(c, 0) for c in range(8, ws.max_column+1))
                gf = wsf.cell(r, 7).value
                div, marg = 200, 1.07
                if isinstance(gf, str):
                    fm = FORM_RE.search(gf.replace(" ", ""))
                    if fm: div, marg = int(fm.group(1)), float(fm.group(2))
                margens.add(marg); divisores[dn].add(div)
                gpred = math.ceil(meters * marg / div) if meters > 0 else 0
                greal = int(num(ws.cell(r, 7).value))
                lev[dn] += meters; prev[dn] += gpred; real[dn] += greal
        print(f"  margens vistas: {sorted(margens)} | divisores por DN: {{ {', '.join(f'{k}:{sorted(v)}' for k,v in sorted(divisores.items()))} }}")
        print(f"  {'DN':>4} {'levantado(m)':>13} {'rolos_prev':>11} {'rolos_real':>11} {'erro':>6}")
        for dn in sorted(set(lev) | set(real)):
            err = prev[dn] - real[dn]
            print(f"  {dn:>4} {lev[dn]:13.1f} {prev[dn]:11d} {real[dn]:11d} {err:6d}")
        wbv.close(); wbf.close()

if __name__ == "__main__":
    main()
