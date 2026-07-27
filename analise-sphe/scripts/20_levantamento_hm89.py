# -*- coding: utf-8 -*-
"""LEVANTAMENTO HM89 (obra nova SPHE, 2 torres, SEM gabarito xlsx).

Entrega SO o que e defensavel:
- Contagens travadas no desenho: 216 aptos (12 finais x 9 pav x 2 torres).
- KITS (conexoes + tubo de kit) emprestados da biblioteca SPHE por unidade x contagem.
  (regra-mae validada: kit transfere entre obras ~±5-14%).
- Tubo de RAMAL: so o metro BRUTO medido neste DXF (informativo), sem split por DN.
- PENDENCIAS: o que falta e por que (ramal por DN, prumada, lavabo, tanque).

Uso: python 20_levantamento_hm89.py
"""
import importlib.util, os, sys, re, math
from collections import defaultdict
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
DIR = os.path.dirname(os.path.abspath(__file__))
spec = importlib.util.spec_from_file_location("lev15", os.path.join(DIR, "15_levantamento_completo.py"))
lev15 = importlib.util.module_from_spec(spec); spec.loader.exec_module(lev15)
ROLO_RE = lev15.ROLO_RE

DEST = r"C:\Users\resen\Clientes\Tecnomontagem\automatização kits\teste 08.07\LEVANTAMENTO_HM89.xlsx"

APTOS = 216  # 12 finais x 9 pav x 2 torres (travado no desenho, Jose confirmou)

# Composicao por apartamento lida no DETIPO (202-TIPO): 1 banho + 1 coz/A.S. + tanque.
# cnt = kits no predio inteiro. Confianca da CONTAGEM (nao da receita).
KITS_HM89 = [
    # (kit na biblioteca, contagem, base da contagem, confianca_contagem)
    ("CHICOTE BANHO TIPO", APTOS, "1 banho/apto (DETIPO: BANHO 24x = 12 finais)", "ALTA"),
    ("CHUVEIRO TIPO",      APTOS, "1 chuveiro/banho (Living: chuveiro=banho 1:1)", "ALTA"),
    ("CHICOTE COZINHA",    APTOS, "1 cozinha/apto (DETIPO: COZ./A.S. 24x)",        "ALTA"),
    ("TRAVESSA TANQUE",    APTOS, "1 tanque/apto (DETIPO: SOB O TANQUE 12x)",      "MEDIA"),
]

# Tubo de ramal medido no proprio DXF (planta-tipo, por pavimento) - script 03 --hid
# AF domina; AQ minima (obra a gas). Metro BRUTO, sem split por DN.
RAMAL_MEDIDO = [
    ("Agua fria (HAF-TUB, todas as sub-camadas)", 700.0, "por pavimento/torre"),
    ("  dos quais EXO-TET (ramal aereo no teto)", 526.0, "por pavimento/torre"),
    ("Agua quente (HAQ-TUB)", 3.5, "por pavimento/torre - obra a gas, AQ minima"),
]

PENDENCIAS = [
    ("Ramal aereo por DN (16/20/25/32)",
     "O projetista SPHE nao marca o DN no traco do ramal (classificacao %%C cobriu so 21%). "
     "O DN e decisao hidraulica tacita do Hederson. Leave-one-out provou que a receita de ramal "
     "NAO transfere entre obras (-100% a +140%). Sem gabarito desta obra, o metro por DN seria chute."),
    ("Prumada vertical (montantes)",
     "Comprimento = altura do shaft x nro de prumadas. Blocos PRUM-* existem no desenho, mas a contagem "
     "exige recursao em blocos aninhados e a altura pe-a-pe nao esta cotada no arquivo lido."),
    ("Travessa manifold / aquecedor",
     "Depende do nro de prumadas/manifolds por shaft (nao contado). Obra a gas: verificar se ha aquecedor."),
    ("Lavabo", "DETIPO rotula LAVATORIO so 2x - incerto quantos finais tem lavabo. Nao contabilizado."),
    ("Terreo / subsolo / barrilete", "Fora do pavimento-tipo (layout proprio). Nao incluidos nesta 1a passada."),
]


def receita_media(bib, nk):
    """media por-unidade dos itens do kit entre as obras que o tem."""
    if nk not in bib:
        return None
    agg = defaultdict(list); meta = {}
    for _, _, rec in bib[nk]:
        for ik, (d, un, q) in rec.items():
            agg[ik].append(q); meta[ik] = (d, un)
    return {ik: (meta[ik][0], meta[ik][1], sum(v) / len(v)) for ik, v in agg.items()}


def main():
    bib = lev15.extrai_biblioteca()
    itens = defaultdict(float)     # item_norm -> qtd (un exata OU metros p/ tubo)
    descr = {}                     # item_norm -> (desc, un)
    rastreio = []
    faltou = []
    for nk, cnt, base, conf in KITS_HM89:
        rec = receita_media(bib, nk)
        if rec is None:
            faltou.append((nk, "kit ausente na biblioteca"))
            rastreio.append((nk, cnt, "-", "SEM RECEITA"))
            continue
        n_fontes = len(bib[nk])
        for ik, (d, un, q) in rec.items():
            itens[ik] += q * cnt
            descr.setdefault(ik, (d, un))
        rastreio.append((nk, cnt, f"{n_fontes} obra(s) | contagem {conf}", base))

    # monta linhas: tubo (RL/BR) vira compra por rolo/barra x1,07; resto soma exata
    linhas = []
    for ik, q in sorted(itens.items(), key=lambda kv: descr[kv[0]][0]):
        d, un = descr[ik]
        m_rolo = ROLO_RE.findall(d)
        if un == "RL" and m_rolo:
            rolo = int(m_rolo[-1]); compra = math.ceil(q * 1.07 / rolo)
            linhas.append((d, "tubo (rolo)", round(q, 1), f"metros x1,07 / rolo {rolo}m", compra))
        elif un == "BR":
            compra = math.ceil(q * 1.07 / 5.8)
            linhas.append((d, "tubo (barra)", round(q, 1), "metros x1,07 / barra 5,80m", compra))
        else:
            linhas.append((d, "peca", round(q, 1), "soma exata (por unid x contagem)", math.ceil(q - 1e-9)))

    out = openpyxl.Workbook()
    ws = out.active; ws.title = "LISTA-MATERIAIS (kits)"
    ws.append(["Item", "Tipo", "Qtde calculada (m p/ tubo)", "Regra", "COMPRA sugerida"])
    for l in linhas:
        ws.append(list(l))

    ws2 = out.create_sheet("CONTAGENS")
    ws2.append(["Base de dimensionamento", "", ""])
    ws2.append(["Total de apartamentos", APTOS, "12 finais x 9 pavimentos x 2 torres (travado no desenho)"])
    ws2.append(["Torre A", 108, "12 finais (planta 103-TIPA) x 9 pav"])
    ws2.append(["Torre B", 108, "12 finais (planta 108-TIPB) x 9 pav"])
    ws2.append([])
    ws2.append(["Kit", "Contagem (predio)", "Base", "Confianca contagem"])
    for nk, cnt, base, conf in KITS_HM89:
        ws2.append([nk, cnt, base, conf])

    ws3 = out.create_sheet("RAMAL-MEDIDO-DXF")
    ws3.append(["Camada / medida", "Metros (por pav/torre)", "Obs"])
    for nome, m, obs in RAMAL_MEDIDO:
        ws3.append([nome, m, obs])
    ws3.append([])
    ws3.append(["ATENCAO", "", "Metro BRUTO medido no DXF, SEM split por DN. Nao usar para compra "
                              "sem a receita por DN do Hederson (ver PENDENCIAS)."])

    ws4 = out.create_sheet("PENDENCIAS")
    ws4.append(["O que ficou faltando", "Por que"])
    for item, por in PENDENCIAS:
        ws4.append([item, por])
    for nk, motivo in faltou:
        ws4.append([f"Kit {nk}", motivo])

    ws5 = out.create_sheet("RASTREIO-KITS")
    ws5.append(["kit", "contagem", "fontes da receita", "base da contagem"])
    for r in rastreio:
        ws5.append(list(r))

    out.save(DEST)

    tubos = [l for l in linhas if l[1].startswith("tubo")]
    pecas = [l for l in linhas if l[1] == "peca"]
    print("=" * 80)
    print(f"LEVANTAMENTO_HM89 gerado (obra sem gabarito - so parte transferivel)")
    print(f"  aptos: {APTOS} | kits contabilizados: {len(KITS_HM89)}")
    print(f"  itens: {len(linhas)} ({len(tubos)} linhas de tubo, {len(pecas)} pecas/conexoes)")
    print("  --- TUBO DE KIT (nao inclui ramal aereo) ---")
    for d, _, q, _, cp in tubos:
        print(f"    {d[:52]:<52} {q:>8.0f} m -> {cp:>4} un/rolo")
    print(f"  --- {len(pecas)} conexoes (top 8) ---")
    for d, _, q, _, cp in sorted(pecas, key=lambda x: -x[4])[:8]:
        print(f"    {d[:52]:<52} {cp:>5} un")
    print(f"  PENDENCIAS registradas: {len(PENDENCIAS)+len(faltou)}")
    print(f"  salvo: {DEST}")


if __name__ == "__main__":
    main()
