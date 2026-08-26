# -*- coding: utf-8 -*-
"""TESTE DO ITEM 5.2 — o arredondamento por rolo explica a divergencia?

Pergunta feita ao Hederson: o desenho da Living da DN16 18% / DN20 41% / DN25 13% /
DN32 28%, mas a compra real foi 12/46/20/22. O que entra na compra que nao esta
desenhado no pavimento tipo?

Resposta (10/08): "Por conta do tamanho do rolo que o fornecedor ja tem definido,
ele tem rolos pre-definidos de 50 metros, 100 metros e 200 metros."

Ele pode ter razao, mas o arredondamento por rolo JA esta no nosso modelo desde
19/06 e explica poucos pontos. Antes de reperguntar, este script mede o efeito.

O ponto cego que ele expos: ate agora pontuamos o extrator contra a COMPRA
(rolos x tamanho / 1,07). Mas a compra ja carrega o ROUNDUP embutido, e o roundup
NAO e neutro entre DNs — PEX16 vem em rolo de 200 m e PEX32 em rolo de 50 m, entao
cada DN e quantizado num passo diferente. O alvo limpo e a RECEITA
(soma receita x contagem), que e o metro liquido antes de qualquer margem.

O script compara os tres mundos:
  A. RECEITA  = Σ(receita_kit x contagem_kit)         <- necessidade real, alvo limpo
  B. COMPRA   = G rolos x tamanho / 1,07              <- alvo usado ate agora
  C. DESENHO  = mix medido pelo extrator (script 30)

Uso: python 32_teste_rolo.py [obra ...]
"""
import sys, os, glob, math, re
from collections import defaultdict
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FIRST = 8
RX_TUBO = re.compile(r"TUBO\s+(?:PEX|PERT)\D*(16|20|25|32)")
TAM_ROLO = re.compile(r"(\d+)\s*M\b")
DNS = (16, 20, 25, 32)

# Mix medido pelo extrator em 10/08 (script 30, melhor combinacao por obra).
# Serve so de referencia; passe --mix 18.6,43.1,15.9,22.5 para testar outro.
MIX_DESENHO = {
    "20241385": {16: 18.6, 20: 43.1, 25: 15.9, 32: 22.5},
    "20241390": {16: 8.1, 20: 66.9, 25: 1.8, 32: 23.2},
}


def linha_contagem(ws):
    best, br = -1, 3
    for r in (2, 3, 4, 5):
        n = sum(1 for c in range(FIRST, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, (int, float)))
        if n > best:
            best, br = n, r
    return br


def planilha(obra):
    arqs = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
            if "PRE-" not in a.upper() and "LEVANTAMENTO" not in a.upper()
            and not os.path.basename(a).startswith("~$")]
    return arqs[0] if arqs else None


def coletar(obra):
    """Uma entrada por LINHA de tubo.

    Cuidado que ja custou um resultado errado: o mesmo DN aparece em mais de uma
    linha com tamanhos de rolo diferentes (na Living o PEX16 e rolo de 200 m).
    Agregar por DN e guardar um unico tamanho quebra a cadeia — o ROUNDUP e por
    linha, porque cada linha e um produto distinto no catalogo do fornecedor.
    """
    path = planilha(obra)
    if not path:
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    liquido = defaultdict(float)
    comprado = defaultdict(float)
    linhas = []
    for ws in wb.worksheets:
        if not ws.title.upper().startswith("RAMAL"):
            continue
        cr = linha_contagem(ws)
        contagem = {c: ws.cell(cr, c).value for c in range(FIRST, ws.max_column + 1)
                    if isinstance(ws.cell(cr, c).value, (int, float))}
        for r in range(cr + 1, ws.max_row + 1):
            desc, un, g = ws.cell(r, 5).value, ws.cell(r, 6).value, ws.cell(r, 7).value
            if not isinstance(desc, str):
                continue
            m = RX_TUBO.search(desc.upper())
            if not m or str(un).strip().upper() != "RL":
                continue
            dn = int(m.group(1))
            tams = [int(t) for t in TAM_ROLO.findall(desc.upper()) if int(t) in (50, 100, 200)]
            tam = tams[-1] if tams else 100
            soma = 0.0
            for c, cont in contagem.items():
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and v:
                    soma += float(v) * float(cont)
            rol = float(g) if isinstance(g, (int, float)) else 0.0
            if not soma and not rol:
                continue
            liquido[dn] += soma
            comprado[dn] += rol * tam
            linhas.append({"aba": ws.title, "lin": r, "desc": desc[:46], "dn": dn,
                           "tam": tam, "liq": soma, "rolos": rol})
    wb.close()
    return {"path": path, "liquido": liquido, "comprado": comprado, "linhas": linhas}


def mix(d):
    s = sum(d.values())
    return {k: 100 * v / s for k, v in d.items()} if s else {}


def erro(a, b):
    return sum(abs(a.get(dn, 0) - b.get(dn, 0)) for dn in DNS)


def linha_mix(rot, m):
    return f"  {rot:<34} " + "  ".join(f"DN{dn}={m.get(dn, 0):5.1f}%" for dn in DNS)


def analisar(obra, mix_des):
    d = coletar(obra)
    if not d or not d["liquido"]:
        print(f"OBRA {obra}: sem linhas de tubo na aba RAMAL\n")
        return
    print("=" * 86)
    print(f"OBRA {obra} | {os.path.basename(d['path'])}")

    print("\n--- cadeia POR LINHA (cada linha = um produto do catalogo) ---")
    print(f"  {'DN':>4} {'rolo':>6} {'liquido m':>11} {'rolos calc':>11} "
          f"{'rolos reais':>12} {'compra m':>10} {'inflacao':>9}")
    ok = True
    for L in sorted(d["linhas"], key=lambda x: (x["dn"], x["tam"])):
        calc = math.ceil(L["liq"] * 1.07 / L["tam"]) if L["liq"] else 0
        compra = L["rolos"] * L["tam"]
        infl = 100 * (compra / L["liq"] - 1) if L["liq"] else 0
        marca = "" if abs(calc - L["rolos"]) < 0.001 else "  <-- NAO bate"
        if marca:
            ok = False
        print(f"  {L['dn']:>4} {L['tam']:>6} {L['liq']:>11.1f} {calc:>11.0f} "
              f"{L['rolos']:>12.0f} {compra:>10.0f} {infl:>8.1f}%{marca}")
    print(f"  cadeia ROUNDUP(liquido x 1,07 / rolo) reproduz a compra: "
          f"{'SIM, erro zero' if ok else 'NAO'}")

    m_liq = mix(d["liquido"])
    m_com = mix(d["comprado"])

    print("\n--- os tres mixes ---")
    print(linha_mix("A. RECEITA (metro liquido)", m_liq))
    print(linha_mix("B. COMPRA (rolos x tamanho)", m_com))
    if mix_des:
        print(linha_mix("C. DESENHO (extrator, script 30)", mix_des))

    print("\n--- quanto o rolo distorce o alvo ---")
    dist = erro(m_com, m_liq)
    print(f"  distorcao do arredondamento (B contra A): {dist:.1f} p.p.")
    for dn in DNS:
        if m_liq.get(dn):
            print(f"    DN{dn}: receita {m_liq.get(dn,0):5.1f}%  ->  "
                  f"compra {m_com.get(dn,0):5.1f}%   ({m_com.get(dn,0)-m_liq.get(dn,0):+.1f} p.p.)")

    if not mix_des:
        print()
        return

    print("\n--- o extrator contra cada alvo ---")
    e_com = erro(mix_des, m_com)
    e_liq = erro(mix_des, m_liq)
    print(f"  contra a COMPRA  (alvo usado ate agora): {e_com:5.1f} p.p.")
    print(f"  contra a RECEITA (alvo limpo)          : {e_liq:5.1f} p.p.")
    veredito = ("o rolo ATRAPALHAVA a leitura" if e_liq < e_com - 0.5
                else "o rolo NAO explica o residuo" if e_liq > e_com + 0.5
                else "o rolo e indiferente aqui")
    print(f"  -> {veredito}")

    print("\n--- teste direto: o desenho, passado pela cadeia de compra ---")
    # Reescala cada LINHA pela razao (share do desenho / share da receita) do seu DN
    # e roda o ROUNDUP linha a linha, que e como a compra de fato acontece.
    prev = defaultdict(float)
    for L in d["linhas"]:
        dn = L["dn"]
        base = m_liq.get(dn, 0)
        if not base or not L["liq"]:
            continue
        fator = mix_des.get(dn, 0) / base
        liq = L["liq"] * fator
        prev[dn] += math.ceil(liq * 1.07 / L["tam"]) * L["tam"]
    m_prev = mix(prev)
    print(linha_mix("desenho -> rolos -> metros", m_prev))
    print(linha_mix("compra real", m_com))
    print(f"  erro depois da cadeia: {erro(m_prev, m_com):.1f} p.p.   "
          f"(antes da cadeia: {e_com:.1f} p.p.)")
    print()


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    obras = args or ["20241385", "20241390"]
    custom = None
    for a in sys.argv[1:]:
        if a.startswith("--mix"):
            vals = [float(x) for x in a.split("=", 1)[1].split(",")]
            custom = dict(zip(DNS, vals))
    for obra in obras:
        analisar(obra, custom or MIX_DESENHO.get(obra))


if __name__ == "__main__":
    main()
