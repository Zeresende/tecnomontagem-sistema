# -*- coding: utf-8 -*-
import openpyxl, glob, os, sys
sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
L = openpyxl.utils.get_column_letter
for obra in ("20251430", "20251533"):
    x = glob.glob(os.path.join(BASE, obra, "*.xlsx"))[0]
    wb = openpyxl.load_workbook(x, data_only=False)
    print("====", obra, os.path.basename(x))
    for ws in wb.worksheets:
        t = ws.title.upper()
        if not (t.startswith("RAMAL") or t.startswith("KIT") or t.startswith("CHICOTE")):
            continue
        for r in range(2, ws.max_row+1):
            e = ws.cell(r, 5).value
            if isinstance(e, str) and "TUBO PEX" in e.upper() and "ROLO" not in e.upper() and "BARRA" not in e.upper():
                cells = [(L(c), ws.cell(r, c).value) for c in range(7, 14)]
                print(f"  [{ws.title}] L{r}:", cells)
                break
    wb.close()
