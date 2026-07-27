# -*- coding: utf-8 -*-
"""TRILHA DXF DO RAMAL: mede tubo AF/AQ no DXF do pavimento TIPO e compara com a
receita real de ramal da planilha da mesma obra (m/apto). Se a razao for estavel,
o ramal de obra nova pode nascer do desenho (pre-medido p/ o Hederson conferir).
Camadas: H(AF|AQ)-TUB* sem LEG/REF/TAB/DET (legendas/tabelas duplicam geometria).
Uso: python 18_dxf_vs_ramal.py [obra ...]  (default: 20241385 20241390 20251670 20251533)
"""
import sys, os, glob, re, math, logging
from collections import defaultdict
from pathlib import Path
import openpyxl
import ezdxf

sys.stdout.reconfigure(encoding="utf-8")
logging.getLogger("ezdxf").setLevel(logging.ERROR)
BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
p17 = __import__("17_parser_contagens")

TUBO_LAYER = re.compile(r"H(AF|AQ)-TUB", re.I)
EXCL_LAYER = re.compile(r"LEG|REF|TAB|DET|PREVIS", re.I)
MAX_DEPTH = 6


def comp(e):
    t = e.dxftype()
    try:
        if t == "LINE":
            return e.dxf.start.distance(e.dxf.end)
        if t == "LWPOLYLINE":
            pts = list(e.get_points("xy"))
            d = sum(math.dist(pts[i], pts[i + 1]) for i in range(len(pts) - 1))
            if e.closed and len(pts) > 2:
                d += math.dist(pts[-1], pts[0])
            return d
        if t == "POLYLINE":
            pts = [v.dxf.location for v in e.vertices]
            return sum(pts[i].distance(pts[i + 1]) for i in range(len(pts) - 1))
        if t == "ARC":
            return e.dxf.radius * math.radians((e.dxf.end_angle - e.dxf.start_angle) % 360)
    except Exception:
        pass
    return 0.0


def coletar(entidades, acc, todos, depth=0):
    for e in entidades:
        t = e.dxftype()
        if t == "INSERT" and depth < MAX_DEPTH:
            try:
                coletar(e.virtual_entities(), acc, todos, depth + 1)
            except Exception:
                pass
        elif t in ("LINE", "LWPOLYLINE", "POLYLINE", "ARC"):
            d = comp(e)
            if d > 0:
                todos.append(d)
                ly = e.dxf.layer or ""
                m = TUBO_LAYER.search(ly)
                if m and not EXCL_LAYER.search(ly):
                    suf = re.sub(r"[_\-]+", "-", ly[m.end():]).strip("-").upper() or "(base)"
                    acc[(m.group(1).upper(), suf)] += d


def medir_dxf(obra):
    d = BASE / "_analise" / "dxf" / obra
    tipos = [p for p in sorted(d.glob("*.dxf"))
             if re.search(r"PVTIPO|TIPO|TIP\b|-TIP-", p.name, re.I)
             and not re.search(r"DET", p.name, re.I)]
    if not tipos:
        return None, None
    path = tipos[0]
    doc = ezdxf.readfile(str(path))
    acc, todos = defaultdict(float), []
    coletar(doc.modelspace(), acc, todos)
    todos.sort()
    med = todos[len(todos) // 2] if todos else 1.0
    fator = 1.0 if med < 20 else (100.0 if med < 2000 else 1000.0)
    return {k: v / fator for k, v in acc.items()}, path.name


def planilha_ramal(obra):
    """m/apto de tubo AF e AQ pela receita real do ramal (media ponderada por contagem)."""
    x = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
         if "PRE-" not in a.upper() and "LEVANTAMENTO" not in a.upper()
         and not os.path.basename(a).startswith("~$")][0]
    wb = openpyxl.load_workbook(x, data_only=True)
    soma = {"AF": [0.0, 0.0], "AQ": [0.0, 0.0]}  # tipo -> [m*cont, cont]
    for ws in wb.worksheets:
        if not ws.title.upper().startswith("RAMAL"):
            continue
        cr = p17.linha_contagem(ws)
        for c in range(8, ws.max_column + 1):
            cnt = ws.cell(cr, c).value
            if not isinstance(cnt, (int, float)) or cnt == 0:
                continue
            h = " ".join(str(ws.cell(r, c).value or "") for r in range(1, cr)).upper()
            tipo = "AQ" if ("AQ" in h or "QUENTE" in h) else ("AF" if ("AF" in h or "FRIA" in h) else None)
            if tipo is None:
                continue
            m_kit = 0.0
            for r in range(cr + 1, ws.max_row + 1):
                e, v = ws.cell(r, 5).value, ws.cell(r, c).value
                if (isinstance(e, str) and e.upper().startswith("TUBO")
                        and str(ws.cell(r, 6).value or "").strip().upper() == "RL"
                        and isinstance(v, (int, float)) and v):
                    m_kit += float(v)
            if m_kit:
                soma[tipo][0] += m_kit * cnt
                soma[tipo][1] += cnt
    wb.close()
    out = {}
    for t, (mc, cc) in soma.items():
        out[t] = (mc / cc if cc else 0.0, cc)
    return out


def aptos_pav(obra):
    x = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
         if "PRE-" not in a.upper() and "LEVANTAMENTO" not in a.upper()
         and not os.path.basename(a).startswith("~$")][0]
    wb = openpyxl.load_workbook(x, data_only=True)
    pred = p17.predios_da_obra(wb)
    wb.close()
    fins = sum(len(v) for v in pred.values() if v)
    return fins


def main():
    obras = sys.argv[1:] or ["20241385", "20241390", "20251670", "20251533"]
    for obra in obras:
        dxf, nome = medir_dxf(obra)
        if dxf is None:
            print(f"{obra}: SEM DXF TIPO")
            continue
        plan = planilha_ramal(obra)
        n = aptos_pav(obra)
        print("=" * 84)
        print(f"OBRA {obra} | {nome} | aptos/pav(matriz)={n}")
        print(f"  planilha (receita real do ramal): AF={plan['AF'][0]:.2f} m/apto  AQ={plan['AQ'][0]:.2f} m/apto")
        for tipo in ("AF", "AQ"):
            p = plan[tipo][0]
            print(f"  --- H{tipo}-TUB por sufixo de camada ---")
            tot = 0.0
            for (t, suf), m in sorted(dxf.items(), key=lambda kv: -kv[1]):
                if t != tipo:
                    continue
                tot += m
                map_ = m / max(n, 1)
                razao = f"{map_/p:5.2f}x" if p else "  -  "
                print(f"     {suf:24} {m:8.1f} m  {map_:7.2f} m/apto  vs plan {razao}")
            if p:
                print(f"     {'TOTAL':24} {tot:8.1f} m  {tot/max(n,1):7.2f} m/apto  vs plan {tot/max(n,1)/p:5.2f}x")


if __name__ == "__main__":
    main()
