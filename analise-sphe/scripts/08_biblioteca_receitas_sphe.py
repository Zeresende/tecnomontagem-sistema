# -*- coding: utf-8 -*-
"""Extrai receita[kit][DN] (metros de PEX por unidade de kit) de cada obra SPHE
e compara entre obras p/ verificar estabilidade da biblioteca do projetista."""
import sys, glob, os, re
from collections import defaultdict
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DN_RE = re.compile(r"PEX\s*(\d{2})", re.I)
FIRST = 8

def norm_kit(s):
    if not isinstance(s, str): return None
    s = re.sub(r"\s+", " ", s.replace("\n", " ")).strip().upper()
    s = re.sub(r"\s*(FINA?IS?|FINAL)\s*[\d/ E]+", "", s)  # tira sufixo de finais
    s = re.sub(r"\s+\d+/\d+.*$", "", s)
    return s.strip(" -")

def linha_contagem(ws):
    best, br = -1, 3
    for r in (2,3,4,5):
        n = sum(1 for c in range(FIRST, ws.max_column+1) if isinstance(ws.cell(r,c).value,(int,float)))
        if n > best: best, br = n, r
    return br

def receitas_obra(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    rec = defaultdict(lambda: defaultdict(float))  # kit -> dn -> metros/unid
    for ws in wb.worksheets:
        t = ws.title.upper()
        if not (t.startswith("RAMAL") or t.startswith("KIT") or t.startswith("CHICOTE")): continue
        cr = linha_contagem(ws)
        nomes = {c: norm_kit(ws.cell(cr-1,c).value) for c in range(FIRST, ws.max_column+1)}
        for r in range(cr+1, ws.max_row+1):
            e = ws.cell(r,5).value
            if not isinstance(e,str) or "TUBO PEX" not in e.upper(): continue
            if "ROLO" in e.upper() or "BARRA" in e.upper(): continue
            m = DN_RE.search(e)
            if not m: continue
            dn = int(m.group(1))
            for c in range(FIRST, ws.max_column+1):
                v = ws.cell(r,c).value
                if isinstance(v,(int,float)) and v and nomes.get(c):
                    rec[nomes[c]][dn] += float(v)
    wb.close()
    return rec

def main():
    obras = {o: glob.glob(os.path.join(BASE,o,"*.xlsx"))[0]
             for o in sorted(os.listdir(BASE))
             if os.path.isdir(os.path.join(BASE,o)) and not o.startswith("_")
             and glob.glob(os.path.join(BASE,o,"*.xlsx"))}
    todas = {o: receitas_obra(p) for o,p in obras.items()}
    # uniao de kits
    kits = sorted({k for r in todas.values() for k in r})
    for kit in kits:
        linhas = []
        for o, r in todas.items():
            if kit in r:
                dns = ", ".join(f"DN{dn}={r[kit][dn]:g}" for dn in sorted(r[kit]))
                linhas.append(f"{o}: {dns}")
        if linhas:
            print(f"\n### KIT: {kit}")
            for l in linhas: print("   ", l)

if __name__ == "__main__":
    main()
