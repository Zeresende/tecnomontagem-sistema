# -*- coding: utf-8 -*-
"""PARSER UNIVERSAL DE CONTAGENS SPHE - deriva contagem de cada coluna de kit a
partir das abas PRÉDIO/RESUMO e valida contra a contagem real preenchida.
Mecanismos (tentados em ordem):
  M1 matriz PRÉDIO: conta marcas (X ou codigo de tipologia) por final, em toda
     aba com PRÉDIO no nome (multi-torre: soma abas da torre correspondente).
  M2 bloco INDUSTRIALIZAÇÃO do RESUMO: contagens prontas por nome de kit.
  M3 totais de ambiente do RESUMO (chuveiro=banho tipo etc.).
Living: térreo/20º/duplex sem coluna própria -> validamos tambem o INVARIANTE
de grupo (soma das colunas que cobrem os mesmos finais == total da matriz)."""
import sys, os, glob, re, unicodedata
from collections import defaultdict
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OBRAS = ["20241385", "20241390", "20251670", "20251430", "20251533"]


def sem_acento(s):
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()


def abas_kit(wb):
    for ws in wb.worksheets:
        t = ws.title.upper()
        if t.startswith("RAMAL") or t.startswith("KIT") or t.startswith("CHICOTE"):
            yield ws


def linha_contagem(ws):
    best, br = -1, 3
    for r in (2, 3, 4, 5):
        n = sum(1 for c in range(8, ws.max_column + 1) if isinstance(ws.cell(r, c).value, (int, float)))
        if n > best:
            best, br = n, r
    return br


def torre_de(nome):
    m = re.search(r"TORRE\s*([A-Z])", nome.upper())
    return m.group(1) if m else None


def parse_predio(ws):
    """Le UMA aba PRÉDIO: {final_label: n_pavimentos}. Marca = X ou codigo TP*."""
    # 1) linha do cabecalho FINAIS
    hdr = None
    for r in range(1, min(ws.max_row, 12) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and sem_acento(v).strip().upper() == "FINAIS":
                hdr = (r, c)
                break
        if hdr:
            break
    if not hdr:
        return {}
    rh, ch = hdr
    # 2) labels de finais: proxima linha nao-vazia a partir de ch, para no 1o buraco grande
    lab_row = None
    for r in range(rh, rh + 3):
        labs = {c: str(ws.cell(r, c).value).strip() for c in range(ch, ws.max_column + 1)
                if ws.cell(r, c).value not in (None, "") and r != rh}
        if len(labs) >= 2:
            lab_row = r
            break
    if lab_row is None:
        return {}
    labs = {}
    vazio = 0
    for c in range(ch, ws.max_column + 1):
        v = ws.cell(lab_row, c).value
        if v in (None, ""):
            vazio += 1
            if vazio >= 3 and labs:
                break
            continue
        vazio = 0
        s = str(v).strip()
        if re.fullmatch(r"[A-Z]?\d+", s):
            labs[c] = s
    if not labs:
        return {}
    # 3) linhas de pavimento = abaixo do label, com marca em alguma coluna de final
    def eh_marca(v):
        if v in (None, ""):
            return False
        s = str(v).strip().upper()
        return s == "X" or bool(re.fullmatch(r"TP\s*\d+[A-Z]?", s))
    contagem = defaultdict(int)
    branco = 0
    for r in range(lab_row + 1, ws.max_row + 1):
        marcas = [c for c in labs if eh_marca(ws.cell(r, c).value)]
        # celulas mescladas deslocam: aceita marca 1-2 col a direita do label
        if not marcas:
            todas = [c for c in range(ch, max(labs) + 3) if eh_marca(ws.cell(r, c).value)]
            if todas:
                orden = sorted(labs)
                marcas = orden[:len(todas)] if len(todas) <= len(orden) else orden
        if marcas:
            branco = 0
            for c in marcas:
                contagem[labs[c]] += 1
        else:
            branco += 1
            if branco >= 4 and contagem:
                break
    return dict(contagem)


def predios_da_obra(wb):
    """{torre|None: {final: n_pav}} somando abas PRÉDIO por torre."""
    out = defaultdict(lambda: defaultdict(int))
    for ws in wb.worksheets:
        t = sem_acento(ws.title).upper()
        if "PREDIO" not in t:
            continue
        m = parse_predio(ws)
        if not m:
            continue
        tor = torre_de(ws.title)
        for f, n in m.items():
            out[tor][f] += n
    return {k: dict(v) for k, v in out.items()}


def parse_industrializacao(wb):
    """{nome_normalizado: qtd} do bloco INDUSTRIALIZAÇÃO do RESUMO."""
    out = {}
    for ws in wb.worksheets:
        if not sem_acento(ws.title).upper().startswith("RESUMO"):
            continue
        ini = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and "INDUSTRIALIZA" in sem_acento(v).upper():
                ini = r
                break
        if ini is None:
            continue
        for r in range(ini + 1, ws.max_row + 1):
            nome, qtd = ws.cell(r, 1).value, ws.cell(r, 2).value
            if not isinstance(nome, str) or not nome.strip():
                if r > ini + 2:
                    break
                continue
            if isinstance(qtd, (int, float)) and qtd:
                out[norm_nome(nome)] = float(qtd)
    return out


def norm_nome(s):
    s = sem_acento(s).upper()
    s = re.sub(r"\bKIT\b|\bTIPO\b", " ", s)
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def finais_do_header(h):
    """['1','2'] ou ['A1','A2'] a partir do header da coluna de kit."""
    up = sem_acento(h).upper()
    m = re.search(r"FINA(?:L|IS)?\b(.*)$", up)
    if not m:
        return None
    # pega a sequencia de finais no comeco do resto: numeros/letras+numero ligados por / , E espacos
    cauda = m.group(1)
    m2 = re.match(r"[\s:\-]*((?:[A-Z]?\d+)(?:\s*(?:[/,]|E\b)\s*[A-Z]?\d+)*)", cauda)
    if not m2:
        return None
    toks = re.findall(r"[A-Z]?\d+", m2.group(1))
    return [t.lstrip("0") or "0" for t in toks] or None


def analisa(obra):
    x = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
         if "PRE-" not in a.upper() and "LEVANTAMENTO" not in a.upper()
         and not os.path.basename(a).startswith("~$")][0]
    wb = openpyxl.load_workbook(x, data_only=True)
    predios = predios_da_obra(wb)
    indus = parse_industrializacao(wb)
    tot_matriz = {t: sum(v.values()) for t, v in predios.items()}
    print("=" * 88)
    print(f"OBRA {obra} | matriz PRÉDIO: " + (" ; ".join(
        f"torre {t or '-'}: {len(v)} finais, {tot_matriz[t]:.0f} aptos" for t, v in predios.items()) or "NAO ACHADA")
        + f" | INDUSTRIALIZACAO: {len(indus)} kits")

    ok = tot = 0
    grupo = defaultdict(lambda: [0.0, 0.0])  # (torre, finais congelados) -> [soma_cont, total_matriz]
    falhas = []
    for ws in abas_kit(wb):
        cr = linha_contagem(ws)
        tor_aba = torre_de(ws.title)
        for c in range(8, ws.max_column + 1):
            cnt = ws.cell(cr, c).value
            if not isinstance(cnt, (int, float)) or cnt == 0:
                continue
            h = " / ".join(str(ws.cell(r, c).value or "").replace("\n", " ").strip()
                           for r in range(1, cr) if ws.cell(r, c).value)
            if not h:
                continue
            tot += 1
            fins = finais_do_header(h)
            deriv, via = None, None
            mapa = predios.get(tor_aba) or (predios.get(None) if len(predios) == 1 or tor_aba is None else None)
            if fins and mapa:
                vals = [mapa.get(f) for f in fins]
                if all(v is not None for v in vals):
                    deriv, via = sum(vals), "M1-predio"
                    up0 = sem_acento(h).upper()
                    ag = "AQ" if ("AQ" in up0 or "QUENTE" in up0) else ("AF" if ("AF" in up0 or "FRIA" in up0) else "-")
                    k = (tor_aba, ag, tuple(sorted(fins)))
                    grupo[k][0] += float(cnt)
                    grupo[k][1] = deriv
            up = sem_acento(h).upper()
            if deriv is None and "TRAVESSA" in up and "MANIFOLD" not in up:
                deriv, via = sum(tot_matriz.values()), "M3-aptos"
            if deriv is None and indus:
                nn = norm_nome(h)
                if nn in indus:
                    deriv, via = indus[nn], "M2-indus"
                else:
                    cands = [q for n, q in indus.items()
                             if len(nn) >= 10 and (nn in n or n in nn)]
                    if len(set(cands)) == 1:
                        deriv, via = cands[0], "M2-indus~"
            if deriv is not None and abs(deriv - cnt) < 0.5:
                ok += 1
            else:
                falhas.append((ws.title, h[:44], cnt, via, deriv))
    # invariante de grupo (colunas que cobrem os mesmos finais somadas == matriz)
    g_ok = sum(1 for s, m in grupo.values() if abs(s - m) < 0.5)
    print(f"  colunas derivadas EXATAS: {ok}/{tot} | grupos de finais com INVARIANTE ok: {g_ok}/{len(grupo)}")
    resg = 0
    for (t, ag, fins), (s, m) in sorted(grupo.items()):
        if abs(s - m) < 0.5:
            resg += sum(1 for f in falhas if f[3] == "M1-predio" and finais_do_header(f[1])
                        and tuple(sorted(finais_do_header(f[1]))) == fins)
    for f in falhas[:14]:
        print(f"    falha [{f[0][:14]}] {f[1]:<44} cont={f[2]:.0f} via={f[3]} deriv={f[4]}")
    if len(falhas) > 14:
        print(f"    ... +{len(falhas)-14} falhas")
    print(f"  (colunas 'falhas' cujo GRUPO fecha exato com a matriz: {resg} -> so falta a alocacao entre colunas)")
    wb.close()
    return ok, tot


def main():
    T = O = 0
    for obra in OBRAS:
        o, t = analisa(obra)
        O += o
        T += t
    print("=" * 88)
    print(f"GERAL: {O}/{T} colunas de kit derivadas exatas via PRÉDIO/RESUMO")


if __name__ == "__main__":
    main()
