# -*- coding: utf-8 -*-
"""Decomposicao do resultado holdout por coluna de kit (receita_lida.csv), usada pelo 63.

Entrada: csv COLUNA;PECA_ID;QTD_POR_KIT (tubo em m/kit, conexao em UN/kit) - o que o conector
leu de cada kit na prancha DTIP. Compara celula a celula com a aba POR_KIT do gabarito.

Regra cravada em 04/09/2026 (antes da rodada): diferenca que cai INTEIRA numa celula listada na
aba PENDENTES do gabarito (os tubos que o Marcelo ficou de cotar) e "explicada, aguarda Marcelo",
nao reprova. Os metros dessas celulas sao descontados do total do tubo para o C3 ajustado.
"""
import csv
import io
import math
import re
from collections import defaultdict
from openpyxl import load_workbook

RX_DN = re.compile(r"(?:PERT|PEX)[ ]*(16|20|25|32)")
TOL_TUBO_CELULA = 0.05      # m/kit: leitura de desenho tem folga de medicao
TOL_CONEXAO_CELULA = 0.0    # UN/kit: inteiro, exato


def _sep(texto):
    return ";" if texto.count(";") >= texto.count(",") else ","


def ler_receita_lida(caminho, norm):
    """-> {(coluna_norm, pid): qtd_por_kit}, avisos"""
    with io.open(caminho, encoding="utf-8-sig", newline="") as fh:
        texto = fh.read()
    out, avisos = defaultdict(float), []
    for r in csv.DictReader(io.StringIO(texto), delimiter=_sep(texto)):
        ch = {k.upper().strip(): k for k in r}
        try:
            col = norm(r[ch["COLUNA"]])
            pid = int(float(r[ch["PECA_ID"]]))
            qtd = float(str(r[ch["QTD_POR_KIT"]]).replace(",", "."))
        except (KeyError, ValueError):
            avisos.append(f"linha ignorada em receita_lida: {r}")
            continue
        out[(col, pid)] += qtd
    return dict(out), avisos


def ler_gabarito_por_kit(pasta, apelido, norm):
    """-> celulas {(coluna_norm, pid): receita}, contagem {coluna_norm: n}, pendentes {(coluna_norm, pid)}"""
    wb = load_workbook(pasta / f"gabarito_{apelido.lower()}.xlsx", read_only=True)
    celulas, contagem, tubo_por_dn = {}, {}, {}
    for row in wb["POR_KIT"].iter_rows(min_row=2, values_only=True):
        aba, coluna, cont, peca, _, receita, pid, _ = row
        if not coluna or pid in (None, ""):
            continue
        col = norm(coluna)
        celulas[(col, int(pid))] = celulas.get((col, int(pid)), 0.0) + float(receita)
        contagem[col] = float(cont)
        m = RX_DN.search(str(peca).upper())
        if str(peca).upper().startswith("TUBO") and m:
            tubo_por_dn[m.group(1)] = int(pid)
    pendentes = set()
    if "PENDENTES" in wb.sheetnames:
        for row in wb["PENDENTES"].iter_rows(min_row=2, values_only=True):
            coluna, _, peca, status, _ = row
            if not status or not str(status).startswith("pendente"):
                continue
            m = RX_DN.search(str(peca).upper())
            pid = tubo_por_dn.get(m.group(1)) if m else None
            if pid is None:
                # tubo de bitola que a obra nao usa em kit nenhum: resolve pelo catalogo no 63
                pid = ("TUBO", m.group(1) if m else "?")
            pendentes.add((norm(coluna), pid))
    return celulas, contagem, pendentes


def resolver_pendentes_por_catalogo(pendentes, catalogo, norm):
    """Troca marcadores ('TUBO', dn) por PECA_ID do tubo Serie 5 daquela bitola."""
    out = set()
    for col, pid in pendentes:
        if isinstance(pid, tuple):
            dn = pid[1]
            pid = next((p for p, x in catalogo.items() if x["sistema"] == "PEX"
                        and norm(x["descricao"]).startswith(f"TUBO PEX {dn} - S")), None)
            if pid is None:
                continue
        out.add((col, pid))
    return out


def decompor(lida, celulas, contagem, pendentes, contagem_resultado, catalogo, gab):
    """Compara celula a celula. -> linhas [(coluna, pid, desc, gab, lido, status)], explicado {pid: metros}"""
    linhas, explicado = [], defaultdict(float)
    cont = dict(contagem)
    cont.update(contagem_resultado or {})
    # so compara as colunas que o conector entregou; o ramal e medido, nao e receita, e so entra
    # se vier no arquivo. Colunas do gabarito ausentes viram uma linha de resumo cada.
    entregues = {col for col, _ in lida}
    ausentes = sorted({col for col, _ in celulas} - entregues)
    ramal = [c for c in ausentes if "RAMAL" in c]
    if ramal:
        linhas.append((f"{len(ramal)} colunas de RAMAL", 0, "(ramal e medido na TIPO, nao e receita)",
                       None, None, "AUSENTE (esperado)"))
    for col in ausentes:
        if col not in ramal:
            linhas.append((col, 0, "(coluna inteira ausente na receita_lida)", None, None,
                           "AUSENTE (coluna de kit nao entregue)"))
    chaves = sorted({k for k in set(celulas) | set(lida) if k[0] in entregues},
                    key=lambda k: (k[0], k[1]))
    for col, pid in chaves:
        g, r = celulas.get((col, pid)), lida.get((col, pid))
        desc = str(catalogo.get(pid, {}).get("descricao", f"PECA_ID {pid} fora do catalogo"))
        tubo = gab.get(pid, {}).get("tipo") == "TUBO" or desc.upper().startswith("TUBO")
        if col not in cont:
            linhas.append((col, pid, desc, g, r, "COLUNA NAO RECONHECIDA"))
            continue
        if g is None:
            status = "EXPLICADA (aguarda Marcelo)" if (col, pid) in pendentes else "A MAIS"
        elif r is None:
            status = "FALTA"
        else:
            tol = TOL_TUBO_CELULA * g if tubo else TOL_CONEXAO_CELULA
            status = "ok" if abs(r - g) <= tol + 1e-9 else "DIFERENCA"
        if status.startswith("EXPLICADA") and tubo:
            explicado[pid] += r * cont[col]
        linhas.append((col, pid, desc, g, r, status))
    return linhas, dict(explicado)


def conferir_totais(lida, contagem, contagem_resultado, res, gab, tubo_em, folga):
    """Consistencia interna do conector: soma(celula x contagem) tem de bater com o levantamento."""
    cont = dict(contagem)
    cont.update(contagem_resultado or {})
    soma = defaultdict(float)
    for (col, pid), q in lida.items():
        if col in cont:
            soma[pid] += q * cont[col]
    avisos = []
    for pid, total in sorted(soma.items()):
        r = res.get(pid)
        if r is None:
            avisos.append(f"PECA_ID {pid}: esta na receita_lida ({total:.0f}) e nao no levantamento")
            continue
        if tubo_em == "rolos" and gab.get(pid, {}).get("rolo"):
            r = r * gab[pid]["rolo"] / folga
        if total > r * 1.01 + 0.5:      # o levantamento nunca pode ser MENOR que a parte dos kits
            avisos.append(f"PECA_ID {pid}: receita x contagem = {total:.0f} > levantamento = {r:.0f}")
    if avisos:
        print("\n  aviso de consistencia (receita_lida x levantamento):")
        for x in avisos:
            print("   ", x)


def imprimir(linhas, explicado, gab, res, tubo_em, folga, tol_rolos):
    """Imprime a decomposicao e devolve o C3 ajustado (bool)."""
    print("\nDECOMPOSICAO por coluna de kit (receita_lida.csv x aba POR_KIT do gabarito):")
    print(f"  {'coluna':22} {'PECA_ID':>7} {'descricao':40} {'gab/kit':>8} {'lido/kit':>8}  status")
    resumo = defaultdict(int)
    for col, pid, desc, g, r, st in linhas:
        resumo[st.split(" ")[0]] += 1
        if st == "ok":
            continue
        gs = f"{g:g}" if g is not None else "-"
        rs = f"{r:g}" if r is not None else "-"
        print(f"  {col[:22]:22} {pid:>7} {desc[:40]:40} {gs:>8} {rs:>8}  {st}")
    print("  resumo: " + ", ".join(f"{k} {v}" for k, v in sorted(resumo.items())))

    # consistencia interna: soma das celulas x contagem tem de bater com o levantamento
    ok_c3 = True
    if explicado:
        print(f"\nC3 ajustado (descontados os metros das celulas pendentes do Marcelo, +-{tol_rolos} rolo):")
    else:
        print(f"\nC3 por bitola a partir da receita_lida (sem celula pendente no gabarito: "
              f"ajustado = bruto, +-{tol_rolos} rolo):")
    for pid, g in sorted(gab.items()):
        if g["tipo"] != "TUBO":
            continue
        r = res.get(pid, 0.0)
        metros = r * g["rolo"] / folga if tubo_em == "rolos" else r
        ajust = max(metros - explicado.get(pid, 0.0), 0.0)
        rolos = math.ceil(ajust * folga / g["rolo"]) if ajust else 0
        passa = abs(rolos - g["rolos"]) <= tol_rolos
        ok_c3 &= passa
        exp = f" (explicado {explicado[pid]:.0f} m)" if pid in explicado else ""
        print(f"  {pid:>7} {g['desc'][:44]:44} gabarito {g['rolos']:>3} rolos | "
              f"ajustado {ajust:>7.0f} m = {rolos:>3} rolos{exp}  {'ok' if passa else 'FALHA'}")
    return ok_c3
