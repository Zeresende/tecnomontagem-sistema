# -*- coding: utf-8 -*-
"""Reconcilia as linhas de CONEXAO (nao-tubo) das obras SPHE completas:
testa se col G (Qtde Total) = f( soma(receita x contagem) ) e com qual regra
(exata / ceil / x1,05 / x1,07). Mesma logica dos scripts 10-12, estendida."""
import sys, glob, os, math, re
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FIRST = 8
OBRAS = ["20241385", "20241390", "20251670"]  # Living, Edition, Pamaris (completas)

def linha_contagem(ws):
    best, br = -1, 3
    for r in (2, 3, 4, 5):
        n = sum(1 for c in range(FIRST, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, (int, float)))
        if n > best:
            best, br = n, r
    return br

def eh_tubo(desc):
    return "TUBO PEX" in desc.upper()

def classifica(g, soma):
    """Retorna a regra que reproduz g a partir de soma, ou None."""
    if soma == 0:
        return None
    candidatos = [
        ("exata",        soma),
        ("ceil",         math.ceil(soma)),
        ("x1,05 ceil",   math.ceil(soma * 1.05)),
        ("x1,07 ceil",   math.ceil(soma * 1.07)),
        ("x1,05 round",  round(soma * 1.05)),
        ("x1,07 round",  round(soma * 1.07)),
    ]
    for nome, v in candidatos:
        if abs(v - g) < 0.001:
            return nome
    return None

def reconcilia(path, obra):
    wb = openpyxl.load_workbook(path, data_only=True)
    stats = {"linhas": 0, "match": 0}
    regras, falhas = {}, []
    for ws in wb.worksheets:
        t = ws.title.upper()
        if not (t.startswith("RAMAL") or t.startswith("KIT") or t.startswith("CHICOTE")):
            continue
        cr = linha_contagem(ws)
        contagem = {c: ws.cell(cr, c).value for c in range(FIRST, ws.max_column + 1)
                    if isinstance(ws.cell(cr, c).value, (int, float))}
        for r in range(cr + 1, ws.max_row + 1):
            desc = ws.cell(r, 5).value
            g = ws.cell(r, 7).value
            if not isinstance(desc, str) or eh_tubo(desc):
                continue
            if not isinstance(g, (int, float)) or g == 0:
                continue
            soma = 0.0
            for c, cont in contagem.items():
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and v:
                    soma += float(v) * float(cont)
            stats["linhas"] += 1
            regra = classifica(g, soma)
            if regra:
                stats["match"] += 1
                regras[regra] = regras.get(regra, 0) + 1
            else:
                falhas.append((ws.title, r, desc[:55], g, round(soma, 2)))
    wb.close()
    print(f"\n{'='*74}\nOBRA {obra} | {os.path.basename(path)}")
    print(f"  linhas de conexao com G preenchido: {stats['linhas']} | reproduzidas: {stats['match']}")
    for nome, n in sorted(regras.items(), key=lambda x: -x[1]):
        print(f"    regra {nome:<12} -> {n} linhas")
    if falhas:
        print(f"  NAO reproduzidas ({len(falhas)}):")
        for aba, r, d, g, s in falhas[:20]:
            print(f"    [{aba}] L{r} {d:<55} G={g} soma={s}")
        if len(falhas) > 20:
            print(f"    ... +{len(falhas)-20}")

def main():
    for obra in OBRAS:
        arqs = glob.glob(os.path.join(BASE, obra, "*QUANTITATIVO*.xlsx"))
        arqs = [a for a in arqs if "PRE-" not in os.path.basename(a).upper()
                and not os.path.basename(a).startswith("~$")]
        if not arqs:
            arqs = [a for a in glob.glob(os.path.join(BASE, obra, "*.xlsx"))
                    if "PRE-" not in os.path.basename(a).upper()
                    and not os.path.basename(a).startswith("~$")]
        if not arqs:
            print(f"OBRA {obra}: xlsx nao encontrado")
            continue
        reconcilia(arqs[0], obra)

if __name__ == "__main__":
    main()
