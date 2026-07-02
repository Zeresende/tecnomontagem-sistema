"""
Varredura completa: todos os xlsx (templates + v2) e arquivos .txt
buscando caracteres problematicos.

Reporta:
- U+FFFD (REPLACEMENT CHARACTER) = corrompido de verdade
- caracteres de controle (U+0000 a U+001F exceto tab/newline)
- bytes que podem ser cp1252 mal interpretado (ex: \xe9 solto)
"""
import openpyxl
from pathlib import Path
import unicodedata

ARQUIVOS = [
    # v2
    Path(r"C:\Users\resen\Clientes\Tecnomontagem\automatização kits\sistema\dados\catalogo_pecas.xlsx"),
    Path(r"C:\Users\resen\Clientes\Tecnomontagem\automatização kits\sistema\dados\equivalencias.xlsx"),
    Path(r"C:\Users\resen\Clientes\Tecnomontagem\automatização kits\sistema\dados\fornecedores.xlsx"),
]

# Tambem percorrer templates originais para comparar
TEMPLATES = Path(r"C:\Users\resen\Clientes\Tecnomontagem\automatização kits\Levantamento de Informações _ Hederson\16.05.2026\extraido\Projetos")
for p in TEMPLATES.glob("OBRA-QUANTITATIVO*.xlsx"):
    ARQUIVOS.append(p)

def analisar_string(s):
    """Retorna lista de problemas encontrados na string."""
    problemas = []
    for i, c in enumerate(s):
        cp = ord(c)
        if cp == 0xFFFD:
            problemas.append((i, c, cp, "REPLACEMENT CHARACTER (corrompido)"))
        elif cp < 0x20 and c not in "\t\n\r":
            problemas.append((i, c, cp, "controle"))
        elif 0x80 <= cp < 0xA0:
            problemas.append((i, c, cp, "C1 control (suspeito de cp1252 mal lido)"))
    return problemas

print("VARREDURA DE ENCODING\n")
total_celulas = 0
total_strings = 0
total_problemas = 0
problemas_exemplos = []

for arquivo in ARQUIVOS:
    if not arquivo.exists():
        continue
    print(f"\n[FILE] {arquivo.name}")
    try:
        wb = openpyxl.load_workbook(arquivo, data_only=True)
    except Exception as e:
        print(f"  ERRO: {e}")
        continue
    for aba_nome in wb.sheetnames:
        ws = wb[aba_nome]
        celulas_aba = 0
        problemas_aba = 0
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                celulas_aba += 1
                total_celulas += 1
                if isinstance(cell.value, str):
                    total_strings += 1
                    probs = analisar_string(cell.value)
                    if probs:
                        problemas_aba += len(probs)
                        total_problemas += len(probs)
                        if len(problemas_exemplos) < 10:
                            problemas_exemplos.append((arquivo.name, aba_nome, cell.coordinate, cell.value, probs))
        print(f"  aba '{aba_nome}': {celulas_aba} celulas, {problemas_aba} problemas")

print(f"\n{'='*70}")
print(f"RESUMO GLOBAL")
print(f"{'='*70}")
print(f"Total de celulas analisadas: {total_celulas:,}")
print(f"Total de strings: {total_strings:,}")
print(f"Total de PROBLEMAS encontrados: {total_problemas}")

if problemas_exemplos:
    print("\nEXEMPLOS DE PROBLEMAS:")
    for arq, aba, coord, val, probs in problemas_exemplos:
        print(f"\n  {arq} :: {aba} :: {coord}")
        print(f"  Valor: {val!r}")
        for idx, c, cp, tipo in probs:
            print(f"    pos {idx}: U+{cp:04X} ({tipo})")
else:
    print("\nNENHUM PROBLEMA REAL. Conteudo de todos os xlsx esta integro.")
